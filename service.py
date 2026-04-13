#!/usr/bin/env python3
"""
Waterqo Swimming Pools — Pool Service Manager
Run: python pool_manager.py
Then open: http://localhost:5050 in your browser
Requires: pip install openpyxl flask
"""

import json, os, sys
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from flask import Flask, request, jsonify, send_file
except ImportError:
    print("Installing Flask...")
    os.system(f"{sys.executable} -m pip install flask -q")
    from flask import Flask, request, jsonify, send_file

EXCEL_FILE = "pool_service.xlsx"
app = Flask(__name__)

# ─── Excel helpers ────────────────────────────────────────────────────────────

def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def h_style(cell, bg="0077B6", fg="FFFFFF"):
    cell.font = Font(bold=True, color=fg, name='Arial', size=11)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border()

def d_style(cell, even=True):
    cell.font = Font(name='Arial', size=10)
    cell.border = thin_border()
    cell.alignment = Alignment(vertical='center')
    cell.fill = PatternFill("solid", start_color="E0F4FF" if even else "FFFFFF")

def ensure_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Client Details"
        ws.merge_cells('A1:C1')
        ws['A1'] = 'Waterqo Swimming Pools — Pool Service'
        ws['A1'].font = Font(bold=True, name='Arial', size=14, color="FFFFFF")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill("solid", start_color="0077B6")
        ws.row_dimensions[1].height = 36
        for col, h in enumerate(['Location', 'Client Name', 'Service Person'], 1):
            h_style(ws.cell(row=2, column=col, value=h))
        ws.row_dimensions[2].height = 24
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 20
        wb.save(EXCEL_FILE)

DEFAULT_PERSONS = ['Indika', 'Dinesh', 'Indranath', 'Dilhan']

def get_persons():
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE, data_only=True)
    if 'Settings' not in wb.sheetnames:
        return list(DEFAULT_PERSONS)
    ws = wb['Settings']
    persons = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            persons.append(str(row[0]).strip())
    return persons if persons else list(DEFAULT_PERSONS)

def save_persons(persons):
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    if 'Settings' in wb.sheetnames:
        del wb['Settings']
    ws = wb.create_sheet(title='Settings')
    ws['A1'] = 'Service Persons'
    ws['A1'].font = Font(bold=True, name='Arial', size=11, color='0077B6')
    ws.column_dimensions['A'].width = 22
    for i, name in enumerate(persons, start=2):
        ws.cell(row=i, column=1, value=name)
    wb.save(EXCEL_FILE)

def get_clients():
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Client Details']
    clients = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1]:
            clients.append({
                'location': str(row[0]).strip(),
                'name': str(row[1]).strip(),
                'person': str(row[2]).strip() if row[2] else ''
            })
    return clients

def sheet_name_for(client_name, location):
    return f"{client_name} {location}"[:31]

def find_sheet(wb, client_name, location):
    """Find sheet by either naming convention."""
    opts = [
        f"{client_name} {location}"[:31],
        f"{location} {client_name}"[:31],
    ]
    for name in opts:
        if name in wb.sheetnames:
            return wb[name], name
    return None, None

def add_client_to_excel(location, name, person):
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Client Details']
    row = ws.max_row + 1
    for r in range(3, ws.max_row + 2):
        if ws.cell(r, 1).value is None:
            row = r
            break
    even = row % 2 == 0
    for col, val in enumerate([location, name, person], 1):
        c = ws.cell(row=row, column=col, value=val)
        d_style(c, even)

    sname = sheet_name_for(name, location)
    existing, _ = find_sheet(wb, name, location)
    if existing is None:
        cs = wb.create_sheet(title=sname)
        _build_client_sheet(cs, name, location, person)

    wb.save(EXCEL_FILE)

def delete_client_from_excel(name, location):
    """Remove client row from Client Details sheet and delete their sheet."""
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Client Details']

    # Find and delete the row
    row_to_delete = None
    for r in range(3, ws.max_row + 1):
        cell_loc = ws.cell(r, 1).value
        cell_name = ws.cell(r, 2).value
        if cell_loc and cell_name:
            if str(cell_loc).strip() == location and str(cell_name).strip() == name:
                row_to_delete = r
                break

    if row_to_delete:
        ws.delete_rows(row_to_delete)
        # Re-apply alternating styles
        for r in range(3, ws.max_row + 1):
            even = r % 2 == 0
            for col in range(1, 4):
                cell = ws.cell(r, col)
                if cell.value is not None:
                    d_style(cell, even)

    # Delete the client's sheet if it exists
    sheet, sname = find_sheet(wb, name, location)
    if sheet is not None:
        del wb[sname]

    wb.save(EXCEL_FILE)
    return True

def _build_client_sheet(ws, client_name, location, service_person):
    # Row 1: Full-width blue title bar
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Waterqo Swimming Pools — Pool Service'
    ws['A1'].font = Font(bold=True, name='Arial', size=14, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill("solid", start_color="0077B6")
    ws.row_dimensions[1].height = 36

    # Row 2: Client Name (A2:B2) and Location (C2:E2)
    ws.merge_cells('A2:B2')
    ws['A2'] = f'Client Name: {client_name}'
    ws['A2'].font = Font(bold=True, name='Arial', size=11, color="003566")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A2'].fill = PatternFill("solid", start_color="BFE9FF")
    ws['A2'].border = thin_border()

    ws.merge_cells('C2:E2')
    ws['C2'] = f'Location: {location}'
    ws['C2'].font = Font(bold=True, name='Arial', size=11, color="003566")
    ws['C2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['C2'].fill = PatternFill("solid", start_color="BFE9FF")
    ws['C2'].border = thin_border()
    ws.row_dimensions[2].height = 26

    # Row 3: Column headers
    for col, h in enumerate(['No.', 'Service Person', 'Price', 'Chlorine Price', 'Paid'], 1):
        h_style(ws.cell(row=3, column=col, value=h))
    ws.row_dimensions[3].height = 24

    MONTHS_LIST = [
        'January','February','March','April','May','June',
        'July','August','September','October','November','December'
    ]
    # 12 monthly blocks, each: 1 month-header row + 8 data rows = 9 rows
    # Starting at row 4
    for mi, mname in enumerate(MONTHS_LIST):
        block_start = 4 + mi * 9
        # Month header row
        mhr = block_start - 1 + 1  # = block_start
        ws.merge_cells(start_row=mhr, start_column=1, end_row=mhr, end_column=5)
        mc = ws.cell(row=mhr, column=1, value=mname)
        mc.font = Font(bold=True, name='Arial', size=10, color="FFFFFF")
        mc.fill = PatternFill("solid", start_color="00B4D8")
        mc.alignment = Alignment(horizontal='center', vertical='center')
        mc.border = thin_border()
        ws.row_dimensions[mhr].height = 18
        # 8 data rows
        for i in range(1, 9):
            r = block_start + i  # row after header
            even = i % 2 == 0
            c = ws.cell(row=r, column=1, value=i)
            c.font = Font(name='Arial', size=10)
            c.border = thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill("solid", start_color="E0F4FF" if even else "FFFFFF")
            for col in range(2, 6):
                cell = ws.cell(row=r, column=col)
                d_style(cell, even)
                if col == 2:
                    cell.value = service_person
                if col in (3, 4):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                if col == 5:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[r].height = 22

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

MONTHS = [
    'January','February','March','April','May','June',
    'July','August','September','October','November','December'
]

def month_start_row(month_name):
    """Each month block: header row + 8 data rows = 9 rows. Months start at row 4."""
    try:
        idx = MONTHS.index(month_name)
    except ValueError:
        idx = 0
    return 4 + idx * 9  # row 4,13,22,31,...

def get_client_entries(client_name, location, month_name=None):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws, _ = find_sheet(wb, client_name, location)
    if ws is None:
        return []
    if month_name is None:
        month_name = MONTHS[datetime.now().month - 1]
    start = month_start_row(month_name)
    entries = []
    for row in ws.iter_rows(min_row=start, max_row=start+7, values_only=True):
        entries.append({
            'no': row[0],
            'person': row[1] or '',
            'price': row[2] or '',
            'chlorine': row[3] or '',
            'paid': row[4] or ''
        })
    return entries

def save_entry(client_name, location, entry_no, person, price, chlorine, paid, month_name=None):
    wb = load_workbook(EXCEL_FILE)
    ws, _ = find_sheet(wb, client_name, location)
    if ws is None:
        return False
    if month_name is None:
        month_name = MONTHS[datetime.now().month - 1]
    start = month_start_row(month_name)
    row = start + int(entry_no) - 1
    ws.cell(row=row, column=2).value = person
    ws.cell(row=row, column=3).value = float(price) if price else None
    ws.cell(row=row, column=4).value = float(chlorine) if chlorine else None
    ws.cell(row=row, column=5).value = paid
    wb.save(EXCEL_FILE)
    return True

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return HTML_PAGE

@app.route('/api/clients')
def api_clients():
    return jsonify(get_clients())

@app.route('/api/clients', methods=['POST'])
def api_add_client():
    d = request.json
    add_client_to_excel(d['location'], d['name'], d['person'])
    return jsonify({'ok': True})

@app.route('/api/clients/delete', methods=['POST'])
def api_delete_client():
    d = request.json
    ok = delete_client_from_excel(d['name'], d['location'])
    return jsonify({'ok': ok})

@app.route('/api/entries')
def api_entries():
    name = request.args.get('name')
    loc = request.args.get('location')
    month = request.args.get('month')
    return jsonify(get_client_entries(name, loc, month))

@app.route('/api/entries', methods=['POST'])
def api_save_entry():
    d = request.json
    ok = save_entry(d['name'], d['location'], d['no'], d['person'], d['price'], d['chlorine'], d['paid'], d.get('month'))
    return jsonify({'ok': ok})

@app.route('/api/persons')
def api_get_persons():
    return jsonify(get_persons())

@app.route('/api/persons', methods=['POST'])
def api_add_person():
    d = request.json
    name = d.get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})
    persons = get_persons()
    if name in persons:
        return jsonify({'ok': False, 'error': 'Already exists'})
    persons.append(name)
    save_persons(persons)
    return jsonify({'ok': True})

@app.route('/api/persons/delete', methods=['POST'])
def api_delete_person():
    d = request.json
    name = d.get('name', '').strip()
    persons = get_persons()
    if name in persons:
        persons.remove(name)
        save_persons(persons)
    return jsonify({'ok': True})

@app.route('/api/download')
def api_download():
    return send_file(os.path.abspath(EXCEL_FILE), as_attachment=True, download_name='pool_service.xlsx')

# ─── Embedded HTML UI ─────────────────────────────────────────────────────────

HTML_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Waterqo — Pool Service Manager</title>
<link rel="icon" type="image/svg+xml" href="data:image/svg+xml,%3Csvg xmlns='ChatGPT Image Sep 9, 2025, 10_44_26 PM-modified.png' viewBox='0 0 100 100'%3E%3Ccircle cx='50' cy='50' r='50' fill='%230077B6'/%3E%3Ctext x='50' y='68' font-family='Georgia,serif' font-size='44' font-weight='bold' fill='white' text-anchor='middle'%3EW%3C/text%3E%3C/svg%3E">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#e8f6fd;color:#003566;min-height:100vh}
:root{
  --navy:#003566;
  --blue:#0077B6;
  --sky:#00B4D8;
  --light:#90E0EF;
  --pale:#CAF0F8;
  --white:#FFFFFF;
  --card:#FFFFFF;
  --border:#BFE9FF;
  --green:#0a9960;
  --red:#d62828;
  --bg:#e8f6fd;
  --sidebar:#f0faff;
}

/* ── Header ── */
header{
  background:linear-gradient(135deg,var(--navy) 0%,var(--blue) 100%);
  padding:0 32px;
  height:68px;
  display:flex;align-items:center;justify-content:space-between;
  border-bottom:3px solid var(--sky);
  box-shadow:0 4px 24px rgba(0,119,182,.25);
}
.header-brand{display:flex;align-items:center;gap:14px}
.logo-circle {
  width: 44px;
  height: 44px;
  border-radius: 50%;
  background: white;
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 1.4rem;
  font-weight: 900;
  color: var(--blue);
  font-family: Georgia, serif;
  box-shadow: 0 0 0 3px rgba(255, 255, 255, .3);
  flex-shrink: 0;
}

.logo-circle img {
  width: 28px; /* Adjust the size of the image */
  height: 28px; /* Adjust the size of the image */
  border-radius: 50%; /* Optional, to make image round */
}
.header-text{}
.header-text h1{font-size:1.25rem;font-weight:700;color:#fff;letter-spacing:.5px;line-height:1.1}
.header-text p{font-size:.72rem;color:var(--light);letter-spacing:1.5px;text-transform:uppercase;font-style:italic}
.btn-download{
  background:var(--sky);color:var(--navy);
  border:none;padding:9px 20px;border-radius:8px;
  cursor:pointer;font-size:.85rem;font-weight:700;
  transition:.2s;text-decoration:none;letter-spacing:.3px;
}
.btn-download:hover{background:var(--light);transform:translateY(-1px)}

/* ── Layout ── */
.layout{display:grid;grid-template-columns:320px 1fr;gap:0;height:calc(100vh - 68px)}

/* ── Sidebar ── */
.sidebar{background:var(--sidebar);border-right:2px solid var(--border);overflow-y:auto;display:flex;flex-direction:column}
.sidebar-header{padding:18px 16px 12px;background:#dff3fb;border-bottom:1px solid var(--border)}
.sidebar-header h2{font-size:.8rem;color:var(--blue);text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;font-weight:700}
.search-box{
  width:100%;padding:9px 14px;
  background:#fff;border:1.5px solid var(--border);
  border-radius:8px;color:var(--navy);font-size:.9rem;outline:none;transition:.2s;
}
.search-box:focus{border-color:var(--sky);box-shadow:0 0 0 3px rgba(0,180,216,.15)}
.search-box::placeholder{color:#9bbccc}

.client-list{flex:1;padding:10px}
.client-card{
  background:#fff;border:1.5px solid var(--border);
  border-radius:10px;padding:13px 15px;margin-bottom:8px;
  cursor:pointer;transition:.2s;position:relative;
}
.client-card:hover{border-color:var(--sky);background:#f0faff;transform:translateX(3px)}
.client-card.active{
  border-color:var(--blue);
  background:linear-gradient(135deg,#dff3fb,#e8f8ff);
  box-shadow:0 0 0 2px rgba(0,119,182,.2);
}
.client-card h3{font-size:.92rem;font-weight:700;color:var(--navy);margin-bottom:3px}
.client-card .meta{font-size:.76rem;color:#5a8fa8}
.client-card .badge{
  position:absolute;right:12px;top:50%;transform:translateY(-50%);
  background:var(--blue);color:#fff;font-size:.65rem;
  padding:2px 8px;border-radius:20px;font-weight:700;
}

.add-client-btn{
  margin:10px;background:var(--blue);color:#fff;border:none;
  padding:12px;border-radius:10px;cursor:pointer;font-size:.88rem;
  font-weight:700;transition:.2s;display:flex;align-items:center;
  justify-content:center;gap:8px;letter-spacing:.3px;
}
.add-client-btn:hover{background:var(--navy);transform:translateY(-1px)}

/* ── Main panel ── */
.main{display:flex;flex-direction:column;overflow:hidden;background:var(--bg)}
.main-header{
  padding:20px 28px;background:#fff;
  border-bottom:1.5px solid var(--border);
  display:flex;align-items:center;justify-content:space-between;
  box-shadow:0 2px 8px rgba(0,119,182,.06);
}
.main-header h2{font-size:1.25rem;font-weight:700;color:var(--navy)}
.main-header .info-chips{display:flex;gap:8px;margin-top:5px}
.chip{
  background:var(--pale);color:var(--blue);
  padding:4px 12px;border-radius:20px;font-size:.78rem;font-weight:700;
}
.main-header-btns{display:flex;gap:10px;align-items:center}

.empty-state{
  flex:1;display:flex;flex-direction:column;align-items:center;
  justify-content:center;color:#7ab8cc;text-align:center;padding:40px;
}
.empty-state svg{opacity:.25;margin-bottom:20px}
.empty-state h3{font-size:1.2rem;margin-bottom:8px;color:#5a9ab5}
.empty-state p{font-size:.88rem;color:#8ab8cc}

.entries-area{flex:1;overflow-y:auto;padding:22px 28px}

.entries-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(290px,1fr));gap:14px}
.entry-card{
  background:#fff;border:1.5px solid var(--border);
  border-radius:12px;overflow:hidden;transition:.2s;
}
.entry-card:hover{border-color:var(--sky);box-shadow:0 4px 20px rgba(0,180,216,.12)}
.entry-card .entry-header{
  background:linear-gradient(135deg,var(--blue),var(--sky));
  padding:11px 15px;display:flex;align-items:center;justify-content:space-between;
}
.entry-card .entry-num{font-size:1rem;font-weight:700;color:#fff}
.entry-card .entry-status{font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:700}
.status-paid{background:rgba(10,153,96,.25);color:#0a9960;border:1px solid rgba(10,153,96,.4)}
.status-unpaid{background:rgba(214,40,40,.2);color:#d62828;border:1px solid rgba(214,40,40,.3)}
.status-partial{background:rgba(230,150,0,.2);color:#c47a00;border:1px solid rgba(230,150,0,.3)}
.status-empty{background:rgba(0,119,182,.15);color:var(--blue);border:1px solid rgba(0,119,182,.2)}
.entry-body{padding:14px}
.entry-field{
  display:flex;justify-content:space-between;align-items:center;
  margin-bottom:9px;padding-bottom:9px;border-bottom:1px solid var(--pale);
}
.entry-field:last-child{margin-bottom:0;padding-bottom:0;border-bottom:none}
.entry-field label{font-size:.75rem;color:#7aaabb;text-transform:uppercase;letter-spacing:.5px;font-weight:600}
.entry-field .val{font-size:.92rem;color:var(--navy);font-weight:600}
.edit-btn{
  width:100%;background:transparent;border:1.5px solid var(--blue);
  color:var(--blue);padding:9px;border-radius:8px;cursor:pointer;
  font-size:.82rem;font-weight:700;transition:.2s;margin-top:10px;
}
.edit-btn:hover{background:var(--blue);color:#fff}

/* ── Summary bar ── */
.summary-bar{
  background:#fff;border-top:1.5px solid var(--border);
  padding:13px 28px;display:flex;gap:28px;align-items:center;
  box-shadow:0 -2px 8px rgba(0,119,182,.06);
}
.sum-item{text-align:center}
.sum-item .val{font-size:1.1rem;font-weight:800;color:var(--blue)}
.sum-item .lbl{font-size:.7rem;color:#8aaabb;text-transform:uppercase;letter-spacing:.5px;margin-top:1px}

/* ── Delete client btn ── */
.btn-delete-client{
  background:transparent;color:#d62828;border:1.5px solid #d62828;
  padding:9px 16px;border-radius:8px;cursor:pointer;font-size:.82rem;
  font-weight:700;transition:.2s;
}
.btn-delete-client:hover{background:#d62828;color:#fff}

/* ── Modal ── */
.overlay{
  position:fixed;inset:0;background:rgba(0,53,102,.55);z-index:100;
  display:flex;align-items:center;justify-content:center;
  backdrop-filter:blur(5px);
}
.overlay.hidden{display:none}
.modal{
  background:#fff;border:1.5px solid var(--border);
  border-radius:16px;width:460px;max-width:95vw;padding:28px;
  box-shadow:0 20px 60px rgba(0,53,102,.2);animation:pop .2s ease;
}
@keyframes pop{from{transform:scale(.92);opacity:0}to{transform:scale(1);opacity:1}}
.modal h3{font-size:1.15rem;font-weight:700;color:var(--navy);margin-bottom:4px}
.modal .modal-sub{font-size:.83rem;color:#7aacbb;margin-bottom:18px}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}
.form-row.full{grid-template-columns:1fr}
.field{display:flex;flex-direction:column;gap:5px}
.field label{font-size:.77rem;color:var(--blue);font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.field input,.field select{
  background:#f0faff;border:1.5px solid var(--border);
  color:var(--navy);padding:10px 13px;border-radius:8px;
  font-size:.9rem;outline:none;transition:.2s;
}
.field input:focus,.field select:focus{border-color:var(--sky);box-shadow:0 0 0 3px rgba(0,180,216,.12)}
.field select option{background:#fff}
.modal-actions{display:flex;gap:10px;margin-top:20px}
.btn-primary{
  flex:1;background:var(--blue);color:#fff;border:none;
  padding:12px;border-radius:8px;cursor:pointer;font-weight:700;
  font-size:.92rem;transition:.2s;
}
.btn-primary:hover{background:var(--navy)}
.btn-cancel{
  flex:1;background:transparent;color:#7aaabb;
  border:1.5px solid var(--border);padding:12px;border-radius:8px;
  cursor:pointer;font-weight:600;font-size:.92rem;transition:.2s;
}
.btn-cancel:hover{border-color:var(--blue);color:var(--blue)}
.btn-danger{
  flex:1;background:#d62828;color:#fff;border:none;
  padding:12px;border-radius:8px;cursor:pointer;font-weight:700;
  font-size:.92rem;transition:.2s;
}
.btn-danger:hover{background:#a81e1e}

/* ── Toast ── */
.toast{
  position:fixed;bottom:28px;right:28px;
  background:var(--blue);border:1.5px solid var(--sky);
  color:#fff;padding:13px 20px;border-radius:10px;
  font-size:.88rem;font-weight:600;z-index:999;
  animation:slide-in .3s ease;box-shadow:0 4px 20px rgba(0,53,102,.2);
}
.toast.success{background:var(--green);border-color:var(--light)}
.toast.error{background:var(--red);border-color:#f4a7a7}
@keyframes slide-in{from{transform:translateX(100%);opacity:0}to{transform:translateX(0);opacity:1}}

::-webkit-scrollbar{width:6px}
::-webkit-scrollbar-track{background:var(--pale)}
::-webkit-scrollbar-thumb{background:var(--light);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--sky)}

/* ── Month selector ── */
.month-bar{
  background:#fff;border-bottom:1.5px solid var(--border);
  padding:10px 24px;display:flex;gap:6px;flex-wrap:wrap;align-items:center;
  box-shadow:0 1px 4px rgba(0,119,182,.06);
}
.month-bar .lbl{font-size:.72rem;color:#8aaabb;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-right:4px;white-space:nowrap}
.month-btn{
  background:transparent;border:1.5px solid var(--border);
  color:#7aaabb;padding:5px 11px;border-radius:20px;
  cursor:pointer;font-size:.76rem;font-weight:700;transition:.15s;
}
.month-btn:hover{border-color:var(--sky);color:var(--blue)}
.month-btn.active{background:var(--blue);border-color:var(--blue);color:#fff}
</style>
</head>
<body>

<header>
  <div class="header-brand">
    <div class="logo-circle" style="width:48px;height:48px;padding:0;overflow:hidden;">
  <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAPpA+kDASIAAhEBAxEB/8QAHQABAAICAwEBAAAAAAAAAAAAAAcIBgkBBAUCA//EAGYQAAEDAwEFAwYFDAsMBwcEAwABAgMEBREGBwgSITETQVEUIjJhcYFCUnKRoRUWI2J1gpKisbKztAkzNDY3OFNzk8HCGCQ1Q1RWY2WUxNLTFyVVdpWj0SYoRIOkw+EnZITwdIXx/8QAGwEBAAIDAQEAAAAAAAAAAAAAAAUGAQQHAwL/xABCEQEAAQIDAwkECAUEAgIDAAAAAQIDBAUREiExBhNBUWFxgZGxNKHB0RQiMjM1cuHwFSNCUrIWJGKCU/El0iZDov/aAAwDAQACEQMRAD8ApkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA5a1znI1qK5yrhEROaqSBo/YptV1ZwPsuhrw+F/oz1EXk0Sp4o+VWtVPYoEfAtHpPcr17XoyTUeo7LZI3dWQo+qlb7UThb8zyWtLbl+zi38Ml9vN9vUqdWpIymid961Fd+OBQI7lqtV0u1R5ParbWV838nTQOld8zUVTaLpnYVsh07wLbtA2V72ejJWRLVPRfHMyuVFJAoqOkoadtNRUsFLA30Y4Y0Y1PYicgNXWn9ge2O+cK0Wz68xo7otZG2kT/AM5Wmf2Pc52t1/CtdJp60tX0kqK1z3J7omOT6TYaAKVWbcdr34dedodND4spLY6TP3zpG4+YzK07k+z6HC3PVGpaxydUhdDC1fcrHL9JaMAQPbN0rYtSY8ostyuGP8ouUqZ/o1aZJb93bYrQ47HZ/bX4/l5JZvz3qSoAMJo9kWyyjx5Ps50m1U6OdaIHO+dWqp69NonRtLhKbSVggx07O3Qt/I098AdGGzWiD9ptVDF8inYn5EO0yCFnoQxt9jUQ/QAMHw+KJ/pxMd7Wop9gDpTWm1Tft1topPlwNX8qHQqdHaRqs+U6Vsc2evaW+J35WnuADDKzZRsvrM+UbOtJPVfhfUeBHfOjcnhV+75sYrs9ts+tLM/yHHD+Y5CUABBdz3T9idXnsNPV1Aq99Pc51x+G5xit23Ktm8+XW7UOp6Jy9EfNDKxPd2aL9JZ4AUxvO47ImX2faIx3hHV2vH47ZF/NMHvm5ntUouJ1vuGnLoz4LY6t8b19z2In4xsGAGru/wC7vtnsvEtVoG5VDW99E6Oqz7Eic5foI9vdgvtjl7K92W5WyTOOCspXwrn2ORDcMfE8MU8LoZ4mSxPTDmPajmuT1ovUDTUDa3qbYzsq1HxLdtBWF73+lLDSpBIvtfHwu+kizVO5vssuaOfZqq+WKRfRbFUpPEntSRFcv4SAa9wWu1ZuS6tpEfJpjV9ourE5pHWQvpXqngmONqr7VT3EP6w2BbXtLcb7joe5zwM5rNQNSrZjx+xK5UT2ogEYg/Sohmp5nwVEUkUrFw9j2q1zV8FReh+YAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAehYbLeL/AHFltsdqrbnWyehBSQOlevuaiqT9s53P9pOoeyqdSTUOlaN3NUnd29TjxSJi49zntX1AVxPW0zprUOp65KHTljuV3qeWY6OmfKqeteFFwnrXkbBtn26bsp0wkc90o6rU9a3mr7jJiJF9UTMNx6ncRONotdss9CygtFupLfSR+hBSwtijb7GtREQDX3oXdB2pX5I5r4tt0zTO5qlVN202PVHHlPc5zSd9Ebmmzi08EupbndtSTJjiYr/JYHfesXjT8MswAMW0ds60Lo9jU0zpOz2x7f8AHQ0re2X2yLl6+9TKQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADA8HVejNJasgWHUumrTd24wi1dKyRzfkuVMt9yoQprfc/2V3vjlsa3TTVQ7m1KaoWaHPrZLlcepHIWJBka/9dbmu0Wz9pNpi52vUsDc8MaO8lqHfevVWfjkD6x0Tq7R1T5PqjTd0tD1XDXVVM5jH/Jfjhd7UVTbwfjW0tNW0slLWU8NTTypwyRSsR7Hp4Ki8lQDTaDZbtB3X9kerUkmisTtP1r8r29oekKZ/mlRY8exqL6yuW0Xcz1zZ+0qdHXah1JTNyqQSf3rU+xEcqsd7eJPYBV8Hs6s0rqXSVxW36msVwtFTzwyrgdHxetqqmHJ60yh4wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD7hiknmZDDG+SWRyNYxjcucq8kRETqoHwCf9ku6ltG1l2NdfIm6TtT8L2lcxVqXt+1gTCp9+rS2+yrdv2YaB7KqjtH1curML5ddESZWu8WR44GepUTPrUCiuy/YTtN2idnPY9PS09ukwqXCvzT0+PFqqmXp8hHFptmW5npC09lWa5u9VqGqTm6kp809Ki+Cqi9o/25b7C0wA8jSml9OaUtqW3TVkt9opExmOkgbGjlTvdhMuX1rlT1wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOjfLPab7bZLbe7ZR3Kil9Onq4Gyxu9rXIqFe9pu59s91F2tXpWpq9K1zsqkcWZ6VV9cbl4m/euRE8CyQA1kbUN2/anoNJaqayfVu2MyvltqzO1G+LmYR7eXVVbhPFSH1RUXCphUNy5Ge1TYVs12jJLPfLDHTXKTP8A1lQYgqc+LlRMPX5aOA1ZgsvtZ3P9cab7Wv0ZVR6pt7cu7FqJDWMT5Crwv+9XK/FK4XKhrbZXTUFxo6iiq4XcMsFREscka+DmuRFRfaB1wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP0poJqmojp6aGSaaRyMjjjarnPcvREROaqTxsV3W9e68SC53uN2l7G/DknrIl8omby/a4eS4VOjncKd6ZLrbIdiuz/ZhTsdp2ztluXDwyXOrxLVP5YXDsYYi/FYjUAp9se3R9c6rSG46wk+tO1uw7spWcdbInqi6R/fqip8VS4myrYxs82aQsdpuxRLXo3D7lV4mqn+Pnqnm58GI1PUSGAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYftK2ZaH2i0HkurdP0tc9rVbFUonBUQ/IlbhyeOM4XvRTMABRHbDucajsyTXPZ3cPq/RNy7yCpVsdWxPBruTJOXyV7kRSsF3tlxs9yntt2oKqgrad3BNT1MTo5I18Fa5EVDcaYbtP2YaI2kW3yPVtjgrHsbww1TU4KiD5EieciZ545oveigamQWY20bour9LJNdNDyyaotTcuWmRqJXRJ8hOUv3nP7UrVPDLTzyQTxPiljcrHse1Wua5FwqKi9FRe4D4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB6Wm7FedS3mCzWC2VVyuFQ7hip6eNXvX18uiJ3qvJO8udsH3P7fbkgvm1GVlwrEVHss9O/wCwRr1TtXpzkX7VMN9bkArPsa2La72p1ifUC2rBbGv4ZrpVorKaPnzRHYy9yfFblfHHUvPsR3b9BbNWwXCWnTUGoWIjluNbGnDE7xhj5pH7ebvtu4mKgo6S30cVFQUsFJSwtRkUMMaMZG1OiNanJE9SH7gAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACMNs2wzQO1GB815t3kV44cR3WiRGToqJyR/LEjfU5F9Sp1JPBgax9t+7zrzZe6Wumpvq1YGryulFGqtYn+lZzWL2rlv2ykPm5aRjJI3RyMa9jkVHNcmUVF6oqFadvO6bpnVqVN70I6DTl7dl7qXhxRVLvDhT9qX1tTH2veZGv8GQa90ZqfQt/lseqrPU22sjVcJI3zJW/GY9PNe31tVUMfAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB3LLa7jerrTWm0UNRXV9VIkcFPBGr3yOXuREA6ZNOwLd21jtRliuUzHWPTXEivuNRGuZk70gZ8NftuTU8VVME97u26VQWhKbUe1COG4XD04rKio+nhXuWZU5SO+1TzfHi7rZQxxwxMhhjZHGxqNYxiYRqJyREROiAYZsm2XaM2Y2VLdpW1shke1EqK2XD6mpXxe/HT7VMNTuRDNQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMd2g6I0vr2wSWTVdop7jSOyrONMSQuVMccb05sd60X6CiW8Fuuan0F5RfdJ+Uai041Vc5GszV0jf9I1qee1PjtT2o3qbDwBpnBsI3h91rT2uEqdQaLSnsGolRz3wtbw0la7r5yJ+1vX4zUwvei5ylENZaXv+jtQVFg1Na6i23GnXz4Zm9U7nNVOTmr3ORVRQPGAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJ/3at22+7SpqbUGomz2fSWeJJccM9ciL6MSL0b4yLy8EXngI72N7KNX7U7+lt03QqlLG5ErLhMipT0rV73O718GplV9mVTYhsL2J6P2TWns7TB5deJW4qrtURp20n2rf5Nn2qerKqvMzXR2mLDo/T9NYNNWynttupkxHDC3HPvc5ernL3uXKr3nsAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADCtruzDSO1DTrrRqe3o97UXyWtiRG1FK74zH4+dq5avehmoMDV7t92Gat2TXR8lZE65aflk4aW7Qx4jdnoyROfZv9S8l7lXniKTcbeLZb7za6m13aigrqGqjWOennYj2SNXqiovJSjG81us3DS61OqtnNPPcLE1FlqbblX1FGnVVZ3yRp+E1OuUyqZFWQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPpjXPe1jGq5zlwiImVVfA7Fpt1fdrnTWy2Uc9ZW1UiRQQQsV75HquEaiJzVS/G6vu1Ueg20+rtbQwV2qFRH01NyfDbvWnc6X7bo34PxlDBN1zdX7dlPrDalQObGuJKKxyphXd6PqE8PCP8LvaXPhjjhhZDDGyONjUaxjEw1qJyREROiH0AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKsb0O69Q6mjqtXbOqSGiv2XS1VtbhkNcvVXM7mSr+C71LlVotcaKst1fPQXClmpKunesc0MzFY+NyLhWuReaKhuQIP3mt36y7Vre+7WzsLXq2CPENZw4ZVIico5sdU7kf1b605Aa1wevrDTV80jqKq0/qO3TW+5UruGWGVOfqci9HNVOaKnJU6HkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPS0xYrxqa+0lisNvnuFxq5Ejgghblzl/IiJ1VV5IiKq8jsaJ0vfdZ6mo9Oact8ldcqt/DHGzoid7nL0a1E5qq8kQ2SbuWw+wbI7Ajm9lcNSVUaJX3JW+/sos82xovvcqZXuRA8ndh2AWjZTa2Xa6pBcdXVMeJ6pEyylaqc4oc93cr+rvUnInAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjDeC2M6d2uab8mrkbRXqlYv1PubGZfEvXgf8aNV6t7uqYU1s7RtE6j2f6pqdOanoHUlbCuWqnOOZndJG74TV8fcuFRUNupge2zZXpnatpV9mvsPZVMaK6hr42p21JIveni1eWWryVPBURUwNUYMv2tbO9SbMtXz6c1JS8ErfPp6hmVhqos8pI170Xw6ovJcKYgZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPa0Tpe+az1NR6c05QSV1yrH8McbOiJ3ucvRrUTmqryREOvpixXbU2oKKw2KilrrjWypFBBGnNzl/IiJlVVeSIiqvJDZZu1bFbRsj0twuSKs1JXMRblXo339lHnmkaL73KmV7kQP23dNi9i2RaY7GHs66/1bE+qNx4eb169nHnm2NF6J1Vea9yJKgBgAAZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgYTtm2Z6b2p6PlsF/h4ZG5fRVsbU7Wklxye1e9PFvRye5U1mbW9neo9mWsKjTeo6bhkb59NUMReyqos8pGL3ovenVFyi80NtJgm27Zfp7aro2Ww3uNIqhmZKCuY1FkpJcekni1eSOb0VPBURUDVEDJdpmiL/ALPdYVmmNR0qwVlOuWPbzjnjX0ZGL3tXH5UXCoqGNGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAO1abfXXa501stlJNWVtVK2GCCJqufI9y4RqInVVU67Gue9rGNVznLhERMqq+BsB3ONgKaDt8WttXUqLqisi/vamkT/B0Tk6L/AKVydfip5vxshku6tsIoNlNhS63ZkNXq6uiRKqdMObSsXn2Ea+HTicnpKngiE4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGBGW8Lsfse1zSK2+rSOkvNKjn2y48OXQvX4Lu9Y3YTKexU5ohrN1rpi96N1PXab1FQvorlRScEsbui+Dmr8JqphUVOSopt/IU3qdh9FtY0ylbbWxU2q7dGvkNQ7DUqG9VgkXwVfRX4Kr4KucjWkDtXa311pudTbLnSzUlbSyuhnglbwvje1cK1U7lRTqgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJ/3Pdh8m0rU6ai1BSO+tK1yp2qOyiV0yYVIU8WpyV6+GE78oEkbkGwVs6Ue1LWFJmNF7Sx0UrfSVOlS9F7viJ998VS6J8wxRwwshhjZHGxqNYxiYRqJyRETuQ+gAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACsO+jsFbrG2T7QNJ0v8A7RUUOa6mjbzr4Wp1RE6ytTp8ZEx1RqFBTcwUX34dhf1Br5tpWkaBG2iqfm700LeVLM5f25ETox6rz8HL4O5BU8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA7lktdwvV3pLRaaSWsr6yZsNPBEmXSPcuERPeBl+wzZpd9qev6TTdta+KlyktwrEbltLAi+c9fFV6NTvVU7sqm0XRemrPo/S9BpuwUjaW3UESRQxpzXxVzl73KuVVe9VVTCt3DZRb9k2z+K0x8E14rOGe61Sf4ybHoN+0ZlUT3r1cpJoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA69zoaO526pt1wpoqmkqonQzwyt4myMcmHNVO9FRTsAwNY29Jsdqtk2uXR0cc02mriqy2ypfz4U+FC5fjMz170VF8USIDbXtc0DZdpWhK7Sl7aqRVCI+CdqZfTTNzwSt9aKvTvRVTopqz2g6SvWhtYXDS9/plgrqGVWO+LI3q2Rq97XJhUXwUyPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC8m4dsZS0Wpm1DUdJ/wBYV0apZYn9YYHJhZsdznpyb9rlfhcoE3SNkMm1HaC2a5QO+tq0ObPcXqnmzLnzIEXxdhc+DUXvVDZbDFHDCyGGNkcbGo1jGJhrUTkiIidEA+gAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAr9vnbGU2i6N+uOxUiP1TZoldGjPSq6dMq6H1uTm5vrynwuVgQYGmcFmd+fY79Z+qvr8sFLw2K9TL5XHGzDaSqXmvTo2Tm5PB3EnLkhWYyAAAAAAAAAAAAAAAAAAAAAAAAB6WmLJc9S6hoLBZ6Z1TcK+dsFPE3vc5cc/BE6qvciKp5peDcA2SpbrTJtRvlMqVlc10FnY9MLHB0fNjxeqcKL8VF7nAWD2J7PLZsx2eW/StuRskkTe1rahEwtTUOROORfVyRETuaiJ3GagAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGB4uudMWjWWkrlpi+0/b2+4QrFK3vb3o5q9zmqiORe5UQ1VbVtEXbZ3ry56TvDczUcn2OZGqjZ4l5skb6nJj2LlOqKbbyvW+7slTXmz9dUWimV+odPxukajEy6ppfSkjwnVU9NvsciekBrsABkAAAAAAAAAAAAAAAAAAAAAEkbuWzSp2p7TqGwcMjLZCvlN0mbn7HTtVOJEXuc5cNT1uz3KbSbfSU1BQU9BRQMp6WmibDDExMNjY1ERrUTwRERCHdz/ZY3ZrsugluFMkeob0jau4q5uHxJj7HCvyEVVX7ZzvUTSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGBrb3y9lH/RztKfcrXT8Gnr659TRo1q8MEufssPqwq8TU+K5E7lILNrW37Z1R7UNmVy0zO2Ntbw9vbp3J+0VLUXgX1IuVav2rlNVtzoqu2XKpt1fA+nq6WZ8E8L0w6ORqq1zV9aKioZHXAAAAAAAAAAAAAAAAAAAnzcl2YfX5tRZe7nTdpY9OqyqnRzctmnzmGPwVMorlTwbhepA9NBNU1EVNTxPlmlejI42Jlz3KuERE71VTadu5bOYdmGyq2adcxn1Skb5Vc5G4XjqXonEmU6o1ERiL4NRe8CRgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACi37ILsw+pGpaXaTaabhors5Ke5oxvKOpRvmyLjoj2phftmqvVxekxzaZpC2680Jd9JXZv97XGnWNH8KKsT+rJE9bXI1yewwNRQPU1bYblpfU1y07eIexr7dUvp52d3E1cZRe9F6oveiop5ZkAAAAAAAAAAAAAAA+4YpJ5mQwxvklkcjWMY3LnKvJERE6qBY7cJ2bfXXtMfrC4wcdr03wyxcScpKx2eyT7xEV/qVGeJsHI93dtn8ezXZLZ9NuY1K9WeVXJ7fh1MiIr+ffw8mIvgxCQjAAAyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADApX+yI7NuxrbdtOtkHmVHDQXbhT4aJ9hlX2tRWKv2rE7ynZt52i6Vt+t9D3jSl0T+9bnTOhc7GVjd1Y9E8WuRrk9bUNTGqbJcNNakuOn7tCsNdbql9NOzuRzHKiqnii4yi96KimR5oAAAAAAAAAAAAAT/uMbPPry2vR32ug7S1aba2sk4ky19QqqkDfnRX/wDyyADZtuf7P/rB2K2yOqg7O63f/rKuynnNWRE7Ni+HCxGpjxV3iBMQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwBRn9kQ2efU3U9t2jW+DhprqiUdwVreSVDG/Y3L63Roqf/L9ZeYwvbhoiDaJsuvmlJWs7aqp1dSPd/i6hnnROz3JxIiL6lVO8DU2D9aynno6uakqonwzwSOjljemHMc1cKi+tFQ/IyAAAAAAAAAAAkzdi0H/ANIm2ayWOeLtLdBJ5dcUVMp5PEqKrV9TncLPvzaYVZ/Y7tC/UnQFz1zWQcNVfJ+wpHKnNKaJVRVT5UnFn+baWmAA4VzUXGU+ccSYzlAOQcKqY6jiTxA5BxxJ4jIHIPlXInVUOUVF6KByAAAOFVE6qiDiTxA5BxxJ4p84yniByDjPrQZTxA5BwqjiTxA5A7gAAAABVwcZA5BxkKoHIOMnIAAAAAAAAAAAAAAAAAAABkEOb2WvXaO2bvtdvnWK9ag46OlVvpRRY+zyp8lqo1F6o57V7jZweFrxd+mzb41T+58OLzu3It0TXPQzHZVreHXlFfLnRNh+p1HeZ7fRSxu4u3jiaxFlVennOV6pj4PD35MyII3HY2RbHKuONvC1t8qmongiJGifkJ3PbNMPRhsZcs0cKZ0fGHrm5biqekABoPcAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYGubfr0Gmkds8t6pIuC3aljWuZhOSToqJO328Stev84QAbJd9zQv147EK6upYO0uOn3fVKBUTmsbUxM32cCq72sQ1tGQAAAAAAAAPQ05aK2/6gt9jtsfa1twqY6WBni97kan0qeeWO/Y/9F/XDtil1JUw8dHpylWZFVMp5RLlkSfg9o72sQC+eitP0OlNI2nTVtTFJbKSOmjXGFcjGonEvrVcqvrVT1wAIA3udj8WsLL9edjt6z6htUK9vBGmXV1MmVViJ3yM5ubjmqZbzy3FMY6Ohkja9kEbmuTLV580NppRfel2dN0LtBW4W2BI7Ff3SVNM1rcNp6hFzNCiJ0Rco9qckwrkT0ToPJDOZn/ZXZ7afjHxjx7EJmuHqinnaJ70OeQ0f+SxL7UHkNH/ksXzHY5gvmsoDnK+uXX8hov8AJYvmHkFF/ksXzKdgc1XHeNZOcr65fVosEt4u1HZ7PbUqrjXTNgpYG/De5e9e5E5qqryREVV6GwvYrs7tezPQ1Np6gXtqhy9vX1S//EVDkRHPRPgt5IiJ3IidVyqwtuS6Bj8lq9pFxha+SZz6Oz8TUXgjavDNMnrc5FYi8lRGO7nFnzmnK3N5v3volufq08e2f09dexZssw80W+cr4yAApqUQ/vI7GKDaZaUulAyOHVFBCrKOV7lSOoYiq7sJPtVVVw7q1V8MoUYqrRHR1c9FXW3yWsppXQ1EEzcPikauHNd60U2jlZ98rZk2ponbSrJTJ5VSMbHeoo2c5qdOTZ+XV0fRVxzZ3ojC68ls9qs1xg70/Vn7M9U9XdPulFZlhJrp5y3xj3qmrQ0f+Sx/MceQ0f8AksXzHZUHSNZVrnK+uXW8ho/8lj+YeQ0ef3NH8x2Tgayc5X1y6/kFF/ksfzE57pOyOLVOp2azutKrLFZqhFpWI9U8rrGKjk5Z9CNcKvcrsJzRHIRTpDT1w1bqu16YtK8Nbc6hIWPVMpE3GZJVTvRjEc5fYbGdJWC2aW0zb9PWeBIKCggbBC3CZVE6udjq5Vy5V71VV7yqcqs4nB2IsW5+vX7o/XhHil8qsVXaucrndD1UAByxYwAGR4evdL2zWekLjpm8NkWir4uB7o3cL2ORUc17V+M1yNcndy5oqcjXXq3Tly0nqm56Zu7pUrrbOsUjmyvRsrerJW8/Re1UcntNmBWvfe0S2pslv2g0cadvbVbRXFU+FTSP+xvXn8CR2OXdIvgW3knmk4bE/Rq5+rXw7Kujz4eSMzPDzctbdPGFT8L/ACk/9O//ANR538rP/Tv/APU56KqA6fpEqvtVdb5kekUTpHy1KNamVxUP/wDUv1u26DfoLZfQ0ddG9LzX/wB+3JXuVXNleiYj5qvoNRrOS4y1V7yqW7Lo9ustr9thqou0t1pT6p1jV6O7NyJExe5cyK1VRerWOL8FA5Z5hvpwdHfPwj4+SwZPZnZm7V4AAKEmwAAAAAAAAAAAAAABgcKuEya+NvOuf+kHabcb1TzJJaqXNDa8dFgY5cyffv4nePDwp3Fod7vXLtKbM32agm7O7aiV1DCqL50cGPs8iexioxOiosjVToUjRGtajGoiNamETHREOhcjcu2aasZXHHdT3dM/DwlA5xid0Wo8VztyH+B+t+7tX+RhOpBW5F/A/WL/AK9q/wAjCdSqZ7+I3vzSlMF9xSAAiG0AAyAAAAAAAAAAAAAAAAAAAAAAADAAAyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA/OpghqaeWmqI2ywysVkjHJlHNVMKip4Khqa2zaPl0DtQ1BpORH9nQVjkpnO6vgd50TvexzV9uTbSUn/AGR/RfYXXT2v6WHDKli2ytcicuNuXxKvrVvaJ7GIYFPwAZAAAAAANjm4lo762dhlLdJ4uCt1BUPr3qqeckXoRJ7OFvGn84a+dHWKr1Pq206coUXym51sVJGuM4V70blfUmcr6kNu9kttJZ7NQ2igj7OkoaeOmgZ8WNjUa1PmRAO2AABH+8Joz6+dlN3tEEPaXGCPy22qjcuSpiRXMRPlJxR+x6kgBT1w9+vD3abtHGmdY8HzXTFdM0z0tXMMjZYmSs9F6I5PV6j7Mv206bbpLazqawxMaymirVqKVrW4a2GdO1Y1PU3iVv3piB3KxdpvW6btHCqImPHepF2ibdc0T0B+dU90dLNI1cK2NyovrwuPpP0PyqmdpTTMTq6NyJ7cHrD5o02o1bLdBWWDTeibJYKZE7K30ENMiomOJWsRFcvrVcqvrU9sx/Ztc3XvZ3pu8PXL661UtQ72via5fpUyA4RfirnatvjrOvevNH2YAAeL6D8qyngq6WakqoY54JmOjlikajmvaqYVqovVFRcYP1BmJ03wNcu1jR8mgdol30qqyPpqaRJaGR+fslLInFGuV6q3mxV8WKYuWn36tMMfbdP61p4W9rTTrbKtzWZc6KVFdEqr3I17VT2ylVzs+SY76dgqLs8eE98fPj4qfj7HM35iODk4AUlWkspuLaajnu2o9YTx8TqZsdspHL0arkSSZceOOyTPhnxLXldtxOojdoHUdGi/ZIr2srk+1fTw4/NUsScg5S3armZXdro0iO7SP/a45fTEYejQABBNwABgDxtc2KDU2jrxp6pRvZXKhmpVVyZ4eNitR3tRVRU9aIeyebqq6xWLTF0vk6Zit1HNVyJ4tjYrl+hD0tTVFymaOOsad75r02Z1ayqV6yUsT3LlysTiX1pyX6UP1RMqfjR58khV2eJWIq+1ea/Sp+qxVE2IKRqvqZnNiha3q571RrUT3qh3mdyjzTrXpHWuPuS6Xbatm1XqiWPFRqCrV8bsrnyaHMcaY+V2rvWj0J8PM0pZqXTumLXYKLPk1to4qSJV6q2NiNRV9a4PTOIZli5xmKuX5/qn3dHuXWxbi1biiOgABovUAAAHGfWcoAAAAAAAAAOFXCHJEu9Trh+jNldXDQ1HZ3e9OW30PC7Do+Jq9rKnNFTgZxYcnRys8TYwmGrxV+mzRxqnT99z4uVxbomqehVTeA1smvdqlyu9NMktso/+rrYrVy10Mbl4pE8eN6uci/F4U7jAj5a1sbGsYmGtRGtTwRD6O3YbD0Ye1Tat8KY0hSr12btc1z0rm7kP8DtX93Kv8jCdSCtyH+B6s+7tX+RhOpyDPvxG9+aVvwf3FIACIbIAAAAAAAAAMp4gAAZAAGAAAAAAAAABw5yNTKqic8c1OUAAAADjKZ6nIAAAAAAAAAHA4k9fzAcgAyAAMagAAAAAAAyAAMAAcZMjkAAAAYAAGQAAAAAAAAAAAAAAAAAAAjbeb0cmudiOpLNHF2lZHTLWUeE59tD9kaietyIrPvlJJBgaZwZ7vCaR+sfbNqbTscSx0sNa6akTHJIJcSRonjhrkT2opgRkAAAAAFidwHSf1e23fVyaLiptP0UlSiqnLtpPscae3DnuT5JsPKz/ALHhpb6k7Ia/UssfDNfrg5WOx6UEGY2/jrMWYAAAAAAKfb9FoWl2haeviYRlxtclKqInV8EnFlfdPj3FfS4O/NaPKtm9nvTIuKS23eNHv+JFMx8a/j9mU+Oucl7/AD2W29eNOseU/KYVTNaNnETPWBvJyL6wCwI1ezdJuLrhsD062SXtJaPt6N/P0ezne1qfgcBK5XPcRubJdE6ls2VWWkvCVPsZNCzH40byxhxjPLPM5jep/wCUz57/AIrrhK9uzTV2AAIlsAAMjAt4axt1DsU1bblRyvbbpKqJG9Vkg+zMRPa6NE95r4Y9JGNkb0e1HJ70ybPbhTsq6KelkTLJo3RuTHcqKi/lNXluVVoKfPVI0T5uX9R0TkRdmbN231TE+cT8kBnVG+mp+4ALwgVk9w64Rx6g1haXPXtKimo6qNvqY6Vj1/GYWwKN7nt0S3bdqGlx/hW21VH7FajZ0/RO+cvIcp5XWebzKav7oifh8Ftyuvaw0dgACsJAAAAineyvDrRsH1C2Kbsqi4Nit8X23bSNa9P6Pj+YlYrTv43dY7BpPT6NylXcJq5y+HYR8KJ886fMS2RWOfzGzR26+W/4NbF183Zqq7FU16+ruM53f7Kt+226SoV4kjir/L5FRM4SnasqZ9SvaxPeYKWD3GLMtVr7UN/cqKy3W2Okaip8OeTjVU9jYUT746pnOI+j4C7c7JjxndHvlV8Bb28RTC4CAA4suIcKdHUN5ten7LVXm9V0FBb6SNZJ55nYaxv9aquERE5qqoic1KWbctvF+19JPZrA6qsml1RWOY13BU1yLyVZXJ6DFT/Fp158SrlESWynJsRmdzS3upjjM8I+c9jWxOLt4enWryT7tP3idDaPmmt1ukk1Nd41Vrqe3vb2MTk7pJ181veio3icipzRCCdT7zO0y6yqlpW0aegR2WJT03lMuPBz5ctX2oxCFWsa1qNa1GtToiJhD6TkdGwXJnL8LEa0bdXXVv8Adw/fFXr+aX7k/VnSEhrtw2u9ssv1+VvEq9PI6Xh/B7LBkmmt5jadantS6PtGoYeJFelTS+Tyq3vRr4sNRfWrFIYBv3MowFynZqs0+UR743tenHYimddqV4tlW8HorW1TBaq1ZdO3uZUYykrXIscz1+DFMnmvXmmEXhcueSKTCnNMmrt7WyMcx7Uc13VHJlFLH7sW3KtoblR6G1rWyVVDUvSG1XKd/FJTyLybBK5fSY5eTXrlWqqNXKKitpud8k4s0TfwfCONPy+U7+1M4LNOcnYu7p61tQAUVMgAAFEd6jWq6w2s1lNTyK616f4rdSpnk6ZF/viRPa9EZnvSNF7y2G37XH1gbMbne4HtS5SolJbWr8KpkyjFwvJeFOJ6p3oxTXzExI42xtcrkRObl6uXvX3rzL5yMy7WqrF1xw3R8Z8t3jKFzjEbNMWo6XKg5ODoKuLgbml4s9s2R1Mdwu9vo5JL1VPayeqYx3D5qIuFXPcTV9dmlv8AOWzf7dF/xGtJ8MD38T4Inu71dGiqfPk1N/ktP/RN/wDQqGN5I28XiK783ZjanXTT9UzZzabVEUbPBsv+uvS/+clm/wBui/4jj669L8WPrls2V/8A30X/ABGtHyWm/wAlg/om/wDofTKWlWRuaWD0k/xTf/Q1f9EWv/NPl+r0/jU/2tn9BWUlfSR1lDVQVVNKmY5oZEex6eKOTkp+5GG6m1G7vmkUa1ETyV68kx/jnknlDxdmLF+u1E67MzHlOieoq2qYnrAAprvoI02q7bNEbPZH0NdWPuV5ROVsoESSViqnJZFyjY05ovnKi4XKIpGW81t3qLXVVeh9B1iMuEarFdLrGuVpV74IV6dr8Z3wOief6FVOfG97nOe+Ryve97lc57lXKucq81VfFS65JyVnE0RfxW6meEdM9s9Ue+exEY3M4szsW98pz1VvRbQbo97LDb7Pp6nXHAqsWrqG+1zuFn4hiD9uW158vau13WI7wbQ0rW/g9lgjsF1tZNgLVOzTZp8YiffOsoWrH4iqdZqS7Yt5DaxbJlfV3S13ti8uzrre2PHsdBwc/aik2bNt5vSN+njt+q6V+la168LZppklonryx9mwnAq8189qImPSUpsFRFRUVEVF6oqGrjOTeX4qnTm4pnrp3e7h7nrZzO/bnfOsdraDG5r2NexyOa5MoqLlFQ+ijm71tsr9nlfDYr/PPWaQldw8K5kktir8OPqqxfGjTpzc3nlrrvUdTT1lHDWUk8VRTTxtkhmiejmSMcmWuaqclRUVFRUOa5vk97LLuxXvpnhPX8p64WPC4qjEUbVL9QARLZAeXqy/2rS2na3UF8qkpLdRR9pPLwK7hTKJ0aiqq5VEwid5XLWu9fGiSU+idLSSu6NrLw/s2IueqQsVXOTHi5i+okcBlWLx/wBxRrEdPCPOXhexFuzGtc6LOVlTT0dLLVVc8VPTwsV8ssr0axjUTKqqryRETvIA2p7zunrM6W26Fpo9SXBuWrWOcrKGF3P4SedNhUTkzDVReTys2vtfaw15Uceqr9UV0CO4o6JidlSRqirjETeSqmVTidxL6zGS7ZbyOtWtK8XO1PVHDxnjPu8UNic4md1qPFkWuNdaw1rcWV+pNQ11VJFJ2lNFFIsENK9Fy10UbMI1ycsOXLuXNS/OyS51152WaUu9ymWetrLPSz1EqoiLJI6JqucuOXNVVTXIvor7DYfsI/gU0T9waP8AQtPPllYtWsLai3TEREzppGnQ+8ou13K65qnVmildN4TeEfpq61GktCeS1N2p1WOvuMre0ho3/wAkxvSSVOq581vRUVco3Od5jaI7Z9s7mkt07GX66OWjticlWNypl82F7o28+9OJWIvUoYxqNbjic7qqq5cqqquVVV8VU0OS+RW8VE4rERrTG6I65657I989z3zLHTZjYo4yz5u2bayyvWtbr66duq5w6GBYv6Lg4foJ73e94G7av1VRaL1XaYFuNYyVaa40PmRv7ONXqkkblXhXDXLxNXCrhOFOpUgkvdX/AIw+lPk1v6rIWnOMpwVeDuVc1ETTTMxMRpwjXoRmBxl6b1NM1bplfgBAciWkAMc2ja0sOgdLVGotQ1KxUsSoyONicUtRKueGKNvwnrheXRERVVURFVPu1bru1xRRGszwhiZimNZelqO92nTtmqbxfLjT2+30zeOaed/C1qd3tVV5Iic1VURMqVl2k71FVLLLQ7PLPGyFFVqXS6MXz05pxRwIqL4KivVPW0hfavtG1HtKv/1SvknYUcL1WhtkT1WGkb0Rft5MdXqnsREwiYh/WdIynklZs0xcxcbVXV0R8/Tv4q9i82qmdm1w62XXvaltNvM3bV+v9Qtd4UdT5Gz8GFGodWj2hbQqSZssGv8AVaPb047rLI1fa16qi/MY0C004PD007MW6dO6EXOJuzOs1SmnRu8ttHssyMvS2/U9Jnzm1ESUs6Jjo2SNOH8Ji+0stsg2v6T2lQyQWuWWhu9PGj6i2VeGzNb0V7MKqSMzy4mry5ZRuUKAHZtNxuNnu1JeLPWyUNyopUmpaiP0o3p9CtVOStXkqKqLyUhMy5MYPF0TNqmKK+iY4eMcPLe3sNml23VEXJ1hs6PE1dq3TOkaBK7U18oLTAueBamZGrIqdUY30nr6moqlR9f7y+t77SQUOm46fTkS00SVdQyNJKiSbh+y9mrstjjyuG8lf5qLlM4SFK+pqrhcZLlcqupr6+VcyVVXM6aZ64xze5VXoVvAcjL1zSrE1bMdUb5+Ue9I383t0bre+WwrZhtL01tGS6yaYdWTU1tnZC+ongWJkyubxIrEd52PlIimaFatw5E+t/Vy963GD9ChZUrucYS3g8bXYt/Zp0490N/C3Zu2aa6uMgAIxsAAAEb7XNsuj9nLVpK+ofcb05vFHa6NUdN0yjpFXlE3mnN3NU6I7Bh28zttdopHaS0nLFJqaaJHVFQqI9ttjcnmuVOiyuTm1q8kTDlTCojqcTSzT1E1VU1E1TU1EiyzzzSK+SV6rlXPcvNyqvepcch5MTjKYv4ndR0R0z8o98+9FY7MosfUo31eiY9Xbyu0y9TuSzyW7TVLxLwMpYEqJ8eDpJUVqr62saYRNtS2mSy9q7aBqNHZz5tXwp8yIifQYgoL7ZyvBWKdmi1THhE++d6AuYy9cnWqpKOnN4DaxZZmOdqOC8QtTHYXOjje1fv40Y/6Sedl28xpXUVRDa9WU31r3KRUayWSXtKKV3LpLhOBV58noifbKpTUKiORUVEVFTCoqclNLG8ncBi6d9EUz107vdwnxe9jMr9qd86w2hpzTrkFK93PbhW6LraXTGq6ySq0rIrYYKiVyufbF6N5r1h7lT4Cc05IqF02qjmo5qoqKmUVO85lmuU3ssvc3c3xPCeif164WXDYmjEUbVLkAEW2AAGQAAAAAAAAAAAAAAAAABgUj/ZItJpT6g01rWCLDayB9uqnInw41441X1q1709jCoZsz3ztLfXTu+X9I4+OptKMukPLOOxz2i/0TpDWYZAAADlrXOcjWornKuERE5qpwSDu46b+uzbjpKyvj7SF9xZPO3HJYocyvRfUrWKnvA2XbJNNN0fsx05plGI19vt0MU2O+XhRZF971cvvMpAAAAwAAMiPt46zLfdh2raFM8bLc+rZhMqroFSZqJ7VjRPea/GvR7UkRco5EcnvNn9ZTxVVJNSzsR8UzHRyNXo5qphU+ZTWLV0E9prKm01WPKKColpJcfGie5i/mnQ+RN/W1ds9UxPnun0hAZ1R9mvwfmAC8oFYTcVua0+0HUdlwnDX2qKrz64JVZj5py35QndZubrZt705mRGRVramilVV68ULntT3vjaX2OWcsLOxmG3/AHUxPw+C15VXtYeI6gAFVSQADI6l6robZaK25VC4hpKeSeRc9Gsarl+hDWBQJihg9caO+fn/AFmwbeVuy2XYVq6qT0pbe6jb7Z1SFP0hQDhRnmN9FvJPYdF5E2dLF271zEeUfqr+dV76aQAF3QTLti9zfZ9sWjq9iomLxDTvVe5s2YXfRIbFU6Grt9RNSNbV068M1O9s7F8HMcjk/IbPqGoirKKCrgcjop42yRuTva5MovzKc85b2dLlm71xMeWk/FY8lr1t1U9T9gAUZNAAMAUq31Lulw2xwW2OZXMtNpijezPJssr3SO/ESIuqprq213d192xavujlRyPustPG5OisgxA1fmjLfyMsbeNquT/TT750j01RWb17NjTrliXUuVuRWfyLZNVXh7G8V3us8zHonNYo+GFqfhRvX3lMpZOyhklVPQarvmTJsY2N6f8ArX2V6ZsToOxmpbbClQzwmc3il/Hc5Sf5ZYjm8FTa6ap90fro0Mmt63Jq6oZafFRNFBBJPPKyKKNqvke9yNa1qJlVVV6IiH2Qtvm3Kvt+xaWGhqn07K+409HVcC4WSF6qrmZ6oi8KIuOqZToqnO8FhpxeIosxOm1MQsF25zdE1dSvG8TtYqNpOpXUdumkbpW3y/3jFhWpVyJlPKXp158+BF6NXOEVVIs7+Zz7sHB2vCYW1hLNNm1GlMfvzUy/eqvVzXU5OG+dPHAxFfNIuGRMar3vX1NTKqZ1sZ2Y3vabqB1HQPWjtVK5PqjcXNykSLzSNiL6Uip3dETmvci3e2c7PNJaBtbaLTdqjgerUSark8+pqF71fIvNefPCYanciENnHKLD5bPNxG1X1R0d8/BuYPLa8RG1O6GvG5W25WyBs9ztVyoIXrhslVRSwsVfDic1E+k6uUXmnNF6KbP54Yp4XwTxMlie1WvY9vE1yL1RUXqhVbed2F2+z2qq1zoai8lgp8y3S1QM+xNj6unhanocPVzE83hyqY4Vzo5XyttYu7Fm9RsTPCddY7uEaPfEZRVbp2qJ1Vq7z5kYyRjo5GorXJhUOUVFRFRUVF5oqd5z3lvQ3BdzdO2iza20G+03eqdPfrCrKepke5VfUQqi9jMqr1cqIrXc1VXMVV9JCZjX1u+6v+sra3ZrpLIjKGskS216qqI1IZlRGuVV6IyRGO9iL4mwVOhyXlPlsYLGTNEaU174+MefumFvy/E8/ZiZ4wAGKbW9Y0+g9nt31ROztX0kH97Q4Ve1ncvDEzCc8K9UyqdEyvcQFm1VduRbojWZnSPFu1VRTGsqq74utk1HtIj01RzI+3acarJFauWvrJERZOi4Xgbws8Ucr0ISPqaWrqZ5qqrdWVNVPI6aomdBI5ZZXuVz3KvDzVXKqnyqK1zmua5rmqrXNc1UVqp1RUXminbcBhKcHhqLFP8ATHv6Z8ZUzFXar12a5cABcImVXBuNYB9rDUouFo63wx5JL/wjsanGUoq7/ZJf+ExrD62Kup8H1H+2N9qH12FVn9w13+xy/wDCfcdNV9q3+8q/GU/+Cl/4TG1BsVdS926r/F90h/8A4j/0ryTiM91qOSPYDpJkrHse2keite1Wqn2V/cpJhxLNN+Nvfmq9ZXez93T3QEJb1e1aTQ+nmacsM/BqW7xOWOVrkRaGnzh03jxKuWs9eVz5uFlvVF7t+m9O3C/3WbsaG3076id/ejWplURO9V6Ineqohrl1pqS5ax1ZctUXflWXGbtFjzygjRMRxJyTkxqIniq5VeaqTfJfKIxuIm7dj6lHvnoj4z4R0tLMsXzFvSnjLxo2MjYjGJhqH0DKNl+hr1tE1fDp2yqyHze2rKyRvFHRw5wr1TPnOXo1vevVURFVOoXb1FmiblydIjfMqvRRVdr2ad8yxiNr5aiKmhZJNUTO4YoYmK+SRfBrWoqqZI3Z9tCWFZk2f6s4E/1RNn8Hh4voL27Mdm2k9nlr8l09bmtqXtRKmvmRH1NSvi9+OnfwphqdyIZiUXFcttK5ixb1p65nj4Rw805byWnZ+vVvavZmvgqZaWeKSGoiXEsMrFZIxfBzXYVPehwbDNrezHTG0iyOo7zSthr4mr5FcoWJ5RSu7sL8JnixeS+3CpQrWmmbxo3VVfpm/QtjrqJ6Irmc45mLzZKxe9rk5+KLlFwqKhYclz6zmlMxEbNccY+Mfvc0Mbl9WG+tE6w8gs1uX7SXxVK7M7xOro3NfPZHu6twnFLT58ETMjfVxpnk1Csp+9urq613Okutrn7C4UM7KmlkxnhlYvE1cd6Z5KneiqbuaZfRmGGqs1cZ4T1T0T8+x4YPETYuxV0dLZ4Dwdnmp6LWeibRqi3pwwXGmbNwZysT+j41Xxa5HNX1op7xxW5bqt1zRVGkxulcomJjWHWutBR3W21NtuNNHU0dVE6GeGRMtkY5MOaqeCopTreB2AyaItlRqrSM9TXWCHzqyjm8+ehZn02u6yRJnnnzmomcuTiVLnH4V9LT1tFPR1kDJ6aeN0U0T0y17HJhzVTvRUVUJLKs3v5bdiq3P1Z4x0T+vVLXxOGoxFOlUNYaLlEUHqavsf1savvem8yObarhPSRuk9J0bXr2bl9rFap5Z2aiumumK6eE71Nrp2KppnoHegvsU2IbCuWxXRP3Aov0DDXe/wDa19ilzrlrN2htzyxXemm7K4zadoaK3Ydh3lEsDWscnranE/HgxSpcrbFeIt2bVHGqrTzhM5PVFG3VPUr9vM61drXaxXugm7S1WVXW2gRHea5Wr9mkTnjznoqZTq1jSMj5ijbFE2Nq5RqYz4n0WfCYajC2abNHCmNP33oq/dm7cmuekJM3Vsf3Q2lfkVv6q8jMkvdY/jEaT9bK79VeeGa+w3vyVf4y9MD7RR3r8J0AQHEVzdS83Kis9pq7rcqllLRUcL56iZ/oxxtRVc5ceCIpr92z7RrhtN1k+9VDZqe103FFaaKTksESrze5E5do/CK7rhMNyqIS9vqbQ3VdfFs1tc32CBI6q9K3PnvXDoYF9SJiRyfzfgpWvvOlck8niza+l3Y+tVw7I6/H075V7NsZMzzNPidADKNmOhL/ALRNUx2GwxtZwoklbWypmKjizjjd8Zy9GsTm5fBEVUt929RZom5cnSmOMoa3bquVRTTG9isskcSIssjWZ6ZXmvsTvOxJTVcdN5XLQ1zKbGe2fSStjx48StwX+2WbItGbPaSJ1st0dXdkbia61bEfUyL34d/i2/atwnJM5XmZ8qJgpWI5bW6a9LVqZjrmdPdpKbt5LrT9ere1eMc17EexzXNXoqLlDkuft13f7HqqhqL1pCiprPqRjXP7OFEjp69evDI1OTXrzxImFyvnZTpTOaKenqJqarp5aapgkdFPDK1WvikauHMci9FRUVFLLlWb2MztzVa3THGJ4x+najcXgq8NO/g+AcnBKNNa7cN/e9q37ow/oULKoVq3Dv3v6u+6MP6FCypx7lJ+J3fD0hccv9mo7gAEG3Aj/b1tFp9m2g57s1IprtUu8mtdM9eUs6pycqJz4GJlzunJMZRXIZ+UE3jNfO2gbS6uppZ0kstqV9Da0a7LHoi4lmTmqLxuTkqfBawnuT2V/wAQxcRXH1Kd8/CPH01aWOxP0e1M9M8Ee1lTVVtbUV1dUyVVZVSunqZ5F86WRy5c5fWqqfkDtWi23C8Xajs9pplqrhXztp6WFFxxSOXCZXuROaqq8kRFVTrutNFO/dEKlEVXKuuZfdhtF21BeYLNYrbU3K4z846enbxOx0Vyr0a1M83OVETxJdpd2DalPQsqZJ9M0krm8S0stbK6Rq/FVzYlbn2Kqess/sZ2aWTZrpdluoGNqLlOjX3K4Ob9kqpUT8ViZVGs6Iniqqq50c8zDlle52acJERTHTMb5+Ue/wBFhw+UW4p1ub5a1tZ6V1Joy8JaNU2ia2Vb0V0XGqOjmanV0cjfNeics96d6IeKbINoui7Br3S9Rp7UNJ21NL50cjMJLTyJ6Msbsea9M9ei80VFRVRdfu0LSN30JrGu0ve0zUUy8cM6M4WVUDlXgmanPkuFRUyuHI5O4seQ59RmdM0Vxs3I6OiY64+MI7H5fOH+tT9l4Kp48/6y1W5rtPfVU6bNb5VOfPTRLJZJpX5WSBqedT8+arGnNvXzMpyRnOqp2bTca+zXajvFoqFp7jQTtqaWT4sjVymfFF6KneiqhIZtltGY4aqzVx6J6p6P17GvgsTOHuxV0dLZ0DHdm+q6HW+h7Tqm3tVkNfAj3RKvOKRFVska+tr0c3PfgyI4vct1Wq5orjSYnSe+FxpmKo1gAB8sgAAAAAAAAAAAAAAAAAMDrXSiprnbKq21kfaU1VC+CZnxmOarXJ8yqahNW2ao05qq7afq8+UW2tmpJMpjLo3q1V+g3CGtvfn039b+8HdKmOPggvNNDcI0xyyrezf8743L7xAgsAGQLR/scmnvL9qN81HIzijtNsSJi49GWd+EX8CORPeVcL+/sdVg+p+yG6X6RmJbtdXI1cdYoWNa38dZALNgAwAAAAAyCmvjeEtLbLtx1fRMR3BLXJWtXx8ojbK7H3znJ7jYOU434bQ6k2n2a8tY1sVytCxZTq6SCVc5+9mb8xa+R1/m8fNE/wBVM+caT6RKMzajaw+vUgQ4AOoqq9PSN0ZYtZWC+SriO3Xakqn88ZayVvF+KqmzA1cVcay0k0aJlXRuRPbjkbKdnN3fqDZ/p2+SuR0lwtdNUyKnxnxNcv0qpQuW9ndZux2x6THxWHJa/q1UveABQE4AACvu/NdkptnFmszJeGW43iNzmZ9OKFj3u+Z/ZlPu8sBvy3davaNYbGiosdttb6lVRej55OHC/ewov3xX9DrnJixzOW2+urWfOfloqma3NrETHUAAsCNHNSRjol6PRWr70wbCd366tvWxPR9c13Ev1Jggeuc5fE3sn/jMU17p4l1Nyq5rW7FkoFRE+pV0qqVqepzknT9MVDlnZ2sFTc/tqj3xP6JnJq9LtVPXCbgAcyWQABgebqi7Q2HTd0vlQirBbqOWrkRO9sbFev0Iay6d0j4GSSuV0sidpIq9Vc7zlVfeql8d7C7utOwjUTYpUjnr2RUEf23bSNY5PwFf8xRJcZ5JhDpHImxs4e5d/umI8o/VXs6r1qpoevoezJqLXGntPvY58dxulPTyo1OkSyIsi+5iOU2VoUc3PbN9VduNHWOVUZZ7fUVq8sornIkDU/8ANev3peMieWmI28VRaj+mPfM/KIbeT29mzNXXIQXvvfwOU33bpP7ZOhBW+/8AwOU33bpP7ZB5F+I2fzQ3cZ9zUpkejpWw3TVOp7dpuyxo+4XGdIYeL0WJjLpHd/CxqK5fUh52PAthuRaHbSWGu2hVsaLUXRXUluz8CmY7D3pz/wAZI3vTpG3HU6nm+Yxl+EqvdPCO+eHz8FXwOG+kXYp6OlOGznR1n0JpCi03ZIuGnpm5kld6dRKvpyvXvc5efq5ImERETIgDjNy5XdrmuudZnfMrhERTGkB8SxsljdHIxr2PRWua5MoqL1RU7z7B8Mtc217SaaH2m33TETVSkpp0loV5/uaVOONMrzXhRVYq+LFMW7yxO/baY6fWelr61XcdfQVFFImOSdi9sjF/85/zFdl6na8nxc4vA2r1XGY398bp98KbjrUWr9VMPiaPtoXxdONqpnwXuNjGxzUi6u2Wab1FJL2s9Zb4lqX4xmdqcEvL5bXGurvyXP3JbpLXbGpaCTCNtV4qqWNPtHcM/wCWZSC5Z4eK8HTd6aavdP66N/Jbkxcqp605nCocg5isYa5dsCqu1/Wqr/2/VJ+MhsaNcu17+F3Wv3fq/wA8u3In2i7+WPVDZz93SxY+ZFxwL4SxfpGn0vU+X9Wfz0f57TpCu0fahtFABwFewAAAB3AVo35NYLT2i0aCpJkR9xf5dcGIvPsI3fY2qng6Tn/8oqnle/qZhtr1Qustq+or9HP21ItWtJRK1/EzyeH7GxW+COVHP9r1MPOz5HgfoWBotTG/jPfPy4eCoZhf56/M9EPmWRsUbpHrhrUyuPBC+W7Hs/TQmzOkWtp0jvt2a2tuaq3DmOcnmQr3okbVRuPjca95UTYXpVus9rWn7HPF2tEk61tcis4mrBAnGrXJ4OdwM+/NhyFZ5aY+YijCUzx+tPw+M+SSyaxGk3Z7gAHPk6Fc9+DRsVdpKg11TRNSqs0raeseiIivpZXI1Mr1XgkVqoncj3qWMPI1pY6fU2kbvp2qVGw3KilpXOVueDjYreJE8Uzn3G/leMnBYui9HRO/u4T7njiLUXbc0T0taaZ7+oHZzwqsFVG6OoiVYpmOTm2Rqq1yL70UHboUmY0nRancY1Ustsv2iaidFfRypcaJiuyvZS+bKiJ3NbIiL7ZSzSGv/dt1C7TW2/TlQr3Np7hK611KInpNnTEaL6u1bEvuNgCdDlPKzB/R8fNcRurjXx4T6a+K2ZZd5zDxE9AACsJBQLecjbFvB6wYxqI1ZaR+PW6khVfpI5JJ3o/4xGr/AG0X6pERt3nb8rnXA2fyU/4wpmO9or73EmeB2PBSV9uepX1WiNluj4JVWC2aXorhUoj+SzSwNbGjk7laxr1T1SEUqmUVO5eR+1fV1NdU+U1UiyS9lFEi4xwsjjbHG1PUjWtRPYet7DRdvW7k/wBGs+Mxp83zav8AN26qY6X4oc4XGURVM02ObN7ztM1Q61W6TyOhpWtkuNe5nElOxV81rU+FI7C4ReXJVXoWxt+7hsiprdHSz6alr5WtRH1VTXzrNIqfCVWvajV+SiJ6iMzLlBg8uuRbuazV1Rv079ZhsYbLruIp2o3QoySVusfxiNJ+HDW/qshkm8DsHqdBUL9TaZqam5adYqeVxVCo6ooUVeT+JETtIs8lVfObyzlMqmObrSY3iNJLnOW1v6rIfWIx1jHZZeu2KtY2KvD6s7p7X1Zw9djF0U1x0r8J0PA2h6notGaJu+qLhzgt1M6bgzhZX9GRovcrnK1qetT3yr+/TqzhhsWhaaTnM5bpXNTKL2bFVkLV7lRX8bsf6NDl2UYH6djKLPRM7+6N8rJibvM2prVkuVfW3W51l2ucvbV9fUPqqqTGOKV7uJ3LuTK4RO5EQ64CnaqYimNI4KXVVNU6y7FtoK663SjtVrplqq+unZTUsKLjtJHrhqZ6Ineqr0RFNg2xnZ7bNm+i6eyUXDPWSYmuNZjDqmdU853qanRre5ETquVWAdyTQra27V+0K4QZjoVdQWriT/Gqn2eVPY1UjRftpE7i2Zzflfms3b30O3P1aePbP6evcsmVYWLdvnKuMgAKWlwprvraRgsm0C26ooo2xwaggeypa1q48qh4fPXuTjjc1MJ/JqveXKIJ33rd5Xsep65qJm3XmmnV2OjXccS/pEJ7k1iarGY29J3VfVnx4e/Rp4+1FyxVCmZwAdfU5a3cN/e/q77owfoULLFadw397+rvujB+hQssce5S/id3w9IXHAezUdwACDbiK96TWkmjNklwdRTOiut2cltoXNyjmPkReORFRUVFbGj3Ivc5G+JQ6NjY42xsTDWphvsJu3y9VLfdrEdhhkVaTTlKkWMJhamZEfIqKnXDOybz6LxEJqdb5L4H6LgKap+1X9afh7t/iqua3+cvbMcICyW4/ouOsut21/WxNe2hctutuUReGRWo6eROXJeFzWIqdznoVrmkSGGSVUyjGq7Hj6jYlsQ0uujtlGnbBJGrKmCjbJVovXyiTMkv47nJ7EQ8OVuNnD4Lm6Z31zp4dPwjxfeUWIruzXPQzMAHK1nCDd8XQrdR7Ol1RRQIt104jqlVannS0n+PZ17kRJE6+gqJ6ROR+dTDFUU8lPPEyWKVqskY9uWuaqYVFTvRUNvA4uvB4ii/Rxpn/wBx4xued23FyiaJ6Wr9FRURUcjkXmip3nJ7GuNOyaQ1pe9LSdpi11slPE6T0nw54onr7Y3MU8c7fbrpuUxXTwnfHcpNyiaKppnoWQ3HtYupL5eNCVUqJBWtW5UCOVExK1GtnaneuW8DkTu4Xr3lszWlorUU+kNZWbVdPxqtrrWTyNZydJD6MrEX7aNXJ7zZTTyxT08c8EjZYpGo9j2rlHNVMoqL3oqHM+WGBizi4v08K498cfdos2VXucs7M8YfoACpJQAAAAAAAAAAAAAAAAABgCnn7JTp7jtekdVxsx2U81umfjrxtSSNPdwS/OXDIX31rB9Xt3bUCsZxTW1Ya+Ll07OREev9G54Gs8AGQNqO7HYvrd2BaNtys4HutrKqRuOaOnVZlz68yGrqx2+a7XqhtVP+3VtTHTx8vhPcjU+lTcNQUsNDQ09FTN4IKeJsUbfBrURET5kA/YAGAAAAAACue/baWz6J05fUY50lDdvJ1VE9GOeJyKq/fRxp7yxhGu9BaZLzsH1XTwo3taakSuaqp08ne2ZcevhYqe8lMlv8xmFmvtiPCd0+rXxVHOWaqexQZPWBlF5t6LzT2BTtKlCclyXm3P7mlw2D2anWZZZrdNU0UuV9HgmerG+5jmFGS1u4dde00/qyw8OPJbhDWovj28XAv0wfSVjldZ5zLpq/tmJ+HxS2T16X5p64WWABypZwKDq3Wtp7bbKq41ciR09LC+eZ/wAVjGq5y/MiiImZ0hiZ0jVQLeEu6Xzbhq6uY5yxxVyULEXonk8bYnY9XGj195grGPe9kcbFe+R7WMa3q5zlRERPeqCSrqLhNLcat/HU1kr6qZ3i+RyvX6XGX7EbM7UG2HSVp5cDrmyqlRyZRY6dFnci+3s8e87hTFOCwkRPCin/ABj9FMr1v4ie2WLXKiqbbc6y2VrOzqqKokpqhnxZI3K1yfOinXM53gLW2z7cNY0TVyj7j5Yi4/l42TL9L3J7jBj1w17n7NF3+6InzjV53qObuTT1BZ/cNubuPWNlkm8xHUlbDHnvc18ci/8AlxlYCaNzG5x0G23ySRzk+qdnqKdiIvJXsfHKmfvWyEZyis89lt2OqNfKYn4NrLK9nE09u5dsBAcdW4AAFa9/C7JHpvStgxlau4y1qrnokEStx887fmKoJ4E276V4S47ZIrZHM50dotMUT4+5ksr3SOX3sSL5iFGomUzyTvOwcm7HMZbaieMxr5zrHu0VLM7m3iKuxaXcPs7m27Vmo3o1Wz1UFviXHNvZMWR/uVZm/glnCKd0yyrZ9hNgfJD2VRckluM323bSK5i/0fZ/MSsc1z7Ec/mN2vt08t3wWTCW+bsU09gQTvv/AMDtN926T+2TsQVvv/wOU33bpP7ZjIvxGz+aDGfcVKe2S01l/vlusNvVG1dzqoqOFyouGukcjeJfU1FVy+pDZRp200VhsNvsltjWKit9NHTU7FXKoxjUa3K964TqUy3NNPpeNsa3aSNzobFb5Khrk9FJ5fsTEX7xZV9xdxOhPcs8ZNeIow8cKY1nvn9NPNo5RZ2bU1z0gAKWlwAAVz376ON2hdNXJUzJBeuwRftZKeVV+mNpUguTvyxo/ZDbnr1jv9K5PeyVPyKpTZDq/JKrXLaY6pn5qvm8fz9ewLUbhda99r1lbVXzIaymqET1yRK1f0SFVyzO4Nn6o67TPLgtq/rJ7cqKYnK7s9Wn+UPjKfaY7pWqAByJaw1y7Xv4Xda+P1fq/wA82NGuXa9/C9rX7v1f55duRPtF38seqGzn7uli58P9Jn87F+kafR8v6s/nov0jTo6u0fahtFABwJewAADC9uWpn6P2Sak1BDI+OpgonR0rmdWzyKkcS+572r7jNCuW/Xekp9Had0417kkuNydUuRF5LHTsXKL9/JGvuJLJ8LGKx1q1PCZ390b590PDE3ObtVVdUKjwxpDEyJOjE4fbjvPoHKIiqiLy58zteqlTOu+VmdxLTyPrdUatliTzEitdNJn1drMn0w/MWpIm3SLH9RdhdklkhSKourpblN9t2r1WNf6JI/mJZONZ/ifpGY3auiJ08t3wXLB2+bsU0gAIdtBwue45Cga8tvVobYttmsLc1/E11yWsavTCVDGz49iK9ye4wkm/fYtzaPbJSVsbEa24WWJ73fGkjlkYv4qsQhA7ZlF7nsDarnppjzjdKm46jYv1RD5fPUUvDWUj3R1FM9s8L29WvYqOaqe9ENm2nbnT3uwW680i5p6+liqovkyMRyfQprMREXkvReSl8t1a6yXbYJpeSZ2ZaWCSicmeiQSvian4LWla5bWNrD2r3VMx5xr8ElktzfVQlAAKc4WBQXek/jEav/8A4X6pERuSRvSfxiNX+2i/VIiNjt2Vew2PyU/4wpuO9pr7w5ayaWRkNNBJPPK9scMUbcuke5URrUTvVVVE95wS/ujaSZqba7DcaqJJKHT0Pl70VuWrUOVWwIvgqLxyJ640PbG4qnCYeu/VwpjX5R4zueeGszeuxRHStdsS0JT7PNnlBYGcElaqeUXGdvPtql6JxuRcJyTCNby9FrTNgDiF+/XfuVXbk6zM6yulFMUUxTHCH5VdPBV0s1LVQRz08zFjlikajmva5MK1UXkqKi4VCn+g9Az7Od8ewWH7I62P8tqbVK/nx076WXDM97mKisXPNcIvLiQuKdSqtdtq7hRXGqoKWetoVetJUSRNdJBxt4X8DlTLeJOS46p1N/LszrwdF23xpuUzTPfMTET4a+TyvWIuTTV0xOrtdxrs2y6nTWO1XUWoI5UkpZKtaaiVrlVq08P2Njm+CO4XP9r1LwbcdSO0jsk1Lf4pHx1EFC9lM9vVs8mI4l/De013RxpDGyFvRjUb7cIWrkThPvMTP5Y9Z+CKzq7pTTbjp3vocMz8R00TpZ5HJHDG1Mq97lRrURO/KqgJI3ZdOpqTbjp+CWLtaW2q+61CcWOHsURIl9f2Z8XIvGKxFOGsV3quFMTPkhMPa525TR1rr7L9K0+idn9l0vTcKpb6Vscr25xJKvnSv5/Gerne8yQdwOG3blV2ua651mZ1nxXamIpjSAAHmyER74CZ3fdQu+LLRL/9ZCS4RJvgfxedS/Lov1yAkcn/ABCx+en1h5X/ALuruUYVMKpwcrniXPicHbFHWu3DP3vau+6UP6FCypWncN/wBq5O76owfoULLHHuUv4nd8PSFxwHs1HcHTvdxpbPZ627V0nZ0lFTvqJ3/FYxqucvzIp3CHt8G/LZdiFxpI5HRz3qohtkbmrzw93FInsWJkie8jcDhpxWJosx/VMQ2LtexRNXUpLdLnVXu7V18r+Hyy5VMtbPjoj5Xq9U92UT3H4BVyq+04O5U0xTERCkVVTVMzLJNllgTVO03TOnnMSSKruUSzsX4UMf2WX8RjjY+hSvcosyXDbDV3WWHjjtNpkex+PQmme1ifOxsqF1DmfLPEbeMptRwpj3zv8ATRZsot7Nja6wAFPSoAAKX77VgS2bV7ffYoUZFe7aiPdnm+endwuX+jfEnuIMLeb9VoZU6AsN8bErprfd2wucnwYpo3Nd+O2IqGnQ6/yZxE38tt68adY8p3e7RU80t7GImY6RURUVHc2qmFTxQvduo6jdqLYjZUmlSSqtSPtdRhOiwrhn/lLGvvKIll9xC+8Fz1VpeWZcSMhudNH7MxTL+hNflZhuey+a440TE/CfXXwemUXdm9s9a1gAOUrQAAAAAAAAAAAAAAAAAAwB5GtLPHqLR16sEqIrLlQT0js9MSRuZ/WeuANNU0b4ZXxStVkjHK1zV6oqdUPgznb9Zfre22axtLWcEcV3qHxN8I3vV7E/BchgxkSXutWf6ubwmi6JWcaR3JtWqY7oGrN/9s2mmvL9j3tXl+3qSuVuW220VE6O8HOcyJPokd9JsNAAAwAAAAAyB0r7b4LvZa61VKZgrKeSnlTxa9qtX6FO6cL0M01TTOsMTETGktXjIZqZvktQxWTU6rBI1eqOYqsVPnap9GZbc7Stk2z6wt2Ux9VX1TUROSNqGtnRPd2ioYYd2w92L1qm7H9URPnGqkX6JouVUz1neTluTXPyPa9XW58vBFcrNJwt+PLFKxzfmY6Qg0zjYBdW2bbdo6uf6L7j5Gv/APIjfCn4z2/Maeb2efwN2j/jPnEax73tga9i/TLYWAgOKLkEa7z94fZNhGqqiLh7WppEoGIvf5Q9sK49jXuX3ElFcd+67pDo/TVgaqpJW3R1WuF5KyCNcov30sa+4lMkw/0jMLVHbE+Eb590NfFXObs1VdipSNRvmp0byT2IT3uQWV1dtOvF8cxrobTakhaq9Wy1EnJU+8henvIERC4e47Y/I9m10v8AJDwS3e6PSN/x4YWpG35npL850jlPiOZy2511aR5zv92quZXb28REz0Ir317alFtkpa6OHgZcbLE97/jyRyyMX5mrGQgWo38rU99s0hfmonBBV1FDJ4qssaPb+gX5yq6Z7z05OXuey21PVGnlOno+czo2cRV2hmewm6LZdtWja9ETndWUi+yoa6Bf0iGGH02qnt8kVxpXrHUUcrKmJyfBdG5HovztJfEWovWqrU/1RMecaNSxXsXKau1tBToD8qKoiq6OGqgcj4po2yMcnRWqmUX5lP1OEaabpXiJ1Dhehyebqm7Q2HTV0vlQmYbdRzVcieLY2K9foQzTTNVUUxxliZ0jVr82z3h1/wBr+r7q5Wua+6y08bmrydHBiBi+9I8+8xFYJ6vho6VjpKipc2CJqdXPeqMaie9yH50rpHU8ckrldI9vG9y9Vc7zlX51Uz3d/s31e23aRoPORsNelfI5EzhKdrpUz6lc1ie87hVNOCwsz0UU/wCMfopsa3sR3yv5p62U1ksNvs1GnDTUFLHSwp4MjajW/QiHeAOHVVTVMzPFcojSNAgnff8A4HaVPG+Un9snYgnffXGx+j5IqfV2kzn2PJXIvxGz+aGvjPuKnkbiNnZT6L1Hf1YqS190SmRV6LHBE3Cp99LInuLGkWbp9uktuwHTDZmcM1VFLWP5Y4u1me9q/guaSmYzu/z+YXq/+Ux5bvg+sLRsWaaewABFPcAAEFb7yZ2O0vqvlJ/bKZdS5u+6n/6O033cpP7ZTFDqnJD8O/7T8FYzf78LM7g/+ENdr4stv+8lZizO4P8A4Q138i2/7ybfKf8ACrv/AF/ypeeU+0x4rVAA5AtYa5Nrv8Lutfu/V/nmxs1y7Xsf9Lutcf8Ab9X+eXbkR7Rd7o9UNnP3dLF16ny/qz+ei/SNPrvPzm5NaqdUljX8dp0dXaPtQ2jAA4EvYAABSjfQvSXLbLHbI5VVlntUUL2dzZZXOkcvvZ2Rddehrp20XV982v6vub1RyPu01PG5FzmODEDF+aMt/Iyxt42q5P8ATT750j01RWb3Nmxp1yxI/KqVyU03Aiq9WK1iInNVXkn0qfsevoWiZctf6YtsreKOsvdFA9PFrp2Z+hFOl11xbpmuejerdqnariGxjSdoisGl7TYoHK6K3UUNIxV72xsRifkPTCA4PVVNdU1TxleIjSNAAHyyKAAKqb+tMjbloqsRvpsr4HL7Oxcn5HFZ0La799KjtHaXr1RMxXh0Of5ynkX+wVKQ65yWr2sstx1ax75VXNo0xEnsLhbjFas2zO90D38TqS+yq1vxWSRRPT6VcU+LObhNQ7t9b0au81HUE7U+U2Zqr+Ih88q7e3llc9UxPviPiZTVpiNOyVpgoByValBd6NP/AHiNX+2i/VIiNiSd6P8AjEav9tF+qREbHbsq9hsfkp/xhTcd7TX3neXG3ILCyh2X12oHxNSa9XKRzJE6ugh+xMT3PbKv3ylOVdwor16NRV+ZDYNu8W2G1bDtG00LeFH2iCpd8uZvauX8J6kDyyvzRgYtx/VVHlG/10buTW9bs1dUM9QAHL1lAABXXfovnkuibDp2N7mvudyWokRF5Oip2K5UX798S+4qN7ic99m7+XbXaG1slR0VrtDEcz4ks0jnO/EZH85Bq9Tr/JrD8xltvrq1nznd7tFTzS5t4iY6nBZ3cRsXFPqvVEkSLhYLbTyd6YRZZW/O+L5isbUyqJ4qXg3ObUy3bCbXVpEsct1qamulz3q6VzGL+Axhr8rMRzWXVUx/VMR8fg9Mot7V/a6oTEADlC0AAAES733Pd51Kn29F+uQEtES73y/+7zqX5dF+uQkjlHt9j89PrDyv/d1dyjDvSd7Tg5d6S+04O2KPC1u4b+9/Vy/6xh/QoWVK17hv73dW/dKH9ChZQ49yk/E7vfHpC44D2ajuCqO/fekkvWlNNxyu+wQ1Fwnj7vOVsUS/pS1xRPe1u0l0283iF3CsdrpaWhiVF6p2fbO/GmVPcbfJGxzmYxVP9MTPw+LyzSvYw89qKEByctTLkT1nVlTWt3DrSkem9V37vqrlFRInqgiR35Z1+YsoQ7ucUEdJsGtNU1nC+41VXVycuqrO9jV/BY0mI4zn17ncxvVdunlu+C6YOjYsU09gACIbIAAI13n7U+8bBtWU8WEkp6NK1q+HYPbMv0RqUG5dWrlF5obMNVWyO9aZulml/a6+jmpn+x7Fav5TWTQO4qGBy9ezRF9qJj+o6PyJu64e7b6pifOP0V/OqPrU1P2JO3V7w6zbeLBmRscFyZUW+ZXd6PjV7ET2yRsT3kYno6WujbFq2xX16LwW26UtW71tZM1VT5slsx1j6Rhrlr+6JjzhFYWvm71NXa2ZAJ0BwxdQAGQAAAAAAAAAAAAAAAAABga5d/mz/UzeFq6xGcKXW3U1X7cNWFf0JAJb39krtXZ6g0bfGt/dFLU0j3eHZvY9qf8Amu+kqEZFw/2NK2cd01reXN/aoKSlYvjxulc5PxG/OXVKufsb9u7HZTqC5q3Dqq9rEi+LY4Y1T6ZFLRgAAAABgAAZAAAUt317SlDtgpLmyPhjulojc52PSlikex34royDi2G/haEl01pW/ovOkuMlEqInVJ4ldn54G/OVP5952Dk3f57LbU9MbvKdPTRUs0o2MRPbvD7hrai21EFzpXcM9FNHVRL4Ojej0/NPg4cxJGqx3ouRWr7FJuYid0tGmdmqJbPqKohq6OGrp5EkhnjbJG9OjmuTKL8yn7EfbuN5W/bDtJVzkVHx29tJJlcqroFWBy+9Y1X3kgnCsTZmxertT/TMx5SvNFW1TFXWFL99u7eXbW6C0smR8VrtDOJnxJppHOd87GR/OXQU13bc7st8206xuKtxi6PpG4Xq2na2BF9/ZqvvLPyNw/OY6q5P9NM+c6R6ao3Nq9mxp1sMfIkcbpF6Marl9yZNhmwOxppzYzpO1LG+ORtsimmY5ObZZU7WRF+/e4oDp60uv2o7TYGuVq3S4U9ErkTPCkkjWqvuRVU2asajWo1qIiJyRE7iV5bX9KLVmOmZny3R6y1clt/ar8EOb5NsZX7CbnWK1XSWurpa2NETwlaxy/gSPKQqmFVPBcGx/afZ36h2caksUTEfNX2upp4kVM+e6NyNX3Owprcp5Flp4pVXKvY1y+1UNjkXe2sJXb/tq184/SXnnVGldNT7DmpI1Y3dHIqL7+QOS5IRsF3eLwl82I6Qr0VVclsjpnqvVXw/YXL+ExTPSCdyO6vrdkdXbZHc7XeKiGNuejJEbMn0yu+YnY4nm1nmMddt9VU+Wu73Lvh69u1TV2BFO9nd3WjYNqFIpUjnr2xUESL8LtpWten9Gr/mJWK07+N2Rmn9KaeRM+VV8tc5UXokEfCnzrOnzHrkVjn8xs0duvlv+D5xdzm7NVXYqmvVcdMlgNxmzuqtoeoL6qosdutbKXn8eok4sp7GwL+EV/Ljbj9mWi2WV96kiaj7vdZXxyJ1dDEjYWp7nMk+c6PyoxHM5bX11aR5zv8AdEq5lVvbxET1J8AByJawgTfrdwbD0k5ZZdqZyKvjh5PZ8TQxTtRs0TJGo5HIj2oqZTmi8+828BivomJov6a7M66PO7b5yiaXh7OLa+zbPdOWeTPHQ2qlpncsc2RNav5D3wgNe5XNdc1z0733TGkRAAD4ZAABBe+9/A3Tfduk/tFMELRb4+0XSN30xFom0XNtyu0NyhqKnyXz4aZI85a+TpxLnHCiqqYXOO+rqHWOSlm5ay+Irp01mZ39W5Vs2qiq9ulyWZ3B/wB367+Tbf8AeSsxZjcH/wAIa7+Rbf8AeT35T/hV3/r/AJUvnKfaY7pWqAByBaw1y7Xf4Xda/d+r/PNjRrl2u/wu61+79X+eXbkT7Rd/LHqhs5+7pYufnNyY3+cj/PafovU+JfRb/OR/ntOjq7R9qG0UBAcCXsAAHXuNVDQ0FRW1DuCGnidLI7wa1FVV+ZDWDBLLPE2oncr5ZszPcvVXPVXKvzqbEdu1c23bGNZ1SuVqpZKtjHJ3PfE5rfpchrwjajI2MTo1qNT3IdE5EWtLV251zEeUT80Bndf2aXJlGyKttVs2qaYul8rWUNtobglTPUPaqtZwMe5uUairzcjU6d5jBwnJeRdb1uLtuq3PTEx5oS3XNFUVQvn/AHQOx7v1vSJ7aef/AIB/dA7Hv896P/Z5v+AobxO68SnPE74y/OVP/ReB/vq84/8Aqlf41e/tj3r4rvA7HuX/ALcUf+zzf8AXeB2Pf58Uf+zzf8BQ7id8ZfnHE5PhL84/0Vgf76vOP/qz/Gr39se9fH+6C2Op11xSf7PP/wAA/ugdj3+e9J/s0/8AwFDeJ3xl+c54nfGX5x/ovA/31ecf/Vj+NXv7Y96yW9ltK0HrjZ7a7dpfUdPcq2mvcNS6FkcjFSNIpmudlzUTlxoVrOVVV71+cFhy3L7eX2Is25mY47+3u0aGJxNWIr26ocFhNxWqVm0DUtFn9utMMqp8iZyf2yvhOW5A9W7ZrlGi8n6clcvtSpg/9VNXlBTtZbejs+MS9ct9opXQABxtb1BN6P8AjE6w9tD+pxEcEkb0f8YjWHtof1SIjc7flfsNj8lP+MKbj/aa+9+VWuKOdfCJ/wCapsl2YRJBs10vAiYSOz0jET2QsQ1tVqf3lUfzL/zVNlez9MaD0/8Acum/RNKpy3n+TZjtn0hJZLxqe4ADnawBwvTkcnC4RFyvIDXlt4un1Z22ayr+HCtujqNPZTsbD+VimFn619wkvFyrbvMuZK+rnqnqveskrn/2j8ju2GtczZot/wBsRHlGikYivbu1T2vznescEsvxWOd8yGx/ZPbHWbZfpa1SNRslJZ6WKRET4aRN4vpya4ZKeSrRKOFMyVD2QsRO9XuRqflNocbWsY1jUw1qIiJ6il8t7mluzR1zM+WnzTOSU7qqn0ADnqeAAAIl3v8A+LzqX5dF+uQEtESb4H8XjU3y6L9cgJHKPb7H56fWHlf+7q7lGXek72nBy70ne1Tg7Yoy124d+93V33Sh/QoWUK1bhv73tXfdKH9ChZU49yk/E7vfHpC5YD2ajucL0Nb20uv+qm0vVdySRZGVF7q3Ruz1Y2VzGfitQ2OVk8dLSTVMq4jiYr3L4IiZX8hq9o5Hy0rJpVVz5cyOVe9XKrl/KWDkRb1rvXOqIjz1+TQzqr6lMdr9jhzuzar/AIqK75kycn4XBUSgqF/0Tk+dMHQVepjWYhsT2C0DbbsV0ZStbwf9S0sjkx0c+NHu+lymbHRsFE222KgtzERG0tNHA3Hg1qN/qO8cIxNznb1dfXMz5yvVEaUxAADxfQAACmsvUlvS0amvNoRMeQ3OqpceHBM9v9Rs0Xoa6ttFG6h2x60p3oiKt7qJkx4S4kT88vHIi5peu0dcRPlP6obOqdbdM9rEe8/Ktbx0c7OeVjd8+OR+pyjeJcL38joquUzpOrZRoO6Le9D2G9OXLq+209Uq+KyRNd/We0R1u0V63HYNo6ocuVZbWU/X+SVYv7BIpwvGWuaxFy31TMeUrzanWiJ7AAGu+wAAAAAAAAAAAAAAAAAGBV39ketnb7J7DdWty6jvSRKvg2SGTK/OxvzlCTZPvzW/y7dvvsyN4nUVRS1Df6djF+h6mtgzA2QbhtD5Ju5Wuo4ceW1tXP7cSrH/APbJ4Ir3SaPyHdy0ZDjHFROm/pJXv/tEqAAAAABgAAZAAARXvX2l922DajSGJHz0UcVfHn4PYyte9fwEf85RF2OJcLlO42YaptUV80zdLJOqpDcKOaleqdUbIxWr9CmsqmR7aaNkjXNkY3geipzRzfNX6UU6PyJv7WHuWuqdfOP0V7OqPrU1v0ABdUGuHuN3VarZndrRJKjpLdeJVYz4sUrGSN/GWT5iwBUPcTurafXGp7Euc11uhrW8+nYyLGvz9s35i3hyDlLY5nMrkdek+cb/AH6rjl9e3h6ZdK+3Gns9lrrtWO4Kaip5KmZ3gxjVc5fmRTWPFNLURpU1D1fNOqzSOXq5z1VyqvvUvtvUXh9m2D6mfC9rZ62BlvjRVxxdvI2JyJ945y+4oYuM8k5dELTyJsbNi5d65iPKP1RedXPrU0JR3UrM68bdrI5Y2yQ2yGouMrXJ04Wdmxfc+Vq+4vihVLcQsvaXXVmpJYl+xR09up5PbmWVv6ItahAcrsRzuYzR/bER8fi38rt7GHieve4U1o6vtTLDrG/WFiKjLbdKmlYip8BsruBfwVabLyhm9XaX2nbzfndmjIbjHTV8OO9HRJG5fw4nm5yKv7OJuWuunXyn9ZeOc0a2Yq6pReADpKsrH7iF0bDqfVlie9eKqpKatjb4dm58b1/Hj+YtohRDdOur7Vt5skaOa2K509VQSuVcclj7VqfhQonvL3nKuV1jm8xmr+6In4fBbMrr2sPHYFKt9S7NuG2SC3Ryq5lqtEUb2Z5Mlle+R34nZF1V6GurbTd3X3bBq66Oc1yOustPG5vRY4MQMX3pHn3nvyMsbeNquf20++d3pq883r2bGnXLEJXpFE+VejGq75kNiOw6w/Wzsh0tZnROhmgtsTp2OTCtmenaSIvr43OKBaTtCah1bZNPO4uG6XKnpHq1OaMfI1Hr7m8S+42YISfLbEaUWrMdOsz4bo9Za2S291VfgAA56ngAAAAAB0b5d7ZY7XPdLxcKa30NO3imqKiRGMYmcJlV8VwieKqiFW9r285W1/bWjZvEtJTL5rrzVRfZX8+sETvRTCek9M8/RRURSSy7KcTmFezZp3dMzwjx+HF4X8TbsRrXKfNqG1HR+zqjSXUNyzWSN4qe3UyJJVT9ebWZ5N5L5zlRvLGclRdrW3XWmvkmt0Mi6esL0VrqGjlXtZ2qmFSaZMKqLz81vC3C4XixkjCqmqKutnrq2qqKysqHcc9TUSLJLK7xc93NT8zpGVcmcLgdK6/r19c8I7o+PHuV3FZncvfVo3Q4Y1rI0jY1rWJ0RqYQ5TkmAcljRmurgs1uEf4Q118i3f7yVlLM7hH+ENd/Itv+8kDyn/Crv/X/AChI5V7THitUADkC1hrk2u/wu61+79X+ebGzXJtd/hd1r/3gq/zy78iPaLv5Y9UNnP3dLGD85ejf52P89p+i9T4l6M/nY/0jTo0K7R9qG0UAHAl7AABE+9zOsO77qVGuw6byWFPXx1UTV+hVKKr1X2l099epWDYukOcJU3ejiX14er/7BStM951HkbTpl8z11T6RCtZzOt6OyAAzTY3s+qtpeq6rT9HdoLXLT291b2stOszXIkjGcOEc3Hp5z6uhZ79+3h7c3bk6UxxRdu3VdqimniwsFjv7kzUH+fNrX/8A1kn/ADR/cmag/wA+LV77ZJ/zSI/1Jln/AJfdV8m5/DMR1K5HBY7+5N1B/nvav/DJP+aF3TdQ/wCe9q/8Mk/5o/1Jlf8A5fdPyP4ZiOpXIFjf7k3UP+e9q/8ADJP+aP7k3UH+fFq/8Mk/5o/1Jlf/AJY8p+R/DMR1K5HBY1d0zUPdri1f+GSf805/uTNQf572r/wyT/mj/UmV/wDl90/Jj+GYjqVxJv3JeW2qsXHpadqE/wDqKc9/+5M1B/nxa/8AwyT/AJpnuwXYVdtm+u5tRVupKK5xSW2WjSKGjfE5FfJE/iyr1/k/pI7N8+y+/grtu3c1qmN26fk28Fgb9q/TXVG6E8ABTl6xqC70f8YjV/tov1OIjckjek/jEav9tF+pxEbnb8r9hsfkp/xhTcf7TX3vxrf3FUfzL/zVNl2guWhrAn+rKb9E01o1v7iqf5l/5qmy3QOV0LYM/wDZlN+iaVTlv91Z75+CTyXjU9sAHO0+Hg7RK91q0BqK6NcrXUdqqqhFTuVkTnf1HvGA7xNalBsN1nOq44rPPD/SN7P+0bGDo5zEUUdcxHvfFz7Ete1G3hoqdvhEz81D9QqI1eFO7kcd53Xio0zrOr3tnECVO0vSNM9uWT6goGOT1duxV/IbJTXdsNgSp20aMiVucXiOT8Fj3f1GxE5xy2q/3Fqn/jPr+iyZNGlqe8ABSUwAAARJvf8A8XjU3y6L9cgJbIk3wP4vGpvl0X65ASGUe32Pz0+sPK/93V3KMu9JfacHLuq+04O2qPC1u4d+97V33Sh/QoWVK1bh373dXfdKH9ChZU49yk/E7vfHpC45f7NR3MZ2sVTqHZbqytauHU9lrJWr62wPU1v07eCmhZ8WNqfQhsH3iqnyXYbrOTpxWieP8NvB/aNfmMcvDkWzkTTphrlXXPpH6orO5+tTAcx061lRT0TUytRUQxY8eKRrf6zg9fQ8LanXumKZyKqTXyhjX2LUMLjdq2KJq6kPajW5THa2VoAgODLyAAAAAClA952Pst4HV7GphFlpX/hUkOfpyX8KI72sPY7f769P8dSUUn/lcP8AZLfyLnTHVx/xn1pRebx/t/GEVHLeSocHKdTpyrLybnsrX7ALFGi57GetjX1f31KuPmVCXiENymbtdi3Z55QXerj+d6O/tE3nFs6p2cwvR/yn1XbDTrZpnsAARj3AAAAAAAAAAAAAAAAAAYEdbzVD9UNgGt6fGeC0TT/0Sdp/YNVht32mUf1Q2b6nt+M+U2erhx48UL0/rNRAgbYtglN5HsQ0PBjCpYKJzk9boGOX6VUzY8DZxB5Ls803TYx2NppY8eyFqHvmQABgAAAABkAABwvQ107ZbS+xbXtX2tzGsRl2mnja1OTY5sTMT8GRDYuUp30rOy37ZYblFFwsu9pilkf8eWJ741+ZnZFu5GX9jG1W5/qp98aT6aorN7e1Y2upCQAOnKukzdauq2nbzp1yypHBXJUUE2fhccSuYnvfGwvuhrL03dUsOprNfnIqpbLlTVrsd6Rytcv0ZNmbV5dTnHLaxs4i3d6408p/VZsmr2rM09Uq37+F2bHpXS+n+fHWXN9YuPiwRKmPwpmfMVOamcIThvsXfy/a9R2qOZXRWq0sa9ncyWaRz3fOxsRBVTIsVNNKnJWsVU9uORauTdjmcttR0zv853e7RFZnXzmJmI7l3tzG0Nt+xCkuHNH3iuqa5yL3Jx9k38WJq+8mkx3ZlZHac2dacsMjGsloLZT08qN6do2NqPX3uyvvMiOWZjf+kYu5d66p8tdyz2aYot0xHUFSd+60sh1bpW/MRVdV0VRRSL3J2T2yM/Sv+YtsQPvv2uSs2SUdzijRVtd3gmlfjm2KRHwr+NIwkOTd/mcytT17vONPXR44+jbw9UKadFA594OwKa9XRt1jsGtbBfpXOSK3XWlqpFT4jZW8afgqpstToauqiPtYJYv5Rit96obINlN7dqPZnpq+ySJJNW2unmmcn8osacf43EULlvY3Wr0dsekx8VgyWvdVQ9XU91hsenLne6n9ot9JLVSfJjYrl+hDWVTukfAySZyulkTtJFXqrnecq/Oql8d7G7utGwbUSRStjnr2RUESL8LtpGsen4CvX3FEuWVx0NjkTY2cPcu9c6eUfq886r1qpoS1uiWZbtt0t1SuOztFFUV7kVOSqrUgantzMq/el6CsG4dZVSk1ZqR7WObLUQW6F3e3s2LJInsVZWfglnyucrMRz2Y1U9FMRHx9ZSOWW9jDx27wAFaSAAABHm2ba3pvZnb2pcFdX3iojV9FbIHYklxy4nL0jjzyVy+C4Rypg8/eC2u0GzOzMp6RkNfqWvYq0FE53msb0WeXHNI0Xu5K9eSfCc2jV6ulzvl4q7zeq+a4XOsf2lTUy+lIvREwnJrUTCI1MIiIiIiYLZyf5OTjtL9/db99X6dvl2RmPzCMPGzT9p720raBqnaJdkuGpq5JIo3ZpaCFFbS0vyWd7vF7su9iYQxdVz1OB7DptmzbsURbtxpEdEKzcuVXKtqqdZDhyo1Gq5ccSo1qImVcq9yInVfUZvsq2Xau2k1f/UVI2mtTH8M92qkVKdnijO+V6YXzW8kXHErclwNkmxLRmz3sq6np3Xa+o3DrpWtR0jVxheyb6MTea+j52FwrnELmvKHC5drRM7VfVHxno9exu4XLbt/fO6FPr7sv1fp7Z63WmoqL6kUstZDS01FUoqVUvaZ89zf8UiY6O85eeUTlnCvUXO33f4HKb1Xuk/tlMj0yHMLuYYWb9zdOs8Orc+cww9Fi5FNLgs1uEfu/XfyLd/vJWUszuDr/ANYa6+Rbv95PPlP+FXf+v+VL7yn2mPFaoAHIFrDXJtc/hc1r/wB4Kz882NmuXa7/AAu61+79Z+eXfkT9/d/LHqhs5+7pYufnN6Lf5yP89p+h8SdG/wA7H+kadGV2j7UNooAOBL2AACAN+p/Dsoszfj6ip0+aGdf6inZcHft/gtsP/eSD9XqSnyHV+SX4bT3yrGcff+ATruPIi7YLtz5/W8/9Zi/9CCidNx5U/wCmS6t8dPSfrMP/AKm/n/4be7mvl3tFK54AOMreAAAAAAAAAAAFAAoLvR/xidYe2h/VIiN1JH3o/wCMRrD20X6nERwdvyv2Gx+Sn/GFNx3tNfe/Gt/cVT/Mv/NU2W6B/eLYPuZTfommtKt/cVT/ADL/AM1TZboH94mn/uZTfomlU5b/AHVnvn4JPJeNT2wAc7T4RRvcyLHu86qVPhR0zPwqmJv9ZK5Em+AuN3vUSfGkok/+shJDKN+Psfnp9YeV/wC7q7pUZk/bHL61Pk5X0l9oO2wo8JA3bY+129aPZ/8AvJnfg00y/wBRsC7ige7EqJvA6Qz/AC9V+pzF/E6HM+Ws/wC9oj/jHrUtGT/ceIACnJUAAAiXe/8A4vOpfl0X65AS0RLvffxetS/Lov1yAkMo9vsfnp9YeV/7uruUYdniXPicBfSX2g7ao61u4b+97V33Sh/QoWVK1bh373tW/dKH9ChZU49yk/E7vfHpC44D2ajuRfvWSdlu+6ud40rG/hTRp/WUOf6a58VL173C43edVeuOnT56mJCia+kvtLjyL9hr/NPpSh86+9p7nBkOzBOLarotF551HQfp2mPmRbK+e1jRWf8AOKh/TNLTivuK+6fRGYf72nvbIAAcJXcAAAAACkG+TGke3OdU6y2ikev4Uqf1F0b5dKW0W6SsqnKjW8mMT0pHL0anrUozvP1U9dtY8tqlTtp7VTvcidG+fLhqepELfyNs1/TJudGkx6ILN8Zaj/ba/Wnf3R29/QjBTlTgHTlfXM3Hlzsgrk+LfapPxYydyBtxzP8A0Q3BfG/VWPwYyeTjOffiN780rpg/uKQAES2QAAAAAAAAAAAAAAAAAGB+VZA2ppJqZ/oyxuYvsVMGnz6j3D+QU3EFCPrGZ/I/QBenT0XYWC3Q/wAnSxN+ZiId4/OmZwU0TPisRPoP0MgAAAAAAAwAAAEC73mz7UGtotIzaZtz6+up62alfGjkY1kUsfGsj3ryaxqwtTK97kRMqqIs9GC7fa/UFp2Qajuulq99BdaGl8qinZGx6tZG5rpOT0VP2tHp05EjlN+7Yxluu1pta6Rrw37t/ZveOIoprtzTVwQFb90zUMtE2W465tlLVqnnQU9sfPGi+CPdIxV9vChiOsd3Ladp+lfV0lLbtQwM4nOS2TOSdrUTOeykROJfUxXL6jxqfbftbp5O0i13Wv8AVNR0z2r7lj/ISrs53qKllRFRbQ7PF2DnI36p2tjvsfTnJAqqqp1VVYq+ppf7n+ocN/M1puR1RH6RPvmUHR/D731eEqyVkEqLV0M0MsFRHxQzRSxqySN/NFa5q80VF7lNjGx68Lf9lOlrw6RJZaq00zpnIucydm1H/jI4xTaVs00Ptm05TX+21dNHXyQ8VvvlCiPVze5siJ+2sRU9F3NqoqIrVyfhsEiu2z/YvX2XVlE6nqNKTViSSNysVTCmahssTl9Jitkwi9UVqoqIqKiQeeZnazXB07MbNymrSaZ7d27x07W7g8NOFrq360z0qlbbru6+7ZNYXN3CqLdZKVip0VlOiQNX3pHn3nlbP7OmoNoGm7E+J0sVfdqaGZrea9l2iOkX3Ma48KGSWaJs9Q9XzTZmkcvVXPVXKq+9xMO57Zlum3SkrFyjLPbqit6clc9EganzSvX70u+KqjA5fVp/RTu8I0hC2o5/Fb+mV5AAcVW8MF3gLOy/bFtXW56K5Vtc08aeMkSdqz8ZjTOjh3NF5J7z1sXZs3ablPGmYnyfNVO1TMKF7O9hW0TXFkivlHRW+0W6pY2Wmmuk72LOxyZRzGMY53D3oruHKKiplD1r9u07VLbSrPSwWK8qnWKhrnMkx38pmMav4R3dWbw+1eg1Jc7R2mnKGS3V09G9lPbnPaqxSOYvN8irz4fVyPY0jvW6kpaprNXaZt9xpFVEdLa3OgmYne7s5HOa9fVxN9p0y5ez/wC9t00TTxins7509Vei3l+uxMzqgG7265WS5vtl5t1ZbK+Pzn01XC6KRE7lRF6p60ynrLn7mN1Sv2IUtBz4rRX1NE5V704+2b+LM1Pce7BU7LtvejXwp5Pd4IsK6N6LFWW+RU5L3Pjdy6p5rsLzcmTy93zZxfNmF61XZKir+qNgrJYKy11eWtfxcLmSxyM7pERsXNE4XJz5LlqQmcZxRmGBqs3qZou0TE6T09G7wnXSfe3MHg5w93aonWmWFb+N2SPTulNP4/ddxlrVVF6JBFwpn3zp8xVFqdMry7ybd9S7tuG2SC2xTK+O02iKN7PiTSvfI752dkQi6GeqxSUrFfUVDmwRNTq571RrU+dyFp5OWeYy21E9Ma+c6+miLzKrnMTMd0L27pdkWzbCrE+WHsqi59rcZvtu2ermL/R9n8xK50dPWunslgt1mpEVKegpYqWLPxI2I1PoQ7xynG4j6TiLl7+6ZnzlZ7VOxRFIADVegY3tM1hbNB6LuOp7qquhpI/scLV8+eVy4ZG31ucqJ6uaryRTJCmm+brh9917Bo2jmVbdYESSpRF82SskbnxwvZxuRE8FkencS+SZbOY4um1P2eM90fPg1sXiIsWprQ1qu/3bVWpK/UV+nbPca+TjlVueCNPgxsRc4Y1MIiZ7ueVVVPMBynNcHZKKKaKYppjSIU2uua5mqri+Xuaxive5GtRMqq9EQsRu+bvb9QQwap2gU80FqeiPorQqqx9S3qkkypzazwYmFXquE5L190vZJHqiuj17qOBJLJRzKlrpXty2snYuFldnrGxyYRPhORVXk3DriFI5Sco6rNU4XCzpP9U9XZHb1z0cOKdy7Lo0i7djuh+FBR0tBRw0VDTQ0tLAxI4YYWIxkbUTCNa1OSIidyH7gHPJmZnWU9wQXvu/wN06/wCuqT+0UxQudvu/wNQfdqk/K4ph3HU+R/4d/wBp+CsZx985LMbhH+ENd/It3+8lZyzO4R/hDXfyLb/vJt8p/wAKu/8AX/KHnlXtMeK1IAOQLWGuTa5/C7rX/vBWfnmxs1ybXf4Xda/d+s/PLtyI9ou/lj1Q2c/d0sYPiXo3H8rH+e0+j5kzhqf6WP8APadHV2j7UNooAOBL2AACAN+tiu2TWh2MozUVMvszFOn9ZTsurvq0yz7FVlRM+TXajlX1Zfwf2ylSHVeSFWuXadVU/BWM4++CcNyRyM201yfymnqhPmqKdSDyWN0WsWl292iFFx5ZRVlOvrxGkn/2yUzunay+9H/GfdvauXzpiaO9esBAcWXEAAAAAAAAAAAAAUF3o/4xGsPbRfqcRG5JG9H/ABiNX+2i/U4iNzt+Vew2PyU/4wpuO9or7341v7iqf5l/5qmy3QSY0NYE/wBWU36JprSrf3FUfzL/AM1TZboH94un1/1ZTfomlU5b/dWe+fgk8l41PbABztPhEW+Dz3f79/P0X63CS6RLvfNzu96kXHovo3ezFZCSOT+32Pz0+sPK/wDdVdyi7uvvAVPOX2g7ZCjs+3cpOy28aPf/APvZGfhU8yf1mwROhrr2KT+S7ZdFy5xm9QR/h8TP7RsUToc15bU/7u3V/wAfjPzWbJ5/kzAACmJcAAAiXe+/i86l+XRfrkBLREm9/wDxedS5+PRfrsBIZR7fY/PT6w8r/wB3V3KMu9JfacHLvSX2nB21R1rdw3H1vau+6UP6FCypWrcN/e/q77pQ/oULKnHuUn4nd749IXLAezUdyK97Riv3etWJ4QwO+aoiX+ook9MPVPWX93lqZ1XsH1jE1Mq22Pl/AVH/ANkoE/m9V9ZcORU64KuP+U+kIfOo/m0z2PkyHZk9se1HRkjlwjdRUCqvgnbNMfPW0RIsWu9Ny9OC9US5/wDnsLXiKdq1VHZPoiLdU01RVHQ2HS6tszHKiSTyYXGWQuVPyH7UuprLUORvlrYnr8GZqs+leRH87USolaickkcn0qfKoipz5p4HJ/4ZZmN0z+/BmnlJi4q3xT5T80sse17UcxyOaqZRUXKKckVUVVVUL+KhqZade9Grlq+1q8j2qfV90jaiTQUs+O9MsVfyoalzK7kfYnX3JbD8pMPXH82maZ84+fuZ2dS73KjtVC+srZUjjbyTxcvc1E71XwMQq9dVETMJboUevRFmVf6jEbtcqy7VnldfLxvRMMY3kyNPBqflXqp9YfKrlVX8zdDwzHlTh7NuYw/1q+6YiO/5ej9r/eKq91/ldQnZxsykEOf2tq96+Ll719yFZt49yLtMjTvbZ6VF97pV/rLEZK37wsiP2p1De+K20bF9vA539ou+Q26aMRs0xpERPwVDLrtd7E13Lk6zMb58YYAcA5TmqFwTi6O5NH2exh7/AOVvFW9Pwmt/qJxIf3OoEj2A2WVE/dFTWyL/ALVI3+yTAcWzqrazC9P/ACn1XbCxs2aYAARj3AAAAAAAAAAAAAAAAAAYAg362G/yafMTkY/9S2fFT5gMgToD4gdxwRv+M1F+g+zIAAAAAAAMAAAB5+o7ZDetP3GzVP7RX0ktLJ8mRitX6FPQOFQ+qappmJjjDExrGktXUCSsgZHM1WSxp2b2qmFRzfNVPnRT7LP/ANzBcL7rXUd2vWoW2e11N2qZqGlooGyyuhkkc9rnPd5rPS9HhcvrOrqndOuEFM+bS2sI6yVE82ludKkaO9ksfRfaxfcdcp5T5bMxTNzSZ7J07tdP0VWvK8RMzVEIl2M7Ub/swvTp7fxVlmqpEdX2tzsMk7lkj+JLjv6OwiOzhFS79rueldqOz6WSiqUuFku9K+mnax6se1r28L43YXLHoi4VOqGvTUVkvOmr3PZdQW2otlygRFkgmRM8K5w5rk817VwuHNVUU72idYap0Tc5blpS9VFsnmREna1qSRTonTtI3Za7GVwuMplcKmTwzjIbeY6YjD1RTc4xPRPVw90/uPXCY+rDzzd2N3oshd90q0SVbXWbXF2oaROsNTRxVD0TwR6cH0ovvJV2QbJ9K7NI6x1kdW1dfWNYyqrKyZHyOa3KtaiNRGsaiuVcIiZ5ZzhCsNfvL7Vay3rRxT6eoHqmFq6e3uWZPXiSRzM/elid1Ogq4NjdvvFznqai536ea61k1RIr3yukdhjlVfGNkZWc5tZtZwf+9va0zMRER09O/dG7d06pLCV4Wu5/Jp39aVgAU5KAUAyKA7y1o+o23bVEDYlZDVTRV0Sr8JJomq9f6RryOi3m8nsa1LtC2l2S66ffRU9M62upK+qqn4ZD2cvEzDW+c9zklfhOnm81TlnFLhul3qOle+3a9oKmpxlsVRanRRqvy2yuVPmU6pl3KHA28JZpvXIirZiNN88N2/SN3irOKy69Xeqqojcr7py83fTd8p75p+4TW250/wC11ES9U72vReT2L3tdlFLxbBNsFr2mWh9PO2K36lo40WuoEdyc3p20Oebo1X3sVcL1a51NdomgNXbPriyj1Ta/J45ncNPWQO7SlqF58mvwmHclXhciOx3HhWuvr7TdaW62qtqKC40j+Onqad/DJGuMLhe9FRVRUXKKiqipg2c0yrDZzYiuiY2v6ao3+E9cenunzw2Ku4KvYuRu6lwdsO7nRa41XXaqt2qqy1XKuVjqiOembUwOVkbY28KZa5nJqfCX2HlbOd16msGprdf79q+ouUtvqoquCmpKNtPH2sb0e3jc5z3OblE5Jw9COrPvRbSKKhZT11u05dJWt4fKZIZYXvX4zkY7hVfko1CQd2fXmuNp2028XjUdfA222e2NiioKGN0VOyad6Kj1RVVXu4YXJlyrjiXCJlSt4ixnmCwdUXLsRbpjTomdOERG7X0Sdu5g712Jpp1qlZFAEBRUuAADzNV3ml07pm6X+tRVprbRy1cqNXmrY2K5UT1rjBrUrK2suldVXW4v466vnkq6l2MZlkcr3fSpdHfSvT7XsWlt0au4r1cKegVWrhUZlZX+5WxK1flFKuaqqr4nSeReFijDV3541Tp4R+s+5Xs5u61U0OFPe2e6Uq9b64tGlKR74nXGfhmmanOGBqcUsnPllGouPFyoneeCWa3E9NcdZqbWUzEXg7O1Ur89OSSzcvasPP1KWLOMb9Bwdd6OMRu753R80dgbHPXopngs7ZLZQ2az0dptlO2moqKBkFPC3oxjURrU+ZDuBAcVqqmqdZ4rlEaAAMCDN9vnsah+7NJ+VxS9OhdDfcz/ANDUKeN6pPyuKYHVOR/4d/2n4Kxm/wB8FmdwjPl+uvkW3/eSsxZncI/whrr5Ft/3k2uU/wCFXf8Ar/lDzyn2mPFakAHIVrDXJtc/hc1r/wB4Kz9IbGzXHtb/AIXda/8AeCs/PLvyI9ou/lj1Q2c/d0sZPiXo3+dj/PafXefMnRv85H+e06MrtH2obRQAcCXsAAEUb3MKy7vmplamXRJSyp6uCqicq/MilFHJh6p05mw7brQNuWxnWVI5iyKtkqnsane9kTnt/Gahrva/tGtkTmj2o750ydL5FXNcHXR1Va+cR8ldzqn+ZTPY5M53friy1bcdG1j/AIVyWl/p4pIk+l6GDHbs1ydZL5bL4xFV1tr6esRE/wBFK1/5EUteKtc9Yrt/3RMecaIrD17F2mrtbOE6A4aqOaiouUXopycJXcAAAAAAAAAAAKABQXej/jEav9tF+qREbEk70f8AGI1f7aL9TiI3O35X7DY/JT/jCm472mvvfjW/uKo/mX/mqbLdA/vE0/8Acym/RNNaVbjyKo/mX/mqbLdAZ+sTT+f+y6b9E0qnLf7qz3z8EnkvGp7YAOdp8Iw3qoFqd3/VsaIqq2lZLy+0mjf/AGSTzDduUCVOxjWsKpn/AKhrXInrSF6p9KIbmXV7GLtVdVUesPO7GtEx2NeD+T3e0+ThjuNjX/GajvnQ+juMKNwe5s6mbS7StJVUi4jhv9BI9fBqVDEX6FU2TIaupJ30qNq41VH08jJmqnXLHo7+o2iMVHNRyLlFTKHPeW9GldmvriY8tPmseS1a26o7XIAKImgAACJN7/nu8am+XRfrkBLZEm+B/F51L8ui/XICRyj8Qsfnp9YeV77urulRh3pL7Qcu9JfacHbFHWt3Dv3vau+6MP6FCypWrcN/e/q77ow/oULKnHuUn4nd8PSFxwHs1HcxLbNTuq9kOsaZicT5LFWtanivYPx9Jrnhdxwxv+MxF+dDZ3dqOO4WuroJfQqYHwu5dzmqi/lNYFDxeQwI5FRzWI1UXuVOS/kLRyIua2r1HVMT5xPyRmdU/Zl+52LZOtLdrfVIuFhrqeTPyZWL/Udc/GuVUopnN5OaziT2pz/qLxMa7kHRGtUQvnV/uufl/jXflU/I+pZGzTOmbjhkRsiexzUd/WdWeriiy1v2R/gnRPapzOmJ0iEBXVFG+rc7HJG5XCInVVXodKorG+jBz+3Xp7jrTSyTLmR2U+KnRD4PamjTi0bmKmd1O4dzcrnKrlXqqqAcH21BeilZtuz+Pa9fm5z2TaSL5qaP/wBSzDuaKnevIqvtYqfK9qeqps5xdJIv6NrWf2SayKP9xVPZ8YTGTR/Mrns+MMZPpmOJFXxPk+Kh3BTyv72sc75kUtSeiNdy/O6vSPot3/SUUjVa59I+fHqklfIi/jEmmM7J6F9s2XaUt0jeGSmstJE9PBzYWIv05MmOG4+5zuKuV9dUz5yvNrdRHcAA1XoAAAAAAAAAAAAAAAMAAAB84b4H0eL9Uk+MgHesUnbWSgm69pTRu+dqKdw8PZ7P5VoHT1TnPa2umfn2xNU9wyAAAAAwAAAAAyAAAqNvi3zV+ntqdv8AqNqrUVsoKy0slbDR3GWGJJmSva/DWqjc8Kx5Mf2a7yGuNN1cNPqiR2qbPlEk7RjWVsTe9WPTDZF6rwvTK8k4kJD38bOslh0pqFqpilr5qF6eKTx8afTB9JVI6lkmEwmYZXbi7RE8Yndv3T18eGitY2/dw+JmaZX01ZpzQW3nZ5T1dLVxztVHPt1zgb9nopsec1UXCp3I+N2M4ToqNVKcbSNm+sNntdLDqK1TLRNd9julNE6SklRei8aJ5i/avwvtTCn1so2iag2bal+q9kd5RSzqiXC3SP4YqxievC8EiZ81+OXRcoqot4Nmu07Ru0KgbPp+6xuquHM1vqFSOqgXvR0ec4TPpJlq9yqRtc47k5VMURzlifd49HlpPfq2o5jMadZ3VNd1JDUXeRaCy09Tca2ZFZFBSQvle5y8kwjUVe82Z6VtMNg0zarFTuV0NuooaSNV72xsRifQh9VF0s9FdaS1T19HT3Cv41paZZWtln4G8T1a3q5ERMqvceghAZ7nteaRRGxsRGs8dddfCOpu4LB04bXSddQAFdbwADIiTewveq9ObLG3zSd2mtc1LcqdKuaJjHKsD1WNU85qonnujXPqK4aW3iNqdkrGSVl3p9QUqcnUtwpWMVUzz4ZIka5Hd2V4k9RbPbzY/ri2N6stKMV8klsllhanVZYk7SNPw2NNeUb0kjZInR7UcnsU6FyUw2FxmCrt3rcVTE9MRrpMRpv49aCzS9ds3KaqKtF89A670Ftx0jWWaqoonyvhRLnZK7CyRJlPOaqek1HYVsjcKi8K+a7CJWrbbsK1HoKrluVkgrL7plyq5s8TFkqaNOvDM1qZc3H+MamOXnY5Zi6x3W6WK9Ul6sdfLb7nRv46epixxMXoqKi8nNVOStXkqclQursP276e17TQWq9SU1j1QiIx9JJJwxVTunFTud6WevAvnN5p5yJxLnEYTF5Bcm/g/rWZ40z0fvr8+1bvWcfRsXd1UKNeVUvCq+VQJ7ZERfm6lwtxe0Ng2d3q/LE5slzuqxserVRJIYY2taqZ6pxul+knC4WvT7HyXWvt1sa+FqySVU0EaLGiJlXK9U5IiJnOTu2yrpLhbqa4UE8dRSVMTZoJY1y2SNyIrXIveioqKhE5xym/iGFmzTb2YmY1nXXhv04Q2MHlsYe5t7WrsAAqKVAABVHfzuaSXfSFkZKqLFFV1srM+KxxsX9IVqJr31aplTtthiaqKtJY6aJyJ3K6WZ/5FaQodj5PW+by2zHZr5zr8VRzKraxNTlEyqetS9O6NbW2/YLYZVibHNXuqK2ZUT01kmfwqv3iMT3FFHO4Wud4Iq/QbEdhkTYdi+iWMTCfUCid71gYq/SpD8tbkxhKKOur0ifm3Mlp+vVPYzMAHNFiAABBm+4v/wCjcKY63qk/K4pgXP33OexqD7tUn5XFL06HVOR/4d/2n4KxnH3wWa3B/wB366+Rbf8AeSsxZncI/d+u/k23/eTa5T/hV3/r/lS88p9pjxWpAByFaw1x7XP4Xda/94Kz882OGuXa+mNr+tU/19Vr+OXbkR7Rd/LHqhs5+7pYsfMnNGJ/pI/z2n0fjWLwU739EarXZ9jkU6OrtH2obSQcJ0OTgS9gAA/C4UsNbQz0dQ1HQzxujkRe9rkVFT5lNYEMMtMzyWdFbLTq6F6L3OYqtX8htGU117a7U+y7Y9YW16ImLtLUMROiMnxO36JC9ciL2ly9a64ifLd8ULnVGtumpiB8yxpLC+NV9Nqt+dMH0M4Ohq7E6b4bD9hd9bqTY/pW79qssktsijnevVZo29nL+OxxmhXjcZ1B5boO9aZlmc6W0XFZomqnJkFQnG1E/wDmNm+csOcUzfDfRsddtdUzp3Tvj3Lthq4uWqao6gAEa9gAAAAAMW1BrShtOvtNaN7J09wvvlEicDkRKeGGJXrI5PW5Eaid+V5+bgyhy4RVVcciqWxzUn/STvfXbVUcjpKGhtdSy2814UpmPjhY5M4xxq+STC/H9RKZdgYxFF67X9mimZ8eFMee/wAHheubE00xxmVrgE6DuIt7qCb0a/8AvEaw9tF+pxEbkk70eP7ojV/tov1OIjc7flfsNj8lP+MKbj/aa+9+Nb+4qj+Zf+apst0Dy0LYE/1ZTfommtKt/cVR/Mv/ADVNlugf3i6f+5lN+iaVTlv91Z75+CTyXjU9sAHO0+HQ1Fb23awXG1vxw1lLLTrnph7Fb/Wd84U+qappmJhiY1jSWrehVVooOJMKkaNVPBU5L+Q/c9fXFvbadd6ktTG4ZRXmsp2pjHmtmdw/QqHjneKK4uUxXHCd6j3Y0rmO18VDO0p5Y/jsc350VDZBspubrzsw0tdZHo+Srs9LNI7Pw3RNV305NcaclyXf3ObtHcdhdso2vc+W1VVTQyqvcrZVe1PwJGFQ5aWdrCUXP7avWP0S+S16V1UpjABzRYgAACJN7/8Ai86l+XRfrkBLZEm9/wDxedS/Lov12Akco/ELH56fWHlf+7q7lGFzlfaA70l9oU7Yo61u4b+97Vy/6yh/QoWVK17h373dW/dGH9ChZQ49yk/E7vh6QuOA9mo7hTWztFtyWfaNqm1MjWOOkvVXHE3wjWVzmfiuQ2TFE97W0utW3m7zLwtjulJS18bUTonB2LvxoVX3ktyKvbOKrtz006+U/rLUzijWzFXVKKT4qG8dPIz4zHJ86H2cp1Q6WrMTpOq3+k7jJcdFaeq+P90WikkeqLzc7sWov0op30TCYRORhuxCq8r2S6ec56OfBHNSP9SxzPRE/BVpmRQL9vm7tVHVMx71RxkTF+uJ6Jn1cAA8muAAD9KRvHVwsx1lb+VCnF8rEuWoLtc0XKVtwqahF9T5XKn0YLdXeu+pdlud2/yGhqKn8CJyp9OCmtI1W0sLXdWxtRfbgsOQUb7lXdHr+ieyen6ldXd8X6hlJNcZobdToqzVk0dNGid7pHoxE/GBmewu1LettOjre1el1ZVuz8Wna6df0aE9iL0WbVV2f6YmfKNU7Yp27lNPa2GxRtiibGxMNaiNRPBEPoA4QvAAAAAAAAAAAAAAAAAAAAAMAQ19cjfj/STKUc+vdv8ALJ84FsNhdT5ZsU0RUZyr9P0PEvrSBiL9KKZmRfuo1fl27toubOeG39j/AEb3M/skoGQAAAAAAAAAAAAAYhte0JQ7RtFTaarquSiR08U8VTGxHvifG9HZRF5c04m+xykB6u3UKqGifPpLVvlVQxuUpLpTtaki+qWP0V9rFTPehaK6QS1VuqaWGqlpJZoXxsnjxxxOVqoj255ZRVynsKb7M95TW+nq5lFrtiagoGPWGoeyFsVbTq1eFVbw4bJhUXzXIir8buW0ZDVmlVqv6FXGlExM0zpv16tY7OuEfjKcNtRz0celDN+tN1sN6qrLfLdUW65UruGammTzm+CoqcnNXqjkyip0U8+WGKVUWWNr1b0VU5p7FL76/wBD6G246JortSVsb3yQLJar1SJmSLPcqLjibxcnRuwqKip5rkylZ9Q7uO1i2VroKG0UF9gz5lRR18cSKndlkytVF9SKqetS5Zbykw2Jp2b8xbuRxid3lM+k74Q+Iy27bq1tb47Hv7jFhjqdpd9v3ZNcy22plPl2VVsk8mUVM/awvT74uSQtumbO77oPSV3fqi3MoLvcq9HrE2oZNiBkbWxormKqZ4lkXCL3k0lB5R4unFZhXXROtMaRE90fPVO4G3VbsUxVxAAQbbAAYHzI1r2OY9EVrkwqL0VCrc+6L2VFwW/aBJ2rM9m2otTVYqc8NVWyIqd3Pn7C0xVnbFtk2jbNtt1ztkc9HdLAsdPU0tBV07WfYntRHIyZiI5F42SIiu4kTwUn8grzCbldvA1xTMxrMTpv074nfv7GnjKbGzFV6NYQXtJ2far2e3ZlBqagbHHMqpS1tO5X0tTjuY9eaO+0dh3fjBikjGSNVsjGvb4OQv1pDU2gdu2gaukmo2VUC4juVrq+U9LJ1avJcp0yyRq88clRUVEr1tA3Zdb2evkfpF0Opba532Jsk7KerjTPR/GqRvwmPORUVfioXbLeUdFVU2Md/LuU9e6J+Xp1IXE5bMaXLG+EMU8V1v76HTMVwr5m11XBSU9PLWSuiR8kjWN81XY5Z8DZlR08FHSQ0lLEyGCGNscUbEw1jWphERO5ERCnew/YftAo9qen7xqjTH1OtFuqVrJZJa2B7uNjHLEjWMe5VXtFavcnJS5JW+V+Os37tu3ZqiYiJndMTGsz2d3vSWVWa7duZucZ6wAFOSgAAKHb2MnabwOoW/ydPRM/8lF/tEWkob1zFbvBakVfhw0bk9nk7U/qIvO25T7DZ/LT6QpmN9or73Dkyit8UVPoNh+widtRsV0VI3miWKjYvtbC1q/ShryRcORfWXk3QLk2v2DWWn7XtJrdLUUUv2qsmcrU/Acwr3LS3M4OiuOir1ifkkclq0rqjsS8ADmaxAAAgzfc/gahz/21SflcUwLV79uoIo7DpvSjHNdNVVrrjMiO5sihYrW5Twc+Tl8hSqaZxzOr8krdVGW0zPTMz8Pgq+b1RN/SHJZncI/whrr5Ft/3krKWa3B/3frv5Ft/3k9+U/4Vd/6/5Q+Mp9pjxWpAByFaxehrv2606022vWkapjN2fJ+Gxjv6zYgUH3o6ZabeB1VlnCyZaSdnrRaWNqr+E1S48iq9MZXT10/GETnFP8jXtRop+FwTNDUJ/o3L8yH7qfMjO0jdH8dqtX3odMhWqJ0mJbOLFVtuFkoa9q8TammjmRfFHNRf6zumE7BrlHd9i+jq2N/Hmz00T3eL440Y/wDGapmxwjE2+avV0dUzHlK8W51piQAHi+wpZvrWVbdtgpbtHCrIbxao3Ok7nzQvcxye5joi6ZXrfm0+tds/s+pYo1dJZrkjJHdzIKhOByr/APMSEsPJfE8xmVGvCrWnz4e/RpZjb5zD1QqEcDn0XqDrkKelfdO1OmmttNvp55Ujo77C+2S8TlRqSr58K473K9qsT+cUvYnQ1fMlqIJY6iknfT1UMjZYJmLh0cjVRzHIviioimxfZLrCm15s9tGqKdGsfVwIlTE3/EztXhlZz54R6OxnqmF7znfLTAzTcoxVMbp3T3xw849Fkye/tW5tz0MqABRkyAAAAfnUSxQQvnnlZFFG1Xve9yI1rUTKqqr0THeNNRDu9zrldJ7M5LRQzcF21CrqKDhXzo4cfZ5U9jFRqd6OkavcRnuGWlr75q29LErUpqaloIHd2HK+R7fdwx/ORJtt14/aLtGrtQROelriRKO1McmMU7FXz1RUTCyOy/nzRFanPBZ/crs62/Ysy5qqKt5uNRWJywqMa5IWp80Ofvi/YvCfwrIZt1bq7kxr38dPCI080Nau/ScbrHClNwAKCmVBd6P+MPrD20X6nERv0JI3o/4xGr/bRfqcRG6nb8r9hsfkp/xhTcd7TX3vxrf3FUfzL/zVNlugf3iaf+5lN+iaa0q39xVP8y/81TZboD94mn/uXTfomlU5b/dWe+fgk8l41PbABztPgAAoTvSWp1q286iTsUihrkp66HHwkfE1j1/pI3kZKWQ37rL2OpNLakjY5UqqWe3zP7kWNySxp7V45fmK4HZsixHP5fZr7NPLd8FPzC3zeIqhwWc3EL6jajVWlZZERVWG507O9cp2Uq+5WxfhFZDPd3jU31p7ZtPXGSR0dJWTLbKzCoiLHPhrVVV6NbIkbl9hnPML9KwF23HHTWO+N/w0Mvu81iKZ8GwUBOgOMLgAAARJvf8A8XnUvy6L9cgJbIl3vv4vOpc9OOi/XICQyj2+x+en1h5X/u6u5Rh3pL7TgKuVX2jvO2qOtbuHL/7P6u+6MP6FCypWrcN/e9q5f9ZQ/oULKnHuUn4nd8PSFxwHs1HcFV9++xKlVpTVEUXJe3ttRJ7USWJPxZfnLUEXb1Gm/rk2JX5kUbX1dsjS6U2e50C8bsetY+0b98eWQ4n6NmFqueGuk+O74vvGWuds1UqHA4a5rkRzFy1yZRfFDk7Mpae92auSbRt3tqqqvorr2qepk0SKn40biVF6kBbtNySm1ndLS5yNbcrd2jU+NJA/iRPwHv8AmJ8Upma29jFVdu/9+Oqs5pRs4mZ69J93z1AARyPAABhm3Cu8g2UX3hfwyViQ0LPX2sicSfgNcVkXqvgTfvO3JGWrT1ja5FWepmr5Uzz4Y29mz53Pf8xCCltyW3s4ba65mfh8FnyyjYw0dszPw+ATjuUWVbhtbrbu+JXw2m0v4X/Elme1rfnYyUg4t/uOae8h2eXXUssatkvVwVsTs8nwQJ2bV/pFmPDlNieYy2511fVjx4+7VYMrt7eIiepYMAHIFsAAZAAAAAAAAAAAAAAAAAAGB8VErYIJJnrhsbVc72ImTUL9ctx+N9Jtg2h1n1P0BqKvVceTWqpmz4cMTl/qNQggbJdxSu8r3b7JBnK0dVVwL6szvk/tk5lYf2OO4dvsfvVuc7LqS+PeieDXwxY+lrizxkAAAAAAAAAAAAAHCmvDbnalse2jWFuXhVPqo+rZwphEbUNbOie7tFT3GxAhbbXsBtu0PUNTqWmv1TabtJSRQIiQMkgesauw56cnKqo5G8nJhE7+hY+TOZ2cBiapvzpTVGnjrH6o/McNViLWlHFV/Y/tS1Jsyuss9qRtdaql/HW2uaRWxzO5J2jHc+zkwmOLCouE4kXCYs9Z953ZXV0Cz3Kvudnna1VdTVVvke7knc6JHsX5yoOuNKX7RWpp9Pako0pq6JOJjmKroqiPulidhOJi+5UXKKiKioeVS0FTdKymtNGjlqLhURUcSJ3ulejET8YvWOyTL8z0v1dO/apnjHbxie/ihrGNv4ermp97Zfpy60990/br3SxTxU9fSx1MTJ2cMjWPajmo5MrhcKmUO+fjRU8NJRw0tOxGQwsbHG1OjWtTCJ8yH7HJK9nanZ4LRGum8AB8sgAAFR9+y0Ng1jpe/t4lWtoJ6KTwTsXpIz3r2z/mLcEf7c9mdPtQ01RWiW6yWqWjrUqo6mOnbK79rexWYVU5Kj8rz+ChL5FjaMFjqLtydKd8T3TDVxlmb1mqiOKiOktR3zSWoYNQacuElBcIfN42plkrFxmORvR7FwnJfBFTCoilsdn+9Boy52+OPWTJ9OXJrUSVyQvnpZXdMsexFc1F64eiY6ZXqVw2ubL9S7M7pHBeWsq7bUvVlFdKdipDMvNeB6Lns5Mc+FVVF58KrhcYU3KLhFwue46Ti8twGc2qbs7+qqOPd+kxuV23ir+CqmifKWxbQO0HSmvHV66UuTrlDQOjZPO2nkZHxvRVRrXPanEqImVx0yniZWQtuZ2b6m7Eqa4OV3aXquqK9UcnNG8XZM+dsTV95NJyzMrFrD4q5atTM00zpv7N0+9Z7NVVduKquMgANF6gAMikW+dRpSbcHzImFrbNSzr96+WP+yhDRYnfvtqQ6x0peUTnV2+ppFXHTsnsen6V3zFdTsmQXOcy2zV2aeU6fBUMxp2cRV2neWU3F9UJT3bUGi6iXDaprbpRtXll7USKZPWuOxXHqcVsPT0lf7npXVNs1LZnMSvts6TRI/0ZEVOF8bvtXsVzV7+ZsZrgox2ErsdMxu743w+MFf5i9FU8Gy8GN7N9aWTXukqTUdin44J04ZYnftlNKiJxxSJ3Obn3phUyioq5IcVuWq7Vc0VxpMcYXGmqKo1gOrdrhRWq2VNzuNTHS0dLE6aeaRcNjY1Mucq+CIhzdK+itdvnuFyrIKOjp2LJNPPIjI42p1Vzl5IhSjeO20T7Rax2n9Pvlg0jTyI5Vc1WPuT2rye9F5pEiplrF5quHO58KNlMnye9mV7Zp3Uxxnq/Xqhr4rFUYejaq4sH2ta1n2hbQLlqmRssVNMqQW+GTk6GlZngRUyuFcqueqZ6vUxQdDk7DZs0WbdNuiNIiNI8FOuXJuVzXV0uCzO4P/hDXad/Bbv95KzFmNwlyJdNcMVeborc5E9SLUp/WQ3Kf8Ku/wDX/KG/lXtMeK1YCA5AtYUu326FabbDQ1jY+FlbY4lV3xnxzSNX8VzC6JWPfytUj7dpG/NanZ09VUUMi475o2vZ9MK/OWLkre5vMqIn+qJj3fOGhmVG1h6uxVjn3jopz3nCnW1RXT3Krx9UNja2t2EdZ7nUUqJnmrHqk7V/85U9xOBTXcm1Oy07SLjpiofwxX+kSSDKquainRy8KdyZjc9V/m0LlIcg5S4WcPmNzduq+tHjx9+q44C7zmHpkABAtwMZ2paZj1ls8vumHoziuFFJFE53RsuMxuX2PRq+49We+WqDUFLp+WuibdKqCSohpUyr3RMVEc9UTo3LkTK4yvTJ6B626q7NdNyN0xvjwl8zpVEw1eR9p2adsx0cqZbIxyYVr0XDkVPFFRT6JO3odIrpPbJdVhYraG9/9a0y81RHSLidufFJEV2E6I9pGJ3DCYmjFWKL1HCqIn99ylYi1Nq5NHU4Jv3R9pLNI6xk0rd6hI7LfpkWF7vRp63CNaqrnk2RERqrz85GdEVVIRPmRjZGOY9vE1yYVF7z4x2Dt43D1WLnCfd1T4M4a/Ni5FcNoiArNu07eoqmCm0Xr+4oyuZiK23aodhtU3okUzl6Sp3OXk/v870rMnG8wy69l96bV6O6eiY64XCxfovUbVEgAU0XsFZd8Xam2Cjl2aWCpXyqpY1b1PE/HYwrzSnTHwnpzd4M5c+Plke8Pt5otIU9TprR9RBXaodmOWZER8Nt8XP7nS+EfcvN3JOF1Np5JZ6iaoqZ5qionkdLPNM9XySyOXLnucvNXKqqqqpeeTGQVV1xi8RGkRvpjrnr7urr7uMNmWPiinm6J3y/NWTSJ2NNG6SeRUjhYxObnuXhaiJ7VRDZVoSwxaX0VZdOQua9tsoYaXjamONWMRqu96oq+8pFuwaXXVW2mzpJGr6Ozqt2qVzjCxqiQpnx7VzFx3o1S+6ckHLXFxVdt4eOjfPjuj095k9nZom5PSAAoyZUF3o/4xGr/bRfqkRG6kj70X8YjWHtov1OIjhep2/K/YbH5Kf8YU3He01978a39xVH8y/81TZboH94un/uZTfommtKt/cVR/Mv/NU2W6A/eJp/7l036JpVOW/3Vnvn0hJ5Lxqe2ADnafAABD29/p5182J3GrhifJUWSeK6Rtb8WNVbKq+pInyL7ijy47uadxs4u1DS3S11dsromzUtXA+CeN3R7HtVrkX2oqmtK+Wer09fbjp6u51Vqq5aKR2FRHdm5URyZ7nNwqepTo/IrF7Vi5h540zrHdP6x71fzq1vpuR3OmfMzElidG5yojk6p3eC+5eZ9guyDidJ1bBdgWuG6/2YWu9SPRbjE3yS5sRUy2pjREeq46cXJ6J4PQz0oNu77TXbNNbOmr3yO07dOCG5sa3iWFU9CoRE5rw5VHInVqryVUQvnR1NPWUcNZRzxVFNPG2SKaJ6PZIxyZRzXJyVFRcoqdTkHKDKqsvxU7MfUq3x8vD00XDA4mMRaiemOL9gAQLcCJd77+Lzqbr6VF+uQEq1lTT0dJNV1dRFT00EbpJZZXo1kbGplXOcvJEREyqqUk3kNs8+0O4LYdPySQ6SpJUcjubXXKVq8pHJ1SJq82tXquHO58KNnuTuX3sVjaK6I+rRMTM9066d89DUxuIos2p2ulDq+koBwdfU1a7cN/e9q77pQ/oULKlatw7972rvujD+hQsqce5Sfid3w9IXHAezUdwfE0ccsL4pWNfG9qte1yZRyL1RT7BBxLca1dc6dl0hrW96Vm41+pda+CNz8cT4V86F648Y3MU8fvLF78WklotTWXW9LCiQ3GL6nVrkaiJ2zMvhcq9VVzFe32RtQrp3nbMpxsY3B273TMb++N0+9TMbZ5m9VS9nQl5TT2urFe3uRsVLWsSdV7oZPscn4r1X3FuJ4+ymkiXPmOVvzFLJWNljdG70XIrV95avZhf/AK5tA2i6yP4qpsXklZzyqTw4Y5V+UnC774js9sfYux3T6x8Vdzi1rRTcjo3fL4slABXUCHCr4dfA5OhqG8U+ndP3HUFXhYbdTun4VXHG9OUbPa56tT3n1TE1TERxlmmma6opp4yrvt3uqXXahcYo3tfBa4o7bGqeLE4pP/Me5PcYQFknme+eqkWWpme6WZ6rzdI5Vc5fnVQX2xaizbpojojRc6KIt0xRHRufL2yuRGU7HSTvVGRMamVc9y8LU96qhsj2a6ah0doGyaYh4F+p1FHDI5icnyImZH/fPVzveUw3WNIrqvbHbppo1dQWFPqpUKqLhZGriBufHjXjwvVI1L4IUHlrjdq5RhaZ4b5754e71WXJrOzRNyekABRU0AAyAAAAAAAAAAAAAAAAAAMCPt5Cu+p+wXXFRnHFZamHP84xY/7RqoNlu+9cfIN23UbEdwvq30tO331Ear+K1xrSEC5X7Gjc8S63s73ek2jqY09natd+Vhc418/sd918i24Vluc7DLjZpo2t8XsfG9PxWv8AnNgxkAAAAAAAAAAAAAAi3eG2gXrZpb9P6koaWGvtkly8iuNEqYkkbJG5zHxv6Nc1Y15KiovEiLjqSkRbvV2lbvsI1IkcSSTUUUdfGq/B7GRsjlT7xHp7zeyyLVWMt03Y1pmYifHc8r+1FuZp4u3JS7NtvOg4Kt8Ud0oeJyRyIqxVdBNhOJuU86N6csp0cmF85qpnAtN7r9osGvbPqOk1bcaijtlaysbRVlNG973sXLMysVvR2F9HuKtaM1XqLRl7+rGlrvNbapyI2ThRHxTtTo2SNfNenNevNM5RUUt1u1bXdUbTbjdaO9WW1UsVrpony1NHJInHJI5yMbwOzhOFj1Xzl6J48rVmGW5hk9mucNd/k9MTprGu7hPfxjf3IzD4qxi6o26frJvABR0wAAAAABj+0e6XSx6Dvt6ssVNLX2+hlq4Y6hrnRyLG1Xq1UaqLzRFRML1VDID8quCKqppaadiPilYrHtXo5qphUPu3VFNcVVRrESxVEzE6I22e620Ntz0FU0dTQwTJLC1t0s1WqOfAruaKi8stzzZI3HNEXzXIqJGup90u2VVa+TTetq600js4p6qibWKzPc1/Gx2E+2yvrKzW6a+6J1RIttuNRbb1ZqqWk8op3Yc10Uisc1UXk5q8PNrkVFTqhPmgN5bXNyutl01XWCy19dcK6nomVrHyQp9kejVe+NEcirjKrhUT1dxf7+T4/Lqpu5Zc/lzv0nTd57p7+KFpxeHvzsX6frQs9ouxw6Z0hZ9O08vbR2yhhpGycPD2nZsRvEqc8KuM+89cJ0Bz6uuquqaquMpqI0jSAAGGQAAQNvvWV9dsnpb1ExqrZrpDNK5eqQyosLsffSRr7imxsn15p6n1Xoy8aaq1RsVyo5aZXq3PArmqjXoni1cOT1oa2p6ero6mahuELoK2llfT1MburJY3K16L70U6VyMxUXMLVYnjTPun9dVczm1pXFzrfIAUuSFe5ojV+ptFXd910td5rdUS4SdqIj4Z2p0SSN3muxlcL1TK4VCWk3qdoqU6sWx6UdL3SdlUI38HtP6yCTg0MTleDxVW3etxM9fS2bWMvWo2aatzKtoe0PWev5WfXVe5KumjfxxUMLEhpY17l7NPSVMrhz1cqZ6mLHB2rTbbjebtR2e0Urqu410zYKWFvw3u6Z8ETqqryREVVNm3btYe3s0RFNMdW6HnVXcvV/WnWZZPoHRMmpNM6y1JUOqIbbpuzy1CSxYTtavh4oouac24RVdjmmW+Jh/5S7OrNDUGgN1PUmmaFUlkistRLV1CN4VqahzMySL7V5Ii5w1GpnkUmIrJsy/iE3rlP2Yq0ju04+PFt47DRh6aKenTeE9bj1yWn2pXm1uVrWV1mSRuV9J8MzeSe6VV9xAp7Gh9SXDR2srTqq1sSSqts/aJEruFJo3IrZI1XnjiYrkzjkuF7jczTCzi8HcsxxmN3fxj3vDB3YtXqap4NlQMe2f6ysGutNU1/wBO1qVFNM3z43YSWB/wo5G5816d6e9FVFRVyE4pct12q5orjSY4wudNUVRrARfvT6fdqHYff2RRsfU26NtzgV3wVgcj349axpInvJQPieKOaF8MzGyRvarXtcmUciphUVD1wl+cPfovU8aZifJ83KIrpmmelq/aqOajmrlqplF8UUHta803Lo3W960pNxqtrq3QxOfjifAvnQvXHjG5qnjHcrdym5RFdM6xMax3SpFyiaKppnodmzXKvst5ob1apUir7fUsqqZy+j2jFyiO8Wr0VO9FU2IbLdb2jaDo2j1JaH4bKnBUU6rl9LOiJxxP9aKvvRUVOSoa5TINBay1NoW9Pu2lrq+gnlRraiNzUkgqWovJska8lxlcOTDkyuFTJBZ/kkZnaiaZ0rp4T19k/D9W/l+O+jTMVcJbIzCNsO0qwbNNNLc7q/yitnyy326J6JNVyInRPisTKK568mp4qqItZq3el2kT0KwwWvS9LM5qtWobBM9Wr4tY6TGfaqoQ1f7xd9QXma9X+51N0uM/KSpqHZXGVXhanRjU54a1ERCs5dyOvTdirFzEUx0ROsz8o96TxGb24p/l75WM3RLjftb7Y9Y68v8AWpU1bbdFSua3KRxJNKr2xRpnzWMSHCJ1XOVVVVVW1JCm5tpiSx7ImXepi7Op1BVPuGHM4XJDhGQp60VrONP5wmsguUF6i5j64txpTTpTHhGnrq38HTVFmJr4zvQrvgaJfqfZi690MHaXTTjnVsaNTzn06pioYnP4qI/vVViRE6lJmuRyI5q5aqZRfFO42hOajmq1yIqKmFRe8197fNn0mznaJU2yCJzbLXK6rtD8LwpEq+dDlero3Ljqq8KsVepaORuZRNNWDrnfG+nu6Y+Pmi84w2ul6nxYEAC9q++Xta9qse1HNXkqKmUUkfZ5ts2i6IpoqG33aK6WyJvDHRXVjpmxpyREZIipI1ERMI3KtTuQjoHhiMLZxNGxepiqO162r9y1OtE6LJR7293bTIyXZ9RvqOHnIy8OazPjwrCq49WSP9e7ftpOrqV1D9UKawUD0VHxWlrmSyJno6Zyq9PvOHKKqLki0YI+xkOXWK9ui1GvbrPrMtmvMcRXGk1PmONkbEZG1GtTuQ5XGO5PH1Ay7Y7oafaJtCoNNMa9KJf74ukrVwsVKxU4ufc564YnrdnoiklfvUWLdV25OkRGstW1bqu1xTHGVoNzHRjrDs4l1NVwujrdSSNqGI9FRzaRmUgTH2yK+TKdUkaTqfnTQQ01PHT08TIYYmIyOONqNaxqJhERE6Iidx+hxPH4urGYmu/V/VP/AKjwhdbNqLVEUR0ABwppvRQbej/jD6wx40X6nERuZrt6uTLvtu1nXx+ilzWlz49hGyFfpYphR3DLaZpwdmmeMU0+kKXjaoqv1zHW/Gu/cVR/Mv8AzVNlezxzX6A069i5a61Uqovq7JprZc1HtWN3R6K1fYvI2Cbut4ZfNh+ka1nWO2x0smevHB9hd+NGpVuW1Ezh7VfREzHnH6JPJZ+tVDPwAc4WEAAAp1vsaOdadd0OtKaJfJL5ElNVORFVG1UTfNVV6JxxIiIn+iUuKYbtm0TT7QNnd003J2bKmVna0Mz0/aalnON+cKqJnkuPgucneS+R5h9AxtN2fszunun5cfBrYuxz9qaGu/3nB9zQ1NNUTUlbTyU1XTyuhqIHph0UrFVrmKncqKiofB2WJ1jVTJiaZ0lzzymCTtiu2vUuzRrbX2H1Z02r+JbfJJwyU2fSWB65RqKvnKxU4VXOOFXKpGIPDFYWzi7c2r1O1TP78HrZv12atqiV8NJ7f9leoIWKuqKez1Kty+mu6eSvj9Sud5i/euU/TVW3nZXp+lfK7VlFdZkaqsprU7yuSRfBODLUX5TkT1lC15pheaeC8zhEROTURqeCJgrH+i8Ft67dWnVu9dEp/Grmzpsxqk/bXtq1JtLV1t7JbNpxsiObbo5OJ9RhfNdO9PSwvNGJ5qLjPEqIpGK95wfrBTVE8VVLBC+VlJD5RUuanKKPjazid4JxPantVCzYbDWcJbi1ap2aY/fjKLu3bl+vaq3y/MDovNFyDZeK1u4b+9/V33Rh/QoWVK1bh373tXfdGH9ChZU49yk/E7vh6QuOA9mo7gAEG3GIbZNHRa92cXjTLlYyoqIeOjkdySKoYvFE7OFVE4kRFx8FVTvNdj2TxSPhqoH09TE90c0L24dHI1Va5qp3KioqG0Ipnvj7P105rNmtLdA1tqv0iMq+FERIa5E648JWpnv85r1X0kLvyOzKLdyrCVzuq3x39MeMeiHzfDbdHORxhBBJ27vqZtp1TNpyskRtHe1b2CuXkysamGezjblnrVGkZHy5Mpyc5qoqKjmrhWqnNFRe5UXmX/E2KcRam3V0qvdt03aJoq4SugvVeSp6lBGuy3anbr9SQ2rU1ZBb78xEYk8zkjgr0+Ojl5Mk8WrhFXmnXCSh5JVdUp5FTqjkb5q+/oUi/Yrw9excjSf3wVK/h7livZrj9X4ZxkhLeO1SyepptF0UqK2le2ruitXl2uPsUK/JRVevrc3vQyzabtQtemaea32GpprlqFU4W9k5JIKFfjyOTk56dzEzzTnjvrpI+SSWSaaaSaaV7pJZZHcT5HuVVc5yr1VVVVUmcowFU1RfuRpEcO3t+SWyzBVUzz1yNOr5/J8hzmtarnLhrUVVVe5ASPu7bPHbRNokFNWQcdhtasq7oqovDImcxwe17k5py81ruecE7isRbwtmq9cnSKY1T9i1N6uKI6VmN0fQ7tJ7MI7pX0/ZXXULkr50cmHRwqmII19jPOwvNHSOTuJkCA4pjMVXi79d+vjVOv6eHBdLVuLdEURwgABrPQAAAAAAAAAAAAAAAAAAAAGBWP8AZGbn5Nsds9ta7D629xqqZ6sjilVfxlYUDLi/sll14q/RVkY79riqquRvjxLGxq/iP+cp0ZEqbpV3+ou8Vo6qV/C2atWjX19vG6JE+d6G0Q086Wuklj1Par3FntLfWw1TMdcxvRyfkNwdPLHPBHPC9HxyNR7HJ0ciplFA+wAYAAGQAAAAAAAAPP1LbIb3p25WWp/aK+klpZPkyMVq/Qp6AUzTVNMxMdDExrGktXFKj2U0bJWq2RicD2qmFRzV4VRfeilw9xm0JS7O71e3wqyS43d8bHr8OGFjWN/HWUrXtmtT7Btb1fb3xoxkN1nqGtanJsUv2dv4siF193KyvsOw/SdBKitldQNqpEXqjp1WZUX1osmDpfKvGRVllE0//smJ8NNfkr+WWdMTVM9CQQAczWEAAAAAAABr83jrO2x7c9V0kUbmw1FUyvjz8Lt42veqertOM9PdRsy3jbxZHrE2SG2QVFxlR32rOyYqetHytX3GY789odTbQdPXxMcFwtklJhE+HBLx5X2pP+KeruIWTiuGrNSSxLljae3wSY9Syyp9MR1G5j//AMf57XfNGz4/ZlW6bH/yGz26rVAA5csgAAAAMAU23ydAOsGs49b2+DFrvrkjrOFOUNa1uEXuwkjUTx85js+khck8rV2n7TqrTddp6+Uraq310SxTRr86ORe5yKiOReqKiL3Erk2ZVZdiouxw4THXH73tbFYeMRbmiWtHvOTL9rezu/bNdTOtN3R09FM5y224tbiOrjTx+LImURzPemUVFMP59FOx2L9u/bi5bnWmeEqfdtVWqppqjeA5Pl7msYrnORrU5qqrhEPV5xvcuVrWq97ka1Eyqr0RC3m6LsnWw21uvtRUaNvFwixbYJU86jpXJ6Sp8GSROa96NwnLLkMI3ZNh09+qqXW+taB0Vmick1tts7MOrXJzbNK1ekSdWsX01wq+bhH2+KBypz6mqJweHn80x/j8/LrWLK8DNH82vj0MY2sW6S77LtVWuBnHNV2erhib4vdC5G/Tg1wUz0kp4pE+Exq/QbRlRFTCpk1vbStKVGiNf3rS88T446Oqe6kV/wDjKV6q6F6L3+auFx0Vqp3GeROIp/m2J47pj0n4MZ1bmaaa2OnzI9rGK97ka1EyqqfSqjUVzlRrUTKqq4RCxu6psXmulXRbQtW0ixW6JUms1DKzzqh3VtTIi9GJ1YnVy4dyTh4rhmOYWcBYm9dnujrnqhD4XDVYivZpZxui7KKrSFrqNZ6hpnU18vEDY4aZ6cL6Slyjka9P5R6o1yovNvC1OS8SE/AHHMfjbmOv1X7vGfdHRC4WrVNqiKKegABpvRWDfd0G58FDtGt0Kr5M1tDdkan+KVy9jMvyXuVirzXD29zSriphcY6Gzi7UFFdbZVWy400dTR1cLoKiGRMtkY5FRzV9SoqlB9uOyy67MNQLG9JqvTtVIqW24q3KIi80glXokqInJej0TKc8onR+SWcU3LUYO7P1qfs9sdXfHp3K/m2DnXnqI70fgfQoLsgnBk+yzRVZtC13QaWpUkZDMva3Cdn/AMPSNVO0fnC4VfQbn4TkPEstqul8vNJZbLQTXC51j+Cmpok86Re9VVeTWomVVy4RERVVS9+wHZdR7MtJuppJY6u+V6tlulWxPNe9EXhjZnn2bMqiZ5rlVVEzhIHPs5oy6xMUz/Mq4R8fD3z4pLLsHN+van7MJBt9JTW+gp6CigZT0tNE2GGKNMNjY1MNaidyIiIh+4ByGZmZ1lawwTbhs7otpOiJrNI6OnuMDvKLbVuRfsE6IuM46sciq1yc+S56omM7B62L9zD3KbtudKo3w+a6Yrpmmrg1jXa23GzXars15o5KK5UMqw1VO/rG9PX0VqphUVOSoqKnU6xeTeG2MUW0iibdLXJDQaopI+CCoei9nUsTmkMuOeM9Hc1avinIpPfrRdtP3qosl+t89tudP+2006YcifGavR7Vxyc1VRTr2TZzZzO1rG6uOMfGOz9yqeNwNWHq1j7LpgICZaAAfvbqOtud0p7Xa6Kor7hUu4YKanYr5JF9TU7u9V6InNTEzERrLMUzVOkPxhinqKiKmpYJqmpnkbFBBExXPlkcuGsaic1VV5IhfPdz2Zs2b6ISKtax9/uXDPdJW4VGuRPNhaqfAYiqnVcqrl78JjO7hsNi0MkeqdVNgqtUyMVIomqj4raxyYVrF+FKqcnPTuVWt5ZV06HNOU2f04ufo2Hn6kcZ65+Ue+e6Fny7AcxG3X9r0AAU5KhjG1TWFFoTQN21RW4clHCqwRZXM8zvNjjTCKvnPVqZxyTK9EMnKlb8N5u1Zq+x6Xipqt9so6P6ou7KCR7ZKh7nxtzwoqZY1rsfzikrk2BjHYyi1V9njPdHz4NfFXuZtTVCuKyVEz3z1crpqmZ7pZ5HLzfI5Vc5V9qqpyfslFW91uuP+xTf8Jz5FX/9m3L/AGGb/hOzRVTop00VzOsxPk/AsluRa4io7nctn1fIjErXOuNrVV9KRGok8SZXrhrZERE7pFK6eQ3Dutly/wBhm/4TsWyW/WW7UV6tdHcYa+3zMq6Z60MuEkjXiTPm82rhUVO9FU0M0wdGPwtViZ48Oyej99TZwdyvD3Yq0nRsz7gedpi6Je9N2y8tgkp219JFUpFIio6PjYjuFUXCoqZweicVqpmmZpnjC3xOsagAPlkCoABVDfJ2YvpK520uyU2aaZGx3yJn+LcmGsqceC8mP+9dj0lK2c+Zs9q6eCrppaWqgjngmYscsUjUcx7VTCtVF5KipyVFKZbfNgV00dUVOoNHU1Rc9NKqySUkaLJUW5O9MdZIk7l5uanXKJxHROTGf0VURhMRVpMbqZnpjq746Ovhx4wOZ4CZnnbcd6DwfMbmSRpJG5Hsd0ci8jkvKA4ByA1HueyNjHySSPRkcbGq5z3KuEa1qc1VV6Ig1IiZ3Q5jjlmmjhghlnnmkbHDDExXPle5cNY1qc1cqqiIiFlNX7Mk2b7oep1uDGrqG6uoJro9rkVI8VcPBA1U5cLEVUz3uVy5wqYyTdf2ITabki1trOlRl7c1fqfQPwvkLFTCvfjl2zkzy+Ai46quMu3wf4u+pvl0X65AUfMM9pxOY2MLYnWiK6dZ652o3d0e+e5YsHgeZtVXK+Mx5KML6S58VBy70l9pwXhXVrdw7972rvulD+hQsqVq3Dl/9ntXfdKH9ChZU49yk/E7vh6QuOA9mo7gAEG3A8TXWmbXrLSVx01eolkoq6FY38PpMXq17V7nNciORfFEPbB90V1W6orpnSY3wxMRMaS1sa70reNE6srdMX2NW1lK7LJUbhlVEq+ZMzxa76FRUXminhl/9umyy17TtONppZG0V4o+J9ur+DiWJy9WOT4UbsJlPUipzQovrDTV90hqCSw6lt0lvr2c2NdzjnZ/KRP6PavinNOioioqHXMizu3mVqImdLkcY+MdnoquPwNVirap+zLx3Na5qtc1rmr1RUyinCRNSNYkV6Rr1YkjuH5s4PpDknkdEzHB8sY1jUaxrWNTojUwhycnf03Y7zqW+w2LT1tnuVym5sgiT0W973uXkxiZTLl5HzXXTRE1VTpEM001VzpG+XxYrRdL/fKKx2WjkrLlXS9lTQMT0nd6qvc1qZc5y8kRFU2B7G9AW7ZxoimsFG/t6ly9vX1aphamocicT8dyckRqdzUTquVXH9gGx627M7S6rqpIrjqasjRtdXNb5rG5z2MOeaRouMquFeqZXHmtbKZy/lJn0Y+rmLM/y4//AKnr7urz7rTl+B+j07VX2pAAVRJAAMgAAAAAAAAAAAAAAAAADAAADXd+yB3f6o7fVoEflLXaqemVvg53FN+SVPoK8EhbyN6S/wC3jWdya/jZ9VZYGO+MyFeyaqerDEI9MgbW93e+fXHsO0ddlfxvfaoYZXeMkSdk9fwmKapDYR+x6X9LnsRqLM9+ZLPdJYmtz0jkRsqL73Ok+YCyAAMAADIAAAAAAAAAADCNYbJdnOrrvLd9RaToK6vma1stQvGyR6NTDeJWKmcIiJz7kRO5DNIYo4YWQxMayNjUa1rUwjUToiH2D1rv3blMUV1TMRwiZ4d3U+YopidYgAB4voABkAAAABgY9rnROldcUENDquy010ggk7SJJeJFjcqYVUc1UVMp6+fLwPvQ+j9N6Jsy2fS1qitlC6Z07oo3OdxSOREVyq5VVVw1E5r0RE7j3ge30i7Nvmtqdnq13eXB87FOu1pvAAeL6AAZAAAAAYHmansFl1NZZ7Nf7bT3G31CYkgnZxNXwVO9FTqjkwqLzRUK7au3TqOSofPpHVs9FEuVbSXKn8oa1c9EkarXIietHL6yzgJHA5ri8BP8ivSJ6OMeUvC9hrV77caqeUu6lrh9U1tVqnTkFPnzpIoZ5Hp7GrwovzktbMd3LRWkauG6XZ82p7rDh0cldG1tPE5M+cyBOSL05uVyoqIqYJpBuYrlJmOJo2KrmkT1REem95WsBYtTrFO8wACCbgYPtV2W6R2kUsLdQUcrKynRW01fSydnUQoq5VEdhUVv2rkcnPOMmcA9bF+5Yri5aqmJjph81UU1xpVGsIQ0luybPLJdoblXz3fUDoXI+OnuMsa0/EioqK6NjG8eMdHKrV70JvwAe2Lx2IxlUVX65qmOt827VFqNKI0AAaj0AAAOpd7ZbrxbZ7ZdaGmrqKobwTU9REkkcidcK1eSnbBmmqaZ1gmNVddYbqWmK6oWbSmorjp9rnZWmmjSsgYmOjEcrXp73uPDt26PKlUx1y2hufTovnspbS2ORyepzpHI1fvVLTAnKOUuZ0UbEXfOImfOY1adWAw9U6zTDDNl+zHSGzqhfBp23r5TM3hqK+pd2lTP6nP7k5ei1Eb34yZmAQ969cv1zcu1TMz0y2qaYpjSmAAHk+gAADHNd6H0pri2toNU2WmuMTMrE96K2WFVxlY5G4cxeSZ4VTOOZkYPu3crtVRXROkx0xxYmmKo0lV/VG6ZEszpdKazmhix5tNdKVJuf86xWrj2tVfWYk7dX2kcWG3rSKt8VmqUX5uz/rLmgsFrlXmdunSa4nviGjXlmHqnXZVZ0xumTLKyXVOtl7Lnx09qpEY5fZLIrsfgE97PNnWjtA0joNMWSCjkkTE9U5VkqJvlyOy5UymcZwnciGVg0MbnONx0bN65Mx1cI8o4+L3s4W1Z+xSAAi2wAAAMAAMAAAAAAAAAAAAAAAAivaNsF2eazqpbhJbpbNdJVV0lba3pC6R2VVVezCseqqvNVbxL4kP3fdMvsLlWz63t9W1V5Nrre6JUT5Ub1RfmQtoCZwvKDMMLTs0XJ06p3+vBq3cFYu76qd6olr3TtVSzYuusrNSRY9KkopJ3/M9zUJv2U7EtE7PZ2XKjp5rpemtVPqlXqj5GZTCpE1ERsaLzTzUzhVRVVCTAMZn+PxlGxcufV6o0j0Ys4KzanWmneESb4H8XfU3yqL9cgJbIk3wP4u+pvlUf65Aa+UfiFj89PrD3vfd1dyjLvSX2nByvpL7Tg7Yo0LW7h3739XfdKH9ChZUrXuHfve1d90of0KFlDj3KT8Tu+HpC44D2ajuAAQbcAAAPE1npPTesbQ606ns1LdKNVyjJm+dG7GOJjk85jvtmqi+s9sH1Rcqt1RVROkx0wxMRMaSrLqzdOoZZnTaU1fVUbFyqU1ypkqURe5EkarHIietHL6zD13Vtona4bftJ9nn0+Op4sePDwf1lygWC1yqzK3Ts7evfENGvLcPVOuzoq/pfdMiSZsuq9aVE8Sp51Na6VIOf869XLj2NRfWWA0NovS+ibX9TdL2WmtsCrmRY0VZJV8XvXLnrzXm5VMgBH43N8Zjt1+5Mx1cI8obFnDWrP2KdAAEa9wAAAAZAAAAAAAAAAAAAAAAAAGAU8/Ul0hsmnbneqjHY0FJLVSZ+LGxXL9CHoEQ74l+TT+7vqmVr+GathZQRpn0u2e1jk/AV6+4DWXW1M1ZWTVdQ9XzTyOkkcve5y5VfnU/EAyBa79jd1D5JrzUumJJMMuNvZVxoq8lfA/hwnrVsyr976iqJJ26zqT61tvukri+TggmrUoplXpwTosWV9SK9F9wG0oAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYAAAAAAAAAAAAAAABkAAYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAiTfA/i76m+VRfrkBLZEm+B/F41N8qi/XICRyj8Qsfnp9YeV77uruUZd6S+04DvSX2g7Yo8LXbh373tXL/rKH9ChZQrXuHfvd1b90of0KFlDj3KT8Tu98ekLjgPZqO4ABBtwAAAAAAAAAAAAAAAZAAAAAAAAAAAAAAAAAAGAAABSpH7JNqHsNJ6V0tHJ51ZWy10rUX4MTOBufUqzO/B9Rbc1z7++pPq3t7qLbHJxQ2Shho0RF5cbkWV6+37IiL8kQK/gAyB+lPNLT1EdRA90csT0ex7erXIuUVD8wBt62d6hi1ZoOxalhVvDc7fDVKifBc9iK5vuXKe494rt+x/6p+rew9bHLJxVFhrpKdEVcr2Ui9qxfZl0iJ8ksSYAAGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYAAAAAAAAAAACJN8Bf/AHeNS9/n0X65AS2RJvgfxeNTfLov1yAkco9vsfnp9YeV/wC7q7lGXekvtODl3pL7ThTtijrW7hv73tXfdGH9ChZUrXuHfve1d90of0KFlDj3KT8Tu+HpC44D2ajuAAQbcAAZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgAAB+dVPFTU0tTUSNjhiYr5Hu6NaiZVV9xqI2g6gl1Xrq+6lmV3Hc6+aqwvwUe9VRvuRUT3GyTe21T9aewHU9XHJwVNbT/U6nwuFV068DsetGK933pq+EAADIAACzH7Hjqz6j7XK/TE0vDBf6FUjbn0p4Mvb+IsxsANQ+zfUs+jtfWLVFPxK+2V0VS5rfhsa5ONn3zcp7zblRVMFbRwVlLK2annjbLFI3o9jkyip7UVAP2AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgAAZAAAAAYAAACJN8D+LxqX5dF+uQEtkS73/APF41NlM+fRfrkBI5R7fY/PT6w8r33dXcow70l9pwcr6S+04O2KPC1u4d+97V33Sh/QoWVK1bh373tXL/rGH9ChZU49yk/E7vh6QuOA9mo7gAEG3AAGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYAAAUy/ZJdWZdpfQ8MvTjulUzPtihXH9MU0JN3o9XfXpt01NdYpe0pIKnyGkVF83soU7NFT1OVrn/AHxGRkAAAAAA2W7lWsfrt2CWiKaXjrbI51rnyvPEeFiX2dm5ie1FNaRaD9jv1p9R9ply0dUzcNNfqXtIEVf/AIiDLkRPbGsmfktAvwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwAAMgAAAAAAAwAAMgAAAAAAAwBEm+Av/u8al+XRfrkBLZEm+Amd3jUvyqL9cgJHKPxCx+en1h5Xvu6u5Rl3pL7Tg5d6S+04O2KMtduHfvd1b90of0KFlCte4d+9zVv3Sh/QoWUOPcpPxO73x6QuWA9mo7gAEG3AAGQAAAAGAABkAAAAAAAAAAAAAAAAAAAAAAAGAMG29awTQmyHUepmypHU01G5lIucL5RJ9jix44e5q+xFM5Kc/skGtOCj07oCll86Vy3Stai8+FMxwovqVVlXH2rQKXKqqqqqqqrzVVOADIAAAAAB7Oh9Q1uktY2jU1vX++bZWR1LEzhHcLkVWr6lTKL6lPGAG4rT91or7YqC922XtqKvpo6mnf8AGje1HNX5lQ7xW39j+1z9cWyafStXMjq7TlR2caKvNaaVVdGvud2jfUiNLJAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgAAAABkAAAAAAAAAAAAAAAAAAAABgAAAIl3vue7zqVPF9En/ANZAS0Qnvp3FtHsTmo3LzuVzpKVqY64k7ZfoiUksmpmrMLER/dT7piXjiatm1VPYpU/03e04C5VVU4O1qQthuHJ/7N6t+6cX6FpZMrDuF1XFR60oV6x1NJN7nxPT/wC2pZ44/wApYmMzu69npC45fOuGo7gAEE3AAGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYHDnI1qucqIic1VeiGqfeG1qu0DbDqHUkcqyUclSsFCvd5PH5kap4ZROJfW5S/W99rn6xdh14qKeZI7jdU+plFhcOR0qKj3J8mNHqi+KIaxhAAAyAAAAAAAAJh3PteJoPbdapqqbs7Zdv+rK3K4ajZFTgevyZEYue5OLxNm5poRVRcouFQ2l7sev02j7HLPfJ5u0uUDPIrllcr5RGiIrl9bmq1/wB+BJgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMAD8quogpKd9RUzRwwsTLnvciNT2qYrNtBsbJ+COKvqI/wCVjg81fYiqir8x7WsPdu/YpmWricdh8LpF6uKdWXg69vrKa4UUVZRzNmglbxMe3op2DymJidJbNNUVRFVM6xIAAyAAAAAAAAAAAAAAAAAAAAAAAMAVT369RNlu2mdIwyOzTskulS3HLLsxQ8/H9u5ezxLRXe4UdqtdVc7hUMpqOkhfPUSv6RxsRXOcvqREVTXHr/VNXrfW121ZWsdG+5TccUS4+wwNThij5eDETK96qq95beSGAm9i5vz9miPfO6PdrPki82vxbs7HTLw06ADvOoKsnPcnvrbbtYrrLLKjIr1bHdm3vfNA7ian9G+VfcXPNaGk7/XaV1TatS2xvHV2uqbUMj4sJKiZR8ar3I5iub7zY5pS+2zU2m7fqCz1CVFBXwNngenXCp0Xwci5RU7lRUOa8ssFVbxNOJiN1UaeMfpp5Ss+UXors7HTD0wAUxLAAAAAyAAAAAAAAAAAAAAAAAAAAAAAAAAMAAY1tQ1bRaE2f3rVtfwrFbaV0rWKuO0k6Rx/fPVrfeBSD9kB14motqdPpKjm4qHTkPBKjV811VKiOf7eFqMb6lRxWs7l7uVbebzW3e4zOnra6ofUVEjur5HuVzl96qp0zIAAAAAAAAAAAWX3ANof1t7S6jRtfPw2/UTESHiXkyrjRVZ7OJvE31rwFaDsWytq7bcqW40E76erpZmTwSs9KORqo5rk9aKiKBuQBh+xjW9JtF2aWXVtLwNfWQIlTE1f2mdvmyM9zkXGeqYXvMwMAADIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwB0L9d6Oy259bWvVGJyYxvN0ju5rU71U7FfV09DRy1lXK2KCFqve9e5EIkuFXctYajjbC1WvfltNE70aePve719FX3Ihv4LCc/VNVW6mOMofN8z+h0RRbjW5Vwj4/LrnxftI++a3vXZYa2ONeLs8qsNK3uV3xn/wD9TCGYfWRZaey1ELmPmqnRL/fUjvPR2OTk7m4XuT35Pc0/aKSy2yOhpG8k85719KR69XL61/8Awdypa99PIyPHG5io3PTOOWT0v4+qaoos/Vpjg8MFklFFE3MV9e5Vxmd+nZHz8tI3MM2OSSSadqlcq8Hlaq1O5FVjFdj3qpm54+jLN9QdN0dsc5r5o2Zme3o6ReblT1Z6eo9g1sZcpuX66qeGrfyqxXh8Hbt3OMRGvy8OAADVSAAAAB8NlidK+JsjFkZhXNRyZbnplO4zoxrEPsHCuajkaqpleaJk5DIADAAAAAAAAAAAAAAMa2m6QpNd6JuGla+tq6KmrkjSSWlVqSIjZGvwnEipheHCoqdFUhxd0/SSrldW6m5+un/5RYgEhhM1xeDomixXNMceh43MPauzrXTqrt/cm6R/zs1P89P/AMof3J2kP87NUfhU/wDyixINr/UWZ/8Amn3fJ5fQcP8A2Qrv/cnaQ/zs1Rn5VN/yiWdkugqHZzpVdO226XK4UvlMlQx1a9jnR8eFVreBrURuUV2MdXOXvMvBr4rN8Zi6ObvXJmPB6W8NatTrRTpIeRrC7PstgnroWxvnRWshbJnhc9zkRM49ufceuRzteuX980dtaqq2FjqqVqd682sT8/6DwwNjnr9NM8Olp5xi5wmDruROk8I753e7j4Mq0Xe5L7Zlqp4WRTxzOhkRiqrVVMc0z3Kioe4eTpC2fUnTtJRuTEqM45l8ZHc3fSuPceseWI2Odq5vhrubOCi7GHo56fraRr3gAPFtAAAAAAAAAAAAAAADIAAAAAAAMAAAOFKZfsi20Lifadmtvn9HFxuaNXv5pDGv4z1T1sUt3qq+W7TOmrlqG7TJDQ26mfUzv7+FiKqoniq9ETvVUQ1MbQtUXDWut7vqq6OzVXOqdO5ucpG1eTWJ6mtRrU9SIIHggAyAAAAAAAAAAAAAC1X7HttJ+o2r63Z3cqjhor1mooOJ3JlWxvnNTuTjYnzxtROpe405Wa5Vtnu9HdrbUPpq2inZUU8zFw6ORjkc1yexUQ2t7Ftd0W0jZtaNW0fAx9VFw1ULV/aKhvKRnjhHIuM9UVF7zAzIAAAAZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPM1RdG2ax1NeqI6RjcRNX4Ui8mp86p7j6oomuqKaeMvO7dptUTcrnSIjWfBgu1C+JU1/1Jhk/vWjVH1Kp8OTqjfvU5+1U8DJtnthdarWtVVx8NfVojpUXrG34LPd1X1qvghhegbS676hSWqVZYaR3lFQ53+Mlcqq1F9+Xe5CWyWx9cWLdOFt+P79/krOS2asbfrzG9HHdTHV/wCuHfr1gAIZagH51M8NNTyVFRKyKKNque9y4RqJ1VTEazaJZopWeTQVlVD1kkZHw8Kex2FU97OHu3vu6dWpicdh8Lpz1cRr++DMgeHetVWe10rJpahZ3yNR0cEGHSuRUyi8PcmOeVwh2Zb9aorJFeZqtsdHMxro3uRcuymURE6qvqTmY+j3dInZnfwfX0zD7VVO3GtMazv4R2vTB0LJdqC80flVvqO2jR3C7LVarXeCovNFO+edVNVE7NUaS9rdyi5TFdE6xPTDzdSXWOzWaor5MOVjcRsz6b15Nb8/0ZIu0dd6K26nlut3qpu1lY5skkbFdxOeqKqvx0amEwndnl0PT2j3Oa66gis1F9kSmekbWp0fUP5fM1F+l3gZHVXKyaHttDQ1EUksjo143QxIrlRuOKR3Pplfp5E3YtxYsRTszNVzojjp7/33Kli784zGTciuKLdnpmNY2uHXHdx6O1i9VUXC/bR3Ptj39rSTtjZI1V4YYWqnGq92Hedy78ohKp8sRuOJqInFzVcdTwdUart9jd2DmvqqxW8SQRKmWp4uVeTU+n1GjduVYuqm3bp4RomMPYt5ZRcvX7mu1OszwjuiN+/18GQAxrRmq4tRvqYVo30s9OjXqnGj2va7KIqORE70VOhhN/ram26rW+W6omalRM/hY6RXJI2N6RuRU6cDlzjwxlBay+uu5Vaq3TDGJzqzasUYi3G1RM92kcJnSY6J3abuKWwYRf8AX0NDcXU9DRpWQwOVKiVZOHp1RnJeJU8eSdxkN51Da7TQxVdXOuJkzDGxvE+Tlnkn9fRDxqwd6nZ1p+1wbVvNMJc29K4+px6vPp8HrA87T15o75QeWUSvRqOVj2SNw9jk7lT3p85xf73b7HStnr5XN43cMbGN4nvX1In5eh5czXt83pv6mx9Js81z21Gzx16HpAxO2a9s1TFUPqknoeyTiakrcrI3OPN4c5X1HtafvVBfKJaqgkc5rXKx7Xt4XMd4Knswp6XMLetRM10zGjysZhhsRMRbriZno6fLi9IGMak1nb7RUvo4opK2rZ6bI1RrY/U5y8kX1JlTtaLv02oLZLWS0C0iMmWNnn8TZEREXiauEynPHtRRVhbtNvnZjSlijMcNXf8Ao9NWtXVGvRx38HugGN6o1dQWWbyRrHVlbhFWGNyIjE+2cvo58Op52rNd2rZojWXtiMTaw1HOXatIZIDH9M6oprva6uumhdRJRuVJ0e7iRqI3i4kcnVMH6T6rsMVpiua17X08r1ZHwMcrnuTqiNxnl7D7nDXoqmnZnXg8qcww1VuLkVxpMa7503cNd/bu73uA6VmulDd6JtZb50lhVVaq4VFa5OqKi80U7p5VUzTOlUaS2qK6blMVUTrE9MABiOpdb0tsrH0NFTLXVEa8Mru04I43fFV2FyvqToelmxcv1bNEay8MVjLOEo271WkfvoZcRLGn1xbR+L04X1nF4p2UPT3KrfxjM11MlVoSrvjIX00rI5Gdmrs8MiLwphe9M4PA2QUH9811e5P2pjKVi+v0n/2CSwlFWGtXblW6Y3eP70QOZXacfiMNYonWmfrz3dHnGsJGAOHORrVc5URqJlVXuIhZ3IMIvG0WgpqtYbfRS3BjVVFlbIjGvVOvBnKu9vT1mX2yshuNupq+mVVhqImyx5TC8LkymT3u4W7apiqunSJaeHzDDYmuq3ar1mnj++nwdgHTvFzorRROrK+dsMTVwirzVy9yInVV9SGL2/aHbam6xUctFVU0c0jY2TSK3krlwnEiLlqKuEz6xawt67TNVFOsQYjMMNh64t3a4iZ6P3w8WaA69xrILfQT1tU5WQwMV71RMrhPV3nU07e6K+0Lqui7VGskWN7ZWcLmuREXmnsVDyi3XNM1xG6Ol7zftxci1NUbU79OnR6YPyhngmc9IZo5FjdwvRrkVWr4L4KfU8scEL5ppGRxsRXOe5cI1E6qqnzpOuj02o0113PsGCV+0mhhqVSlt1RU0res/GjOJE6q1q81THPngzenlZPBHNGuWSNRzV8UVMoe17DXbMRNynTVqYXMMNiqqqbNeuzx/fT4P0AB4twAAAAGAAAAAADg5PJ1fqC26V0vctR3ibsaC20z6idydeFqZwid6r0RO9VRAKtfsh+0nyOzW/ZnbKjE9dw1t14XejC132KNflPRXqn2jfEpEZDtJ1bctda5u+rLs7++rjUOlVnFlImdGRp6mtRrU9SGPGQAAAAAAAAAAAAAAAALMbhG0761tfyaHulRwWrUT0SmVy+bFWImGf0ieZ7UYVnPuCaWnnjnglfFLG5Hsexytc1yLlFRU6Ki94G5UEYbsu0yHajssobzNIz6sUn96XWNMIqTtRPPx3NemHJ7VTuUk8wAAMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEa7V7mktzp7Yj/sVKzyiZM8uN2Ubn2Nyv3yElKQ2rPrg1u+JVV7Ku4OR381GvP8AFZj3kplVuOcm5Vwpj9+7VXOUt6uLFFijjXMR+/HRIez+1rbNNwdq3FTU/wB8T+pXJyT3JhPcZCE6Ajrtybtc1z0pzDWKcPZptUcKY0DhVwcPe1jFe9yNa1MqqrhETxI61LqGs1PXN0/pviWCXKSz9Ekb3rnujTvXq7onXn64bDVX6t26I4z0Q8Mfj7eDoiZ31TwpjjMuvrS+S6kucNhtD0dTOmSPjz5s8mevrY3Cr61T1Jnp69tdNZKuhp4ZWLElIuW/4xXI5VdI75WeXswftfKddF6gtlTTQMkgipeFskiK1skvnI9VVOi4VFRPDknQ7OjrPW6gvv1wXVrnU7X9qjntVEnenoo1PiN+lUTrzJ6iqizRTcpnS3ET3zP7/enCl3aLuLvV2blOt+qqI7KaY0ny9dNeOmvkXXT81nstPXVsiR1dXMjWUzUTzG8KqquXqruSdOSes69vp7hepae3Qqj46GFyt4uTIG81Vy/bOXl/+Mns7TauSs1XFb4WPkfTxNZHGielJKueXuRqZ9pnGlbDBZbP5J5sk83nVMuP2xy9fcnRE8D5uY2bWHpuV76qt8dn7h6WMppxONrs2d1ujdM9cx0eNUeGncw3ZHUql1r2cXDFLSxzKir0VFVM/Mv0GQ3LXVohpe0o1lqVlgc+nkaz7G96KqcCr1Rc8+adOZgE1kvluuE1pho66SV6LC10LVSOeLPJVcnJGryzleXM+9V2yCxTUlE1Vlq20qSVisXzHOV2Go1PHkqetEQ+ruFsX7+3M67XCI6ojjPu/fD4w2ZYzB4ObdNOzsa6zPXM8I7t8+/v+NNXSmtV3+qVdDPXVaI5YYo0TL5XqvE9V9irjCL1OzXzTaq1lTsqIH06TSxwdgq5WONmXPRfXyd85ImjrFBZ7TAj4I/LnM4qiXhTiVy81bnwTp7iP7q6p0vriSvngV7UqpJ4VcqtZKyRHZTix1TiVPd6xaxNF+9XzcfWiJ0nXj3QzicDeweFtRfq/lzVE1Rpw754zOm7q3dMs/1rfUsVo44eF1bOvBTMXx73Kngic/mTvI7tdNRQ0zr/AKic+pbM9X01Iq/ZK1+eb3faZ93uwi9fVV3qrrcIq25QLSNfD/esLsp9jVeqZ6qq9+E6IexpfSNfeZ2V117emo0aiN7RcTStROSIi+gxPn9XefNmzRhMPrXOkzx6+6P384zisXdzPG7Nmnain7MTwj/lVHpE9ndPd0Jeau5aynfNSUrEmouFOyaqdkyNycLUXvTz1/8A+cjwNb1UMuop2wNRtNRMSliRqdVRcux4rxLj14O8mn9V2O61P1LglkRY3xsqYnMVHRquU5OXKP5J7+h3tB6TrHXOO43akkpoaZeOGKVUV8knx3c15J1581Xn3c/raw9qub8VRppGkRP77vPVibeNxVqnB1UVbW1MzMxu8+mOnyiGOXi1PtUVHT1eUrKmJ00sXdExVRGt9bvSyftDba+42au1DVyP8npYEZA53WXhwiNb4MTx7196kmah0zab7NDNXxSLJEnCjo5FYrmZyrHKnVP/AO952Lxaoq3T1TaIUZTxyU6wx8LcNZywnLwTlyNb+LRNNPXrvnqjXob08mqoruTr9XT6sa75nTdr47/HojcxnZOrYrLdJ5HtZGlauVcuEREiZlVPFt/FrXW61M6Ktvp042tX+SRfMbjxevnL6kwfhb9P6teyayLTyUtJNJxzuc9vZKqJjKKnnKi4Tl851NN6idp6jubGU39+zPazikciNhVqKio5OuUVV5d5sczrVcuWpiap000nhE7v3+rR+lbNuxYxNM0W6dZnWJ+tMb9NOrh59jz7rFUx3Orp6yNzalKh7XMb61yiNx3Kipg93SupLfZrrdq6oifT09UkfY0sEfF5zcpjuwuMZVcZPZ2dadn8oXUN2ZIs8mXU7ZfT870pXJ3KvRE7k9vLwbXeoLNqK91FTbfK5H1EjY25anZubI/vXoioqc08D0ruUX4rtRG1pEa6Tp0xw8nlZsXMHVaxNVWxtTOmsazEaTpM6aTMzrO7umep5V+Y6ZlXeIKGspaWsnesa1PDl0jsuXh71T6E6ZJqo0kSjhSZjGSJG3jaz0WrjmieojjTlPcdZaiiu9050FI7LGtTESuReTGeKZRFc7vwieyTSMzS79i3PGOPZ2a+qe5O4fSbmIjXZq001jTXTXWdOiJmdzxdZ3r6h2R9SxEdUyL2VO1eivXvX1JzVfYRbTQyViLRUUElzvNY7tZ5HdKduc816I53Vzl6JyQzLa5SVUlLQV0THvp6Z0iTcKZ4OJEw5U8EwqKvdkxu0aor4aCO06et1LHMrUR0lOx00sju96p0RflZQ28vtzThortxrMzv36RGnX06dOnwR2dX4uY6bV+ZimI3REazOvHTo1nhrPDTdve3fqWHS+hJLQlQ2W4XF/2VyfCVccaonc1Gpwp7u9TD4oaiqhndEmYaGF0z1Xk1iKvP75y/k9R+l6pLpS3VkN1ZUTXKoY1WI9/aOkyuEai9OS8sJyTJJFu0sym0XU2fialXVwu7ebuWVU/NTkiJ4J6z1qvU4S3FVVW1NU66+s90Q1qMLczK/VTRRsU26dNOnpmIntmd8/8AqXn7Ild9S7k1V5JWcv6NmTNUkYsjo0e1XtRFc1F5oi5x+RSJbE7V9mqZ6C32yobPMqcbJIFdGjk5cbX5RqJ61Xw5Eh6Ss8tot8nllS6ruFVIs9XMq54pFREwng1ERERPURuY2aYuVXJqjfppEfvgn8hxVdVmixFuY2ddZndHHdEdc+nT0OjtDvslotDYaR/DW1irHE7+Tb8J/u5Y9aoR3abe+uiS2WenbV3CZUdUTuRUio2Zzw8Xe5erl69yZ5qZLtdoanyuluaNetKkLoXvRMpE7iyir4IvTPqQ8yg1NequlZaNM26mpkRMf3mzjcnrVzvNbnxXJv4KiaMLFVrTWeMzOkR39O7qQubXabuYVUYiZiIjSmIjWZieOz0az1+Eb41j09cQ01l0lbtO09Q18vbNfMnwnomXOcqd2X4PU0hU0un9nSXeuyxjmvqpMdX8SrwInrVOFEQwzUtgq7JHRy10/bVdd2iyu4ldwuTGEVy83LhV5+rkenZG1Gq7larY6NzLRZ4ollbnk97WoiZ9aqnTuRF8Rcs01YePra06zVVPXx9Z3QWMVct46rS3s17MUUU8dNdJjXujfPkk6mkdLTRyvjWNz2I5WL1aqp0Uj3aLfqivrl03a0c9vEkc/AvOaRekSepPhfN0RTPrlO6mt1TUsbxOiie9G+Koirgg+0+XVNVEy2rNPXTNVfsLvOVX+k5XfBRcrlVVOpp5VYprqquz/Tw6tf0SnKPGV2qKMNTr9fjpxmOqO2dfh0sj8rt+lo3UdsjprhenN4KqrenFFD4xt8ceCe/wMp0ZdIaHZ7BX1zWQQUrXswxOSta9WtREVeq4RETxPBn0FV0unVlhe2oubXtf2LFxGjO9jc9V78r1x3GO3KqutPYaW1XClmo6GGZz2dtC6PjflVwqryXCquP/AMG7VatYunZoq1na3z0zunhHp0IqjE4nLK5ruUbMbE7MdETMxxnr3a1b9Z8ndkkq9WXea5XSqSht9I3Mj1XzKVi9GN8ZHd6//hD947za5pYrBQ2WBtlq5WU8jpFVKiRXOREkz3Ki4XnleXd0OvpvT9wvkMaSdvBaYXLJxImFkcvXs0Xq5enEvJE6HXv/AJHbdWyLSwtipqGeHEbFymWI1ypnqqqvf3qbGzbqr5qP6Y3RHCP19O/fOjt37dmMTVGm3O+qd81a6zPH+nSNO3u3RmO1e6xxW2KzxyfZqhzZJWp3RNXPP2uRE9ynntrvra2e0sEDkZdLo1ZUVF5s4+avX5LeFE9eDxtX2+7Pp26iu0To0r3ua+L/ACWP/FNd4ZTOfBV9Z3LFaa/V9xnulbmKlZFwROa3ha5UTDWMRfgp1Ve9fo1qLNq3h6dqfqxOs9s9X76m/exOIv465sUzFyqIpp7KZ3zV2bvKZmOMaO7schbHV3dWNRrUZA3l3r565XxX1nV1veqnUN3bZbY101LHN2bWNXlUSovNV+0bhfVyVe5DxbHfayyx3GhgiaysqWpHIjs9pC5iORV4U69V+hTp2qKqqZm0dqZJPPJHwcMLufAvXLu5F71XBs/Rf9xXfq7NPLi0P4h/sbWCt68Z2tOP2p0pj4+EdcPbq6i1We31Fpt0cFxuFRG6GurnpljUVMLHH/8Ajl45XkSFoerqK3SlvqKpsaSOjx5jcIqIqo1cexEMMr9CVtFp5KmBy1VxY9HSQRLhnZ4wrGJ3qnJcr1wfhaqHV91t0FiY2e3W2JOCSV8axKrc5wvwnLzxhMJ4mniKLOItfVrjdO+Z/flCUwV3FYHEfXtTvp0ppp4cY01nr46zMzO/ekyiraStWbySdkyQyLFIrFyjXpjLc+KZOwdKyWyjs9rgt1DHwQQtwniq9VVfFVXKndIGvZ2p2eC52prmiOc+106cAAHw9AAAAAAAAHBTL9kN2nIrqPZbaajpwVt4Vq9/WGFf0ip8gtHta1tbdnez666tuatdHRQ/YYeLCzzLyjjT1q5U9iZXuNUmqL3ctS6iuF/u9QtRX3CofUVEi97nLlceCJ0RO5ERBA80AGQAAAAAAAAAAAAAAAAAAEwbpm1Ndl+1CCaumVtgu3DSXRufNY3PmTfeKqr8lXJ3mzaN7JI2yRua9jkRzXNXKKi9FRTTSbANxHa19dmjF0Je6vjvdiiTyV0jvOqaPo32rHyavqVnrMCzAAAAAyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA4cqo1VTwIm2URtm1PC9yZWOikkTPirmJn6VJaUiOjkfozWz0qInuhTtGNRMZkgeqOarc4RVRURFT1KSuXfXtXbdPGY3e/5q1nulvEYa/X9imqdZ6tdNPRLh1LrcaK10b6uvqGQQt73LzVfBE6qvqQw27bRqdkSpbaGRXqn7ZVKjGN9yKqr9B4lBZ9Q6vrG11bLIyDuqJmcLWp4RR/1/Sp8WstqiNu/OzT7/AN/vR64jPqKquawUc5XPVwjvn9x2w+71e7vrC4NtNtgfFTOXKQquFc3PpzKnRv2vj4rgzzSun6Sw0PZRYlqJMLPOrcK9fD1NTuQ7FgstBZKLyahiVOJcySOXL5F8XL//AFD0T4xWLiunmrUaUR7+975dllVqv6Tiatq7PT0R2R+/jrw5rXIqORFTwU5AI9MutNQUU1dDXy0kL6qFqtjmViK9iL1RF7jsp0AMzMzxfMU0066RxcKqImV5ETWXOpNoLal6cUTqh1SvLpFHhI0+fg+dSQta1i0GlbjUNXD+wVjOfwnean0qYrsgokR9xrlbybwUsa+xOJ35W/MSuC/lYa5e6eEfvxhXc1/3OPw+F6PtT4cPSY8UhHDkReqIpyCJWR8Piie5rnxsc5voqqZVPYfYBnVjSBQAYZAAAwh0ZLPapK/y99tpHVSKi9ssLVfnuXOM5O8D6pqqp4S+K7dFf2o1Dx6/TFhrrgtfV2yCWoXHE5c4djpxJnC+9D2AZouV0TrTOnc+bti3ejZuUxVHbGr4ijjiibFFG2ONiYa1qYRE8EQ+wD5esRpwD4jiijz2cbGcS5XhaiZU+wNWNI4uFa1XI5Woqp0XHQ5AMMgAAKiKmFTKHzGxkbeFjGsTwamD6ANHTu1roLtS+S3GmZURcSORrs8lToqKnNF9h9W230VtpG0lBTR08LVyjWJ3+K+K+tTtA++cq2dnXd1PPmbe3zmzG1w1036d4dWgt1BQdp5DRU9N2ruKTso0bxL4rjqdoGIqmI0h9TRTMxVMb4D5ljjkbwSMa9vg5MofQPnXR9TGrr3Gqp6Chmral6RwwMV7lXuREIy0BTfVrWMtwrIUXs+Orc1eaNke7zE9yZ+ZDv7Urx5RVx2OB32OHEtVjvd1Yz3ekv3p6myei7KyVFwc3DqydeFftGean08S+8mrVE4bB1XJ41bo7v8A1rPkquIvRmGa0WKd9NvWZ74+U6R5sxc1HNVrkRUXkqKEajU4WoiInREOQQy1aPzSCFJlmSJiSuTCv4U4lTwyfnR0NFRcfkdJBT8a5f2UaN4l9eOp2AZ2p001fOxTrrpvAAfL6AAAAAAAAAAAAIY3uNrDNmGzWVtuqWs1HeEdTW1qL58SY8+f7xFTH2zm92QKwb9m1ZNY67bouz1PaWTT8jmzOY7Laisxh7vYxFVievj7lQrafT3Oe9z3uVznLlzlXKqvifJkAAAAAAAAAAAAAAAAAAAAAAyDZ3q676F1pbNVWOZY6ygmSREzhsrOj43fauaqtX1KY+ANu2zbWFo17om2arskvFSV8KP4FVFdC9OT43fbNdlF9ngZEa8NybbF9YWtPrTvtUrNOXyVrUc9fNpKpcNbJ6mu5Nd96vwVNh5gAAAABkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADrXCgoq+Hsa6kgqY+vDKxHInznZAiZidYfNVNNcbNUaw8qj05YaOZJqa0UccidHJEmU9ngeqAfVddVc61Tq+bVm3ajS3TER2RoAA+HoAAAAAML2t1PBZqOjRVzUVSK5PFrEVy/Twne2ZU/YaQppFTDqh8kzvXly4+hEMd2vrUuuFuZHTzSN7KRI1ZGrkWRzmojeXfyM8slJ9T7PR0PL7BAyNcd6oiIpKXpijA26euZn9+5XcLTN3OL1yeFERHnpPzdwAEWsQAAAAAAAAAAAAAAAAAAAAAAAyAAMAAAAAAHTvVfDa7VU3Cf9rgjV6p3uXuRPWq4T3ncMC2tXH7HR2eN37Y7yibC/BauGp73Ln702cJZ5+9TR0fBo5ni/omFru9McO+d0MDf5XWVDnL9kraybu75Hu/JlfmQnC10cdvtlNQw+hBE2NPXhMZI12Z23y7Ua1r25hoGcXqWV6Yb8zeJfehKZIZxe1rptR0fv09UFyWws0268RVxq3R3Rxnxn0AAQy1gAAAAAAAAAAAAAAFA6V+utvsVlrLzdqqOkoKKF09RNIuGsY1Mqpqy29bSbhtS2j12papZIqPPY26lc7Pk9O1V4W+1cq53rcvdgn7f52xeX167LNPVS+S0r2yXuVi8pJUwrIM+DeTnfbYTq1SoZkAAAAAAAAAAAAAAAAAAAAAAAAAAANhG5FtmTXGkk0XqCrRdRWWFGwvkfl9bSpyR/Pq9nJrvFOFeeVxr3PZ0Vqa8aO1Vb9TWCqWmuNBMksL8ZRe5WuTvaqKqKneiqBt/BhWxbaLZtqGgqPU9oe1j3p2VbS8WXUtQiJxxr8+UXvaqL3mamAAAAAGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgAAZAAGAABkAAAABgAAZAAGAAAAAGQAAAAGAAAAAGQAAAAGAAAAAACHtfTySa2uDFa98jeyihjRMucisRURqetzlJhOu+io31rK19LA6qY1WsmWNFe1vgi9UQ3cDiow1c1zGu7RFZvl1WYWabdNWzpOvumPi83Rln+oliipZOFah6rLUOTor16p7E5InsPaANW5cquVzXVxlIWLNFi3TbojdEaAAPh6gAAAAAAAAAAAAAQ9vV7X6fZVoB7qGWJ+pbojoLZCq5VnLzp1T4rMpjPVytTpnEi661TZdFaUuGp9QVaUtuoYuOR3wnL0axqd7nKqIid6qhq12ybQrztN17XapvDuFZV7Olp0XLaaBFXgjb7M5Ve9VVe8QMRqZ5qqplqamaSaeV6vkkkcrnPcq5VVVeaqq88n5gGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAStuzbXq/ZNryOse+WXT9e5sV2pW8+JmeUrU+OzKqniiqnflNm9muVBebTS3a1VcVZQ1cTZqeeJ2WSMcmUci+w05FptyLbmul7rDs51VWNbYq6VfqbUyuwlHO5fQVV6RvX3NcuejlVAvkADAAAAADIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMAAAAAAAAAAAAAAAAAfMj2Rxukkc1jGornOcuERE6qqn0U/wB+bbmtDDUbLtJ1jfKZmcN8qonc4mL/APDNVPhKnp+Ceb3uwET74221+0nVX1u6frFdpO1Sr2SsXCVs6ZRZl8WplUZ6sr8LCV/AMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAL47lG3lmprZT7OtXVq/V6kj4bbVTP510LU/a1VesrUT75qZ6oqrac0326tq7dcKe4UFTLTVdNI2WGaJytfG9q5a5FToqKhsg3U9udHtW059TLs+Km1bbok8shTDUqmJy7eNPBVxxNT0VXwVDAnAAAAAAABkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYAAAAAAAAAAjPeF2u2bZHo11zq+Cqu9Ujo7ZQcXOeRE9J2OaRtyiuX2InNUAxPe325UuzDTTrHZJ2SauuUK+TNRUXyKNeSzvTx6o1F6qmeiKi646meapqJampmkmnler5JJHK5z3KuVcqrzVVXnk9LWGo7xq3UtdqO/1j6y5V0qyzSu8e5ETuaiYRETkiIiHkmQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD19HakvOkdTUOo9P1r6O5UMqSQyt+lFTorVTKKi8lRVQ8gAbSd3fbDZNrmkUrqbs6S9UjWsudv4srE9fhtzzWN3PC93NF5oSeaitm+tb/s/1fR6n05VrT1tM7CtXmyaNfSjenwmr3p7FTCoimzjYftR0/tW0ZFfbM9IqmPEdfQudmSklx6K+LV5q13RU8FRUTAz0AAAAAABkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgAAAAAAAAADHNpGtbBs/0jWan1JWJT0VMmGtTm+aRfRjYnwnLjkntVcIiqgdDa/tG09sw0bUak1BN5rfMpaVjk7WqlxyjYn5V6ImVU1jbWdoGoNpWtKvU+oZ0dNKvBBAxV7OmiRfNjYncifOqqqrzU9DbltRv+1bWkt+vD1hpY8x2+ha7MdLFn0U8XLyVzu9fUiImBGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAy/ZLtD1Hsz1hT6k05U8ErPMqKd6r2VVFnnG9O9F8eqLhU5oYgANsWxnabpzano+LUFgm4ZG4ZW0Ujk7WklxzY5O9OuHdHJ70TNjUtsl2iak2Zavg1Hpuq4JG+ZUUz1VYqqLPOORO9PBeqLzTmbLtiu1LTO1XSjL3YZuzqI0RtdQSOTtqSRU9F3i1eeHJyVPBUVEwM7AAAAAAAZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGAAAAAAADHNo2ttObP9K1OpNT17aSigTDWpzkmf3Rxt+E5e5PeuERVQPvaBrDT+hNK1epdS1zaO30rea9XyvX0Y2N+E9e5PyIiqa0d4Ha/f9rmrFuNwV1JaaZXNttua/LKdi/CX40jsJl3uTCIiDb/tg1Dtc1StfcVdR2mmc5tutrH5ZTsX4S/GkXll3uTCIiEamQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMl2ba51Js91VT6j0xXupayLk9i8452Z5xyN+E1fD3phURTGgBtL3f9smnNrmmvK7e5tHeaZqfVC2Pfl8LvjN+PGq9He5cLyJNNQGi9UX3RupKTUWm7jLb7lSu4o5Y16p3tcnRzVTkrV5KhsU3adv8AY9rFuS21qQ2vVdPHmooeLzKhETnJCq81Txb1b605rgTWAAAAAAAyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwAAAAAAARpt82yaZ2R6d8rub0rLvUtXyC2RvRJJ16cTviRovVy+xMryA9na7tI0zsw0nLqDUlVwt5spaWNUWaqkxyYxPyr0ROamtbbZtV1PtW1S68X6bsqWJVbQ2+Jy9jSxr3J4uXlly819SIiJ5+1XaFqfaVqqXUOp61Z5l8yCBmUhpo88o429yfSq81VVMSMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHatVwrrTcqe52ysno62mkSWCeB6sfG9OaOaqc0U6oAv1uu7ztDrFlPpPaBU09v1Hyjpq5cMgr16Ii9zJfV0d3YXDSzxpnLYbsm9RV2HyPSO0qolrLQmIqa7uy+alTojZe+RifG9JPtk6YF6Ade211Hc6CC4W6rgrKSoYkkM8EiPZI1eitcnJUOwAAAAAGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgAAAB8yPZHG6SR7WMaiuc5y4RETvVSoW81vWRUK1OktltVHPUYWOqvjPOZEvRW0/c53+k6J8HPVAkfeY3ibHswoprJY30921dI3DabPFFRZTk+ZU7+9GIuV78JhV166u1JfNW6gqr/qO5T3G5VTuKWeV2V9SInRrUTkiJhEToebUzzVNRLU1M0k08r1fJJI5XOe5VyrlVeaqq88n5mQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAS/u+be9V7Jq5lHG51101LJxVNslfybnq+F3+Ld3+C96dFTYZst2i6T2ladbfNKXJtTCio2eB6cE9M/4sjOrV8F5ovcqoalD39B6x1JobUUF/wBL3We3V0K+kxfNkb3se1eT2r4LyA28gr9u77zWmdoraaxakWCw6pd5qRudimq1/wBE9V5OX4jufgru6wJgAAAABkAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAYAAAAAAPE1vqzT+itOVGoNT3SC3W6nTzpJF5ud3Na1ObnL3NRFVSOdv28Bo/ZTSyULpG3fUro8w2yB6eYvc6Z3+Lb345uXuTHNNfW1faXq7abqF141TcXTK3KU9LHltPTN+LGzPL1quVXvVRoJO3kd5TUG0l9Tp/Tvb2XSSrwrFnFRWp4yqi8mr/ACaLjxV3LFfwDIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAWZ3fN62/wCj0p7Brzyi/wBibwxxVeeKrpG9Oq/trU8FXiTuXkiFZgBt+0ZqrT2srDDfdMXamudvm9GWF2eFe9rkXm1yd7VRFQ9o1IbNdoOrdnd9S8aTu81BMuEmj9KGdqL6MjF5OT6U7lReZerYPvTaP14kFn1QsOmdQvVGo2WTFJUu6J2ci+iq/Ed7EVxgWFAAAAGQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgAAAAAAAhjbrvFaI2YRz25kzb5qNqKjbbSyJiJ3+mfzSP2c3erHMCWr7drZYrVUXa819Nb6CmYr5qiokRjGJ61UphvBb3VXX+Uae2WLLR0uVZLe5GYmkTv7Fip5ifbO87wRvUgHbDtc1rtSu3lepbkvkcb1dS26DLKan+S3vX7Z2V9eORgJkfpUzz1VTJU1M0k88r1fJJI5XOe5Vyqqq81VV7z8wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAnTYZvMa42c9ha7lI7UenmYalHVSL2sDf8ARS81T5Lst8ETqXm2R7XtC7ULck+mbszy1rOKe21OI6qDxyzPnJ9s1Vb6zVKdi211bbK+Gvt1ZUUVXA7jhnp5Fjkjd4tcioqL7ANyAKLbFt8S+WhILTtJo3XqiTDEudM1raqNOmXt5NkRPHzXePEpcbQGudJ68s6XbSV8pLpTcu0SJ2JIlX4MjFw5i+pyIYGRgAAABqAAMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMAAAAAAAHlaq1JYdK2eW8aju9HaqCL0p6mVGNz3Imeqr3ImVXuA9UxTaTtE0fs7s63PVt7p7exUXsYVXimnVO6ONPOd82E71Qq5to3ylVJrTstoOHq1bxXxc/bFCv5X/glR9SX69alvE13v90q7nXzrmSoqZVe9fVleiJ3InJO4aCwG3Pex1ZrFs9n0WybTFkflrpmvTy2oavxnpyjT1MXP2ypyK3vc573Pe5XOcuXOVcqq+J8gyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHqaX1DfdL3iK8adu1Za6+L0J6WVWOx3ouOqL3ovJe88sAXE2Pb5tVD2Ns2nWrylnJv1Wt0aNk9skPJq+1mMJ8FS2uiNaaV1vakumlL9RXal5cToJPOjVe57Fw5i+pyIpqGPS03f73pq7RXbT92rbXXRehPSzOjeieGU6ovei8l7wNwwKM7Jd8y+21IbftGtLbzTJhq3Cha2KpRPF0fJj19nB7y2WzXapoLaJSpLpTUdJWTcPE+kcvZ1MfjxROw7HrRFTwUwM0AAAAAAANQAAAAAAAZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGAAAAAAADGtea80foS3eX6t1DQ2mJUVWNmfmST5EaZe9fkooGSnmam1BY9M2mW7ahu1Fa6GL056qZI258EVeqr3InNe4qJtZ30XO7W37NbHwJzalzujeftZCi/Mrl9rSqettZap1rdlumqr7XXaq58LqiTLY0XuY1PNYnqaiIZFv9sO+XbKLt7Zsztf1Rn5t+qtexzIU9ccXJzvFFdw/JUqHrzW2q9dXh121ZfKu61XPgWZ/mRIvwWMTDWJ6moiGPAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAH60lTUUdTHVUk8tPPE5HRyxPVr2KneipzRT8gBP+y/ew2m6RSOkvc8WrLczCcFwcqVCJ9rOnnKvrejy02zPel2Wax7Klrrk/TNxfhFgumGRKv2syeZj5StX1GtoAblKaeCpgZUU00c0MjUcySNyOa5F70VOqH6GpbZ/tN15oGdH6T1PcLdHxcTqZsnHTvX7aJ2WL7cZLI7Od9evg7Kl1/paOrYnJ1baXcEmPFYnrwuX2PanqMaC7AI22fbdNlmuEjjsuraKKsf0o65fJp8+CNfhHr8lXEkgAAAAAAAAAAAAAAAGQAAAAAAAAAAAAAAAAAAAAGAAAAAAAAAAAAAAAYNtA2ubONCNkbqbVlupKlmc0kcnbVGfDsmZcntVET1lc9ou+xTR9pS6A0q+Z3NG1t3fwtz4pDGuVT2vT2AXEcqNarnKjWomVVeiEQ7Td47ZVoXtaeovzbzcWZTyK1Ik7kXwc9F7NvsV2fUUF2j7ZNpO0BZI9SaprZaN6/uGnd2FNjwWNmEd7XZX1mAGRZTahvg6+1EktHpKlptK0LsokrFSercny3Jwt+9blPjFd7xdLnebjLcbvcKu4Vsy5kqKqZ0sj19bnKqqdMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAM90Ftj2maH7OPTusLlBSx9KSZ/b06J4JHJxNT3IimBAC3mhd9q8U6Rwa10jS1zejqq2SrC/wBqxv4kcvsc1Cd9Ebzex3VPBGmpkstS/H2C7x+T49snOP8AHNZgA3H26voblRsrLdWU1bTP9CanlbIx3sc1VRTsmnvTuor/AKcq/K9P3u5WmozntKKqfC5farVTJMGj963bFp/gjqbzR36Bv+LudI1y4+WzhevvVTGg2SAp7pPfet70ZHqzQ1TCqenNbKpsmfZHIjcfhqS1pbeg2MX7hYuqHWqZ3+KuVK+HHteiKz8YCaAeRp7U+mtRRJLYNQWq7MVM8VFWRzJj71VPXAAAAAAAAAAAAAoAAAAAAAAAAAAAAAAAA8+9XuzWSn8ovV3t9shxntKupZC353KiEZap3k9jOn0c2XWdNcJm9I7dE+p4vY5icHzuAl0FSNWb7mnoEfHpbRdzr3dGy3CoZTtT18LONVT3p7iH9Yb3W1298cdtqrXp6F3JEoaRHPx63yq/n60RBoNidTPBTQPnqZo4YY0y+SRyNa1PFVXkhFmt94jZBpPjjrNYUdwqWcvJ7Xmrcq+HEzLEX5TkNbeqtY6r1XOs2pdSXa7vzlPK6t8jW+xFXCe5DwhoLl66321XtIND6Nx14Kq7zflhjX/7hAOvdvW1fWiSRXXV9bT0j+S0lvVKWLHxVSPCuT5SqRkDI5VVVVVVVVXmqqcAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB9wySQytlikdHI1ctc1cKi+pTNtObX9qOnka20691BDG30Yn1r5Y0+8eqt+gwYAT9YN7rbJbOFKyvs95RP8tt7W598KsM/sW/BeI+Ft80BQ1PxnUdwfDj2Ncx/wCUqEAL7WbfX2e1GG3XTOpKFy9VhbDOxPfxtX6DMrTvW7E67CTalqre5fg1Ntn/ACsa5PpNa4Gg2p2zbnsfuGPJ9ounmZ/yiqSD9JwmSW/XOibhj6n6w09V56dhc4X5+ZxqIBjQbk6app6lnHTTxTN8Y3o5PoP1NO+n/wDCsXtLI7M/Qi9wF9wQdoz0GexCUbL6DQMgB8J6HuPIu/oOA9k5Ia1h6DyAdpPoS+8C7800ULFfNKyNidXPciJ9J41frLSFvz5fqqxUmOvb3CJmPncan9Y/4WceIZG1e57bNklvz5RtF007HXsK9ky/iKpit23ptiVBlGaslrXp8Gmt1Q76VYjfpNaQAv5ed9PZrTZbbbFqa4PToroYomL71kVfxTCL5vw1juJlk2ewRfFkrLkr8+1rWN/OKdACw9/3wtr1x4koHWGzIvRaWg43J/Sueir7iOtRbbdrV/RzbltAv3A/0o6apWmY71K2LhRU9WCPQB+1XVVNZUOqKuolqJnelJK9XOX2qvM/EAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP/Z" alt="Waterqo Logo" style="width:48px;height:48px;border-radius:50%;object-fit:cover;">
</div>
    <div class="header-text">
      <h1>Waterqo Swimming Pools</h1>
      <p>Pool Service Manager</p>
    </div>
  </div>
  <div style="display:flex;gap:10px;align-items:center">
    <button class="btn-download" style="background:rgba(255,255,255,.18);color:#fff;border:1.5px solid rgba(255,255,255,.35)" onclick="openManagePersons()">👤 Service Persons</button>
    <a class="btn-download" href="/api/download">⬇ Download Excel</a>
  </div>
</header>

<div class="layout">
  <!-- SIDEBAR -->
  <div class="sidebar">
    <div class="sidebar-header">
      <h2>Clients</h2>
      <input class="search-box" type="text" id="search" placeholder="🔍  Search clients..." oninput="filterClients()">
    </div>
    <div class="client-list" id="clientList"></div>
    <button class="add-client-btn" onclick="openAddClient()">
      <svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="3"><line x1="12" y1="5" x2="12" y2="19"/><line x1="5" y1="12" x2="19" y2="12"/></svg>
      Add New Client
    </button>
  </div>

  <!-- MAIN -->
  <div class="main">
    <div id="emptyState" class="empty-state" style="flex:1;display:flex">
      <svg width="80" height="80" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.2">
        <path d="M3 9l9-7 9 7v11a2 2 0 01-2 2H5a2 2 0 01-2-2z"/>
        <polyline points="9 22 9 12 15 12 15 22"/>
      </svg>
      <h3>Select a client</h3>
      <p>Choose a client from the sidebar to view and manage their service entries.</p>
    </div>

    <div id="mainContent" style="display:none;flex-direction:column;flex:1;overflow:hidden">
      <div class="main-header">
        <div>
          <h2 id="clientTitle">—</h2>
          <div class="info-chips">
            <span class="chip" id="locChip"></span>
            <span class="chip" id="personChip"></span>
          </div>
        </div>
        <div class="main-header-btns">
          <button class="btn-delete-client" onclick="openDeleteClient()">🗑 Delete Client</button>
          <button class="edit-btn" style="width:auto;padding:9px 18px;margin:0" onclick="openBulkEdit()">✏ Edit Entries</button>
        </div>
      </div>
      <!-- Month selector -->
      <div class="month-bar" id="monthBar">
        <span class="lbl">Month:</span>
        <button class="month-btn" onclick="selectMonth('January')">Jan</button>
        <button class="month-btn" onclick="selectMonth('February')">Feb</button>
        <button class="month-btn" onclick="selectMonth('March')">Mar</button>
        <button class="month-btn" onclick="selectMonth('April')">Apr</button>
        <button class="month-btn" onclick="selectMonth('May')">May</button>
        <button class="month-btn" onclick="selectMonth('June')">Jun</button>
        <button class="month-btn" onclick="selectMonth('July')">Jul</button>
        <button class="month-btn" onclick="selectMonth('August')">Aug</button>
        <button class="month-btn" onclick="selectMonth('September')">Sep</button>
        <button class="month-btn" onclick="selectMonth('October')">Oct</button>
        <button class="month-btn" onclick="selectMonth('November')">Nov</button>
        <button class="month-btn" onclick="selectMonth('December')">Dec</button>
      </div>
      <div class="entries-area">
        <div class="entries-grid" id="entriesGrid"></div>
      </div>
  
    </div>
  </div>
</div>

<!-- ADD CLIENT MODAL -->
<div class="overlay hidden" id="addClientOverlay">
  <div class="modal">
    <h3>Add New Client</h3>
    <p class="modal-sub">A new service sheet will be created automatically.</p>
    <div class="form-row">
      <div class="field"><label>Client Name</label><input id="nc_name" placeholder="Mrs. Example" /></div>
      <div class="field"><label>Location</label><input id="nc_location" placeholder="Colombo" /></div>
    </div>
    <div class="form-row full">
      <div class="field"><label>Service Person</label>
        <select id="nc_person">
          <option value="">— Select —</option>
        </select>
      </div>
    </div>
    <div class="modal-actions">
      <button class="btn-cancel" onclick="closeModal('addClientOverlay')">Cancel</button>
      <button class="btn-primary" onclick="submitAddClient()">Add Client</button>
    </div>
  </div>
</div>

<!-- DELETE CLIENT CONFIRM MODAL -->
<div class="overlay hidden" id="deleteClientOverlay">
  <div class="modal">
    <h3>Delete Client</h3>
    <p class="modal-sub" id="deleteClientSub"></p>
    <p style="font-size:.9rem;color:#d62828;background:#fff0f0;border:1px solid #f4a7a7;border-radius:8px;padding:12px 14px;margin-bottom:8px;">
      ⚠️ This will permanently remove the client and all their service entries from the Excel file. This cannot be undone.
    </p>
    <div class="modal-actions">
      <button class="btn-cancel" onclick="closeModal('deleteClientOverlay')">Cancel</button>
      <button class="btn-danger" onclick="confirmDeleteClient()">Yes, Delete</button>
    </div>
  </div>
</div>

<!-- EDIT ENTRY MODAL -->
<div class="overlay hidden" id="editEntryOverlay">
  <div class="modal">
    <h3 id="editTitle">Edit Entry</h3>
    <p class="modal-sub" id="editSub"></p>
    <input type="hidden" id="edit_no" />
    <div class="form-row full">
      <div class="field"><label>Service Person</label>
        <select id="edit_person">
        </select>
      </div>
    </div>
    <div class="form-row">
      <div class="field"><label>Price (Rs.)</label><input id="edit_price" type="number" step="0.01" placeholder="0.00" /></div>
      <div class="field"><label>Chlorine Price (Rs.)</label><input id="edit_chlorine" type="number" step="0.01" placeholder="0.00" /></div>
    </div>
    <div class="form-row full">
      <div class="field"><label>Payment Status</label>
        <select id="edit_paid">
          <option value="">— Unpaid —</option>
          <option value="Paid">Paid</option>
          <option value="Partial">Partial</option>
        </select>
      </div>
    </div>
    <div class="modal-actions">
      <button class="btn-cancel" onclick="closeModal('editEntryOverlay')">Cancel</button>
      <button class="btn-primary" onclick="submitEntry()">Save Entry</button>
    </div>
  </div>
</div>

<!-- MANAGE SERVICE PERSONS MODAL -->
<div class="overlay hidden" id="managePersonsOverlay">
  <div class="modal" style="width:420px">
    <h3>👤 Service Persons</h3>
    <p class="modal-sub">Add or remove pool service team members.</p>
    <div id="personsList" style="margin-bottom:16px;display:flex;flex-direction:column;gap:8px;max-height:240px;overflow-y:auto"></div>
    <div style="display:flex;gap:8px;margin-bottom:4px">
      <input id="newPersonName" class="search-box" placeholder="New person's name..." style="flex:1;padding:10px 13px" onkeydown="if(event.key==='Enter')submitAddPerson()" />
      <button class="btn-primary" style="flex:0 0 auto;padding:10px 18px" onclick="submitAddPerson()">+ Add</button>
    </div>
    <div class="modal-actions" style="margin-top:16px">
      <button class="btn-cancel" onclick="closeModal('managePersonsOverlay')">Close</button>
    </div>
  </div>
</div>

<script>
let clients = [], currentClient = null, currentEntries = [], allPersons = [];
const MONTHS = ['January','February','March','April','May','June','July','August','September','October','November','December'];
let currentMonth = MONTHS[new Date().getMonth()];

function populatePersonSelects(){
  ['nc_person','edit_person'].forEach(id => {
    const sel = document.getElementById(id);
    const cur = sel.value;
    sel.innerHTML = id === 'nc_person' ? '<option value="">— Select —</option>' : '';
    allPersons.forEach(p => {
      const o = document.createElement('option');
      o.value = o.textContent = p;
      sel.appendChild(o);
    });
    if(cur) sel.value = cur;
  });
}

async function loadPersons(){
  const r = await fetch('/api/persons');
  allPersons = await r.json();
  populatePersonSelects();
}

async function openManagePersons(){
  await loadPersons();
  renderPersonsList();
  document.getElementById('managePersonsOverlay').classList.remove('hidden');
}

function renderPersonsList(){
  const el = document.getElementById('personsList');
  el.innerHTML = '';
  if(!allPersons.length){
    el.innerHTML = '<div style="color:#8aaabb;font-size:.85rem;text-align:center;padding:10px">No service persons added yet.</div>';
    return;
  }
  allPersons.forEach(p => {
    const row = document.createElement('div');
    row.style.cssText = 'display:flex;align-items:center;justify-content:space-between;background:#f0faff;border:1.5px solid var(--border);border-radius:8px;padding:9px 14px';
    row.innerHTML = `<span style="font-weight:600;color:var(--navy)">👤 ${p}</span>
      <button onclick="deletePerson('${p}')" style="background:transparent;border:1.5px solid #d62828;color:#d62828;border-radius:6px;padding:4px 10px;cursor:pointer;font-size:.8rem;font-weight:700">Remove</button>`;
    el.appendChild(row);
  });
}

async function submitAddPerson(){
  const name = document.getElementById('newPersonName').value.trim();
  if(!name){ showToast('Enter a name', 'error'); return; }
  const r = await fetch('/api/persons', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({name})});
  const d = await r.json();
  if(!d.ok){ showToast(d.error || 'Error', 'error'); return; }
  document.getElementById('newPersonName').value = '';
  await loadPersons();
  renderPersonsList();
  showToast(`"${name}" added!`, 'success');
}

async function deletePerson(name){
  await fetch('/api/persons/delete', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({name})});
  await loadPersons();
  renderPersonsList();
  showToast(`"${name}" removed`, 'success');
}

async function loadClients(){
  const r = await fetch('/api/clients');
  clients = await r.json();
  renderClients();
}

function renderClients(filter=''){
  const f = filter.toLowerCase();
  const list = document.getElementById('clientList');
  list.innerHTML = '';
  const filtered = clients.filter(c =>
    c.name.toLowerCase().includes(f) || c.location.toLowerCase().includes(f) || c.person.toLowerCase().includes(f)
  );
  if(!filtered.length){
    list.innerHTML = '<div style="text-align:center;color:#7aaabb;padding:30px;font-size:.83rem">No clients found</div>';
    return;
  }
  filtered.forEach(c => {
    const d = document.createElement('div');
    d.className = 'client-card' + (currentClient && currentClient.name === c.name && currentClient.location === c.location ? ' active' : '');
    d.innerHTML = `<h3>${c.name}</h3><div class="meta">📍 ${c.location} &nbsp;·&nbsp; 👤 ${c.person}</div><span class="badge">→</span>`;
    d.onclick = () => selectClient(c);
    list.appendChild(d);
  });
}

function filterClients(){
  renderClients(document.getElementById('search').value);
}

function selectMonth(month){
  currentMonth = month;
  // Update active button
  document.querySelectorAll('.month-btn').forEach(btn => {
    const short = btn.textContent.trim();
    const full = MONTHS.find(m => m.startsWith(short.substring(0,3)));
    btn.classList.toggle('active', full === month);
  });
  if(currentClient) loadEntriesForMonth();
}

async function loadEntriesForMonth(){
  const c = currentClient;
  const r = await fetch(`/api/entries?name=${encodeURIComponent(c.name)}&location=${encodeURIComponent(c.location)}&month=${encodeURIComponent(currentMonth)}`);
  currentEntries = await r.json();
  renderEntries();
}

function highlightCurrentMonthBtn(){
  document.querySelectorAll('.month-btn').forEach(btn => {
    const short = btn.textContent.trim();
    const full = MONTHS.find(m => m.startsWith(short.substring(0,3)));
    btn.classList.toggle('active', full === currentMonth);
  });
}

async function selectClient(c){
  currentClient = c;
  renderClients(document.getElementById('search').value);
  document.getElementById('emptyState').style.display = 'none';
  document.getElementById('mainContent').style.display = 'flex';
  document.getElementById('clientTitle').textContent = c.name;
  document.getElementById('locChip').textContent = '📍 ' + c.location;
  document.getElementById('personChip').textContent = '👤 ' + c.person;

  highlightCurrentMonthBtn();
  const r = await fetch(`/api/entries?name=${encodeURIComponent(c.name)}&location=${encodeURIComponent(c.location)}&month=${encodeURIComponent(currentMonth)}`);
  currentEntries = await r.json();
  renderEntries();
}

function fmtPrice(v){ return v ? 'Rs. ' + parseFloat(v).toLocaleString('en-LK', {minimumFractionDigits:2}) : '—'; }

function renderEntries(){
  const grid = document.getElementById('entriesGrid');
  grid.innerHTML = '';
  let totPrice = 0, totChlor = 0, paidCount = 0;
  currentEntries.forEach(e => {
    const isEmpty = !e.price && !e.chlorine && !e.paid;
    const paid = e.paid || '';
    const statusClass = paid === 'Paid' ? 'status-paid' : paid === 'Partial' ? 'status-partial' : isEmpty ? 'status-empty' : 'status-unpaid';
    const statusText = paid || (isEmpty ? 'Empty' : 'Unpaid');
    if(e.price) totPrice += parseFloat(e.price);
    if(e.chlorine) totChlor += parseFloat(e.chlorine);
    if(paid === 'Paid') paidCount++;
    const card = document.createElement('div');
    card.className = 'entry-card';
    card.innerHTML = `
      <div class="entry-header">
        <span class="entry-num">Entry #${e.no}</span>
        <span class="entry-status ${statusClass}">${statusText}</span>
      </div>
      <div class="entry-body">
        <div class="entry-field"><label>Service Person</label><span class="val">${e.person || '—'}</span></div>
        <div class="entry-field"><label>Price</label><span class="val">${fmtPrice(e.price)}</span></div>
        <div class="entry-field"><label>Chlorine</label><span class="val">${fmtPrice(e.chlorine)}</span></div>
        <div class="entry-field"><label>Payment</label><span class="val">${paid || 'Unpaid'}</span></div>
        <button class="edit-btn" onclick="openEditEntry(${e.no})">✏ Edit Entry #${e.no}</button>
      </div>`;
    grid.appendChild(card);
  });
  document.getElementById('sumTotal').textContent = 'Rs. ' + totPrice.toLocaleString('en-LK', {minimumFractionDigits:2});
  document.getElementById('sumChlorine').textContent = 'Rs. ' + totChlor.toLocaleString('en-LK', {minimumFractionDigits:2});
  document.getElementById('sumPaid').textContent = `${paidCount} / 8`;
  const balance = totPrice + totChlor;
  document.getElementById('sumBalance').textContent = 'Rs. ' + balance.toLocaleString('en-LK', {minimumFractionDigits:2});
}

function openEditEntry(no){
  const e = currentEntries.find(x => x.no == no);
  document.getElementById('editTitle').textContent = `Edit Entry #${no}`;
  document.getElementById('editSub').textContent = `${currentClient.name} · ${currentClient.location}`;
  document.getElementById('edit_no').value = no;
  document.getElementById('edit_person').value = e.person || currentClient.person;
  document.getElementById('edit_price').value = e.price || '';
  document.getElementById('edit_chlorine').value = e.chlorine || '';
  document.getElementById('edit_paid').value = e.paid || '';
  document.getElementById('editEntryOverlay').classList.remove('hidden');
}

async function submitEntry(){
  const body = {
    name: currentClient.name,
    location: currentClient.location,
    no: document.getElementById('edit_no').value,
    person: document.getElementById('edit_person').value,
    price: document.getElementById('edit_price').value,
    chlorine: document.getElementById('edit_chlorine').value,
    paid: document.getElementById('edit_paid').value,
    month: currentMonth
  };
  await fetch('/api/entries', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(body)});
  closeModal('editEntryOverlay');
  await loadEntriesForMonth();
  showToast('Entry saved successfully!', 'success');
}

function openBulkEdit(){ openEditEntry(1); }

function openAddClient(){
  document.getElementById('addClientOverlay').classList.remove('hidden');
}

async function submitAddClient(){
  const name = document.getElementById('nc_name').value.trim();
  const location = document.getElementById('nc_location').value.trim();
  const person = document.getElementById('nc_person').value;
  if(!name || !location || !person){ showToast('Please fill all fields', 'error'); return; }
  await fetch('/api/clients', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({name, location, person})});
  closeModal('addClientOverlay');
  document.getElementById('nc_name').value = '';
  document.getElementById('nc_location').value = '';
  document.getElementById('nc_person').value = '';
  await loadClients();
  showToast(`Client "${name}" added!`, 'success');
}

function openDeleteClient(){
  document.getElementById('deleteClientSub').textContent =
    `Are you sure you want to delete "${currentClient.name}" (${currentClient.location})?`;
  document.getElementById('deleteClientOverlay').classList.remove('hidden');
}

async function confirmDeleteClient(){
  const c = currentClient;
  await fetch('/api/clients/delete', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({name: c.name, location: c.location})
  });
  closeModal('deleteClientOverlay');
  currentClient = null;
  document.getElementById('mainContent').style.display = 'none';
  document.getElementById('emptyState').style.display = 'flex';
  await loadClients();
  showToast(`Client "${c.name}" deleted.`, 'success');
}

function closeModal(id){ document.getElementById(id).classList.add('hidden'); }

function showToast(msg, type=''){
  const t = document.createElement('div');
  t.className = 'toast ' + type;
  t.textContent = msg;
  document.body.appendChild(t);
  setTimeout(() => t.remove(), 3000);
}

document.querySelectorAll('.overlay').forEach(o => o.addEventListener('click', e => { if(e.target === o) o.classList.add('hidden'); }));

loadPersons();
loadClients();
highlightCurrentMonthBtn();
</script>
</body>
</html>"""

if __name__ == '__main__':
    ensure_workbook()
    print("\n" + "="*55)
    print("  🏊 Waterqo Swimming Pools — Pool Service Manager")
    print("="*55)
    print(f"  📂 Excel file: {os.path.abspath(EXCEL_FILE)}")
    print(f"  🌐 Open in browser: http://localhost:5050")
    print("  Press Ctrl+C to stop")
    print("="*55 + "\n")
    app.run(host='0.0.0.0', port=5050, debug=False)
